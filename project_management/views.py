import json
from django.shortcuts import redirect, get_object_or_404, HttpResponse
from django.shortcuts import render
from django.db.models import Q
import math
from timezonefinder import TimezoneFinder
from .models import *
from .models import Link
from project_management.serializers import TaskSerializer
from project_management.serializers import LinkSerializer
import datetime
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.http import JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.contrib.auth.models import User, Group
from django.db import connection, IntegrityError
from django.contrib.auth.decorators import login_required
from users.models import Profile
from datetime import date, timedelta
import os
from project_management_system.settings import MEDIA_ROOT
from .action import createTaskLogHistory, createProjectLogHistory
from django.core.files.storage import default_storage
from django.contrib import messages
import csv
from django.contrib.sites.shortcuts import get_current_site
from collections import Counter
from dateutil.parser import parse
from django.utils.timezone import localtime, now
from django.db.models import Max
from .models import FIELD_TYPE, ProjectField, FieldChoice, SectionTemplateMetadata, SectionMetadataValues, Task, Project
from django.db.models import F, IntegerField
from django.db.models.functions import Cast
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from collections import OrderedDict
from pytz import timezone, UnknownTimeZoneError
from django.utils.timezone import now as django_now
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from io import BytesIO
import pytz
import tzlocal
from django.views.decorators.csrf import csrf_exempt
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


class GanttChartView(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):
        project_obj = get_object_or_404(Project, proj_id=kwargs["proj_id"])
        project_id = project_obj.proj_id
        project_name = project_obj.proj_name

        request.session['project_id'] = project_id
        request.session['user_id'] = request.user.id

        context = {
            'client_name': connection.tenant.name,
            'project_id': project_id,
            'project_name':project_name,
        }
        return render(request, 'project_management/gantt_chart_view.html', context)


@login_required
@api_view(['GET'])
def data_list(request, offset):
    if request.method == 'GET':
        project_id = request.session['project_id']
        user_id = request.session['user_id']
        project_obj = get_object_or_404(Project, proj_id=project_id)
        #tasks = Task.objects.all()
        links = Link.objects.all()

        user = User.objects.get(id=user_id)
        common_group = get_object_or_404(Group, name='COMMON')
        admin_group = get_object_or_404(Group, name='DOC_ADMIN')

        if admin_group in user.groups.all():
            tasks = Task.objects.filter(project=project_obj, task_deleted=False).distinct()
        else:
            tasks = Task.objects.filter(Q(project=project_obj, task_deleted=False, project__proj_doc_group__in=user.groups.all()) |
                                       Q(project=project_obj, task_deleted=False, project__proj_doc_group__in=[common_group]) |
                                       Q(project=project_obj, task_deleted=False, project__proj_owner=user) |
                                       Q(project=project_obj, task_deleted=False, created_by=user)).distinct()

        taskData = TaskSerializer(tasks, many=True)
        linkData = LinkSerializer(links, many=True)

        return Response({
            'tasks': taskData.data,
            'links': linkData.data
        })


@login_required
@api_view(['POST'])
def task_add(request):
    if request.method == 'POST':
        serializer = TaskSerializer(data=request.data)

        if serializer.is_valid():
            task = serializer.save()
            return JsonResponse({'action':'inserted', 'tid': task.id})
        return JsonResponse({'action':'error'})


@login_required
@api_view(['PUT', 'DELETE'])
def task_update(request, pk):
    try:
        task = Task.objects.get(pk=pk)
    except Task.DoesNotExist:
        return JsonResponse({'action':'error2'})

    if request.method == 'PUT':
        serializer = TaskSerializer(task, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse({'action':'updated'})
        return JsonResponse({'action':'error'})

    if request.method == 'DELETE':
        # task.task_deleted = True
        # task.save()
        task.delete()
        return JsonResponse({'action':'deleted'})


@login_required
@api_view(['POST'])
def link_add(request):
    if request.method == 'POST':
        serializer = LinkSerializer(data=request.data)
        if serializer.is_valid():
            link = serializer.save()
            return JsonResponse({'action':'inserted', 'tid': link.id})
        return JsonResponse({'action':'error'})


@login_required
@api_view(['PUT', 'DELETE'])
def link_update(request, pk):
    try:
        link = Link.objects.get(pk=pk)
    except Link.DoesNotExist:
        return JsonResponse({'action':'error'})

    if request.method == 'PUT':
        serializer = LinkSerializer(link, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse({'action':'updated'})
        return JsonResponse({'action':'error'})

    if request.method == 'DELETE':
        link.delete()
        return JsonResponse({'action':'deleted'})


class TreeGridView(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):
        project_obj =  get_object_or_404(Project, proj_id=kwargs["proj_id"])
        project_id = project_obj.proj_id
        project_name = project_obj.proj_name
        project_description = project_obj.proj_description

        user = self.request.user
        username = str(user.username)
        staff_user = user.is_staff
        super_user = user.is_superuser

        delete_permissions = 'False'
        if username == 'vspladmin' or staff_user == True or super_user == True:
            delete_permissions = 'True'
        else:
            delete_permissions = 'False'
        user_list = [user.username for user in User.objects.all()]
        show_hide_column_name_list = []

        common_group = get_object_or_404(Group, name='COMMON')
        admin_group = get_object_or_404(Group, name='DOC_ADMIN')

        proj_temp_obj = project_obj.proj_origin
        try:
            proj_temp_id = project_obj.proj_origin.proj_temp_id
        except AttributeError:
            messages.error(request, 'Oops! Project template field is empty. Please assign template to the project.')
            return redirect('project-card-view')

        proj_temp_field_obj = ProjectTemplateField.objects.filter(proj_temp=proj_temp_obj, proj_temp_field_hide=False).extra(
            select={'int_from_position': 'CAST(proj_temp_field_position AS INTEGER)'}
        ).order_by('int_from_position', 'proj_temp_field_position')

        proj_temp_show_hide_obj = ProjectTemplateField.objects.filter(proj_temp=proj_temp_obj).extra(
            select={'int_from_position': 'CAST(proj_temp_field_position AS INTEGER)'}
        ).order_by('int_from_position', 'proj_temp_field_position')

        user_groups = request.user.groups.all()
        for item in proj_temp_show_hide_obj:
            # check if task name field is true then break loop and go to the next field
            check_task_name = " ".join(item.proj_temp_field.field_title.split("_")).lower()
            if check_task_name == "task name":
                continue
            field_id = item.proj_temp_field.field_title
            field_title = " ".join(item.proj_temp_field.field_title.split("_")).capitalize()
            proj_temp_field_hidden = item.proj_temp_field_hide
            if proj_temp_field_hidden == True:
                field_hidden = "True"
            else:
                field_hidden = "False"
            show_hide_column_name_list.append({'id':field_id, 'text':field_title, 'field_hidden':field_hidden, 'proj_temp_id':proj_temp_id})

        columns = []
        for item in proj_temp_field_obj:
            proj_temp_field_hide = item.proj_temp_field_hide

            # check if task name field is true then break loop and go to the next field
            check_task_name = " ".join(item.proj_temp_field.field_title.split("_")).lower()
            if check_task_name == "task name":
                continue
            type = item.proj_temp_field.field_type
            field_id = item.proj_temp_field.field_title
            field_title = " ".join(item.proj_temp_field.field_title.split("_")).capitalize()

            if proj_temp_field_hide == True:
                hidden = 1
            else:
                hidden = 0

            if project_obj.proj_lock:
                proj_lock = "true"
                column_editable = "false"
            else:
                proj_lock = "false"
                # check if field is disabled or not
                column_editable = "true"
                try:
                    if len(set(item.proj_temp_field_editable.all())) >= 1:
                        if len(set(item.proj_temp_field_editable.all()).intersection(set(user_groups))) >= 1:
                            column_editable = "false"
                        else:
                            column_editable = "true"
                except:
                    pass

            if type == 'string':
                string_column = {
                    'id': field_id,
                    'type': "string",
                    'editorType':"input",
                    'minWidth': 250,
                    'header': [{'text':field_title, 'align': 'left' }],
                    'hidden': hidden,
                    'editable':column_editable,
                }
                columns.append(string_column)
            elif type == 'number':
                number_column = {
                    'id': field_id,
                    'type': "number",
                    'minWidth': 100,
                    'editorType': "input",
                    'header': [{'text':field_title, 'align': 'left' }],
                    'format':"#.00",
                    'hidden':hidden,
                    'editable': column_editable,
                }
                columns.append(number_column)
            elif type == 'textarea':
                textarea_column = {
                    'id': field_id,
                    'type': "string",
                    'editorType':"textarea",
                    'minWidth': 100,
                    'header': [{'text':field_title, 'align': 'left' }],
                    'hidden': hidden,
                    'editable': column_editable,
                }
                columns.append(textarea_column)
            elif type == 'bool':
                bool_column = {
                    'id': field_id,
                    'type': 'boolean',
                    'header': [{'text':field_title, 'align': 'left' }],
                    'hidden': hidden,
                    'editable': column_editable,
                }
                columns.append(bool_column)
            elif type == 'date':
                date_column = {
                    'id': field_id,
                    'minWidth': 120,
                    'header': [{'text':field_title, 'align': 'left' }],
                    'type': 'date',
                    'format': '%d %M, %Y',
                    'align': 'center',
                    'hidden': hidden,
                    'editable': column_editable,
                }
                columns.append(date_column)
            elif type == 'single-select':
                if field_id == 'status':
                    # get or create status choices
                    status_list = ['Not Started', 'In Progress', 'Completed', 'On Hold', 'Delayed']
                    for s in status_list:
                        get_status_choice, create_status_choice = TaskStatus.objects.get_or_create(task_status=s)
                    options = [item.task_status for item in TaskStatus.objects.all()]
                elif field_id == 'priority':
                    # get or create priority choices
                    priority_list = ['Low', 'Medium', 'High']
                    for p in priority_list:
                        get_priority_choice, create_priority_choice = TaskPriority.objects.get_or_create(task_priority=p)
                    options = [item.task_priority for item in TaskPriority.objects.all()]
                elif field_id == 'urgency':
                    options = ['Urgent & Important', 'Less Urgent but Important', 'Urgent but less Important',
                                    'Neither Urgent nor Important']
                else:
                    options = [choice.f_choice for choice in item.proj_temp_field.field_choice.all()]
                single_select_column = {
                    'id': field_id,
                    'minWidth': 140,
                    'header': [{'text': field_title, 'align': 'left'}, {'content': 'selectFilter'}],
                    'editorType': 'combobox',
                    'options': options,
                    'hidden': hidden,
                    'editable': column_editable,
                }
                columns.append(single_select_column)
            elif type == 'multi-select':
                if field_id == 'assignee':
                    options = user_list
                else:
                    options = [choice.f_choice for choice in item.proj_temp_field.field_choice.all()]
                multi_select_column = {
                    'id': field_id,
                    'minWidth': 250,
                    'header': [{'text': field_title, 'align': 'left'}, {'content': 'selectFilter'}],
                    'editorType': 'multiselect',
                    'options': options,
                    'hidden': hidden,
                    'editable': column_editable,
                }
                columns.append(multi_select_column)
            elif type == 'people':
                options = user_list
                multi_select_column = {
                    'id': field_id,
                    'minWidth': 250,
                    'header': [{'text': field_title, 'align': 'left'}, {'content': 'selectFilter'}],
                    'editorType': 'multiselect',
                    'options': options,
                    'hidden': hidden,
                    'customType': 'people',
                    'editable': column_editable,
                }
                columns.append(multi_select_column)
            elif type == 'signature':
                string_column = {
                    'id': field_id,
                    'type': "string",
                    'editorType': "input",
                    'minWidth': 250,
                    'header': [{'text': field_title, 'align': 'left'}],
                    'hidden': hidden,
                    'editable': column_editable,
                    'customType': 'signature',
                }
                columns.append(string_column)
            else:
                pass

        if admin_group in user.groups.all():
            task_obj_list = Task.objects.filter(project=project_obj, task_deleted=False).distinct()
        else:
            task_obj_list = Task.objects.filter(Q(project=project_obj, task_deleted=False, project__proj_doc_group__in=user.groups.all()) |
                                               Q(project=project_obj, task_deleted=False, project__proj_doc_group__in=[common_group]) |
                                               Q(project=project_obj, task_deleted=False, project__proj_owner=user) |
                                               Q(project=project_obj, task_deleted=False, created_by=user)).distinct()

        task_data_list = []
        for task in task_obj_list:

            sort_order = task.sort_order
            id = task.id
            task_name = task.text
            task_progress = round(task.progress * 100)
            try:
                assignee = ', '.join([user.username for user in task.assigned_to.all()])
            except:
                assignee = ""
            try:
                task_status = task.status.task_status
            except:
                task_status=""
            try:
                task_priority = task.priority.task_priority
            except:
                task_priority=""
            try:
                due_date = task.end_date.strftime('%d %b, %Y')
            except:
                due_date=""
            try:
                start_date = task.start_date.strftime('%d %b, %Y')
            except:
                start_date=""
            try:
                section = task.is_section
            except:
                section = False
            if section:
                is_section = "True"
            else:
                is_section = "False"

            dict_section_label = task.section_label
            if dict_section_label:
                section_label = dict(dict_section_label)
            else:
                section_label = ""

            parent_value = task.parent
            if parent_value == '0':
                parent = 'treegrid'
            else:
                parent = parent_value
            task_data = []
            extra_column_dict = {}
            task_extra_column_list = task.task_extra_column

            if task_extra_column_list:
                extra_column_dict = task_extra_column_list["column_details"]
                extra_column_dict.update({'id': id})

            task_data.append({'id': id, 'task_name': task_name, 'assignee': assignee,
                              'due_date': due_date, 'start_date': start_date, 'status': task_status,
                              'priority': task_priority, 'parent': parent, 'sort_order': sort_order,
                              'progress': task_progress, 'is_section': is_section,
                              'section_label': section_label,
                              })

            task_data.append(extra_column_dict)

            d1 = {}
            for i in task_data:
                d1.update(i)
            task_data_list.append(d1)
            task_data_list = sorted(task_data_list, key=lambda x: x['sort_order'], reverse=False)

        time_zone = get_object_or_404(Profile, user=self.request.user).client_tz

        # notification new code
        number_of_days_list1 = [str(x) for x in range(1, 31)]
        number_of_days_list2 = ['45', '60', '75', '90']
        number_of_days_list = number_of_days_list1 + number_of_days_list2
        number_of_hours_list = [str(x) for x in range(1, 24)]

        # project task notification
        try:
            proj_task_notification_obj = get_object_or_404(ProjectTaskNotification, notify_proj_id=project_id)
        except Exception as e:
            print("error:->",e)
            proj_task_notification_obj=""

        groups = ', '.join(map(str, project_obj.proj_doc_group.all()))
        if proj_task_notification_obj !="":
            notification_action = 'Update'
            groups_on_off = ', '.join(map(str, proj_task_notification_obj.notify_groups.all()))
            notify_other = proj_task_notification_obj.notify_other
            notify = proj_task_notification_obj.notify
            notify_when_list = [item.notify_choice for item in proj_task_notification_obj.notify_when_choice.all()]
            task_starting = False
            task_due = False
            task_completed = False
            delayed_by = False
            reminder_on_whatsapp = False
            x_days_before_start_date = False
            start_date_selected_day = ''
            x_hours_before_start_date = False
            start_date_selected_hour = ''
            due_in_x_days = False
            due_date_selected_day = ''
            due_in_x_hours = False
            due_date_selected_hour = ''
            for item in notify_when_list:
                if item == 'Task starting':
                    task_starting = True
                elif item == 'Task due':
                    task_due = True
                elif item == 'Task completed':
                    task_completed = True
                elif item == 'Delayed by':
                    delayed_by = True
                elif item == 'Reminder on whatsapp':
                    reminder_on_whatsapp = True
                elif 'Due in' in item:
                    if 'Due in' and 'days' in item:
                        due_in_x_days = True
                        due_date_selected_day = str(item.split(" ")[2])
                    else:
                        due_in_x_hours = True
                        due_date_selected_hour = str(item.split(" ")[2])
                elif 'before start date' in item:
                    if 'days before start date' in item:
                        x_days_before_start_date = True
                        start_date_selected_day = str(item.split(" ")[0])
                    else:
                        x_hours_before_start_date = True
                        start_date_selected_hour = str(item.split(" ")[0])
                else:
                    task_starting = ''
                    task_due = ''
                    task_completed = ''
                    delayed_by = ''
                    reminder_on_whatsapp = ''
                    x_days_before_start_date = ''
                    start_date_selected_day = ''
                    x_hours_before_start_date = ''
                    start_date_selected_hour = ''
                    due_in_x_days = ''
                    due_date_selected_day = ''
                    due_in_x_hours = ''
                    due_date_selected_hour = ''
        else:
            notification_action = 'Create'
            notify_other = []
            notify = ''
            task_starting = ''
            task_due = ''
            task_completed = ''
            delayed_by = ''
            reminder_on_whatsapp = ''
            groups_on_off = ''
            x_days_before_start_date = ''
            start_date_selected_day = ''
            x_hours_before_start_date = ''
            start_date_selected_hour = ''
            due_in_x_days = ''
            due_date_selected_day = ''
            due_in_x_hours = ''
            due_date_selected_hour = ''

        current_site = get_current_site(request)
        current_site_domain = current_site.domain

        if request.user.is_superuser or admin_group in user.groups.all() or request.user.username == 'vspladmin':
            show_hide_column_restriction = True
        else:
            show_hide_column_restriction = False



        context = {
            'current_site_domain':current_site_domain,
            'columns':columns,
            'show_hide_column_name_list': show_hide_column_name_list,
            'task_data_list': task_data_list,
            'project_id': project_id,
            'time_zone': time_zone,
            'client_name': connection.tenant.name,
            'notification_action': notification_action,
            'groups': groups,
            'groups_on_off':groups_on_off,
            'notify_other': notify_other,
            'notify':notify,
            'task_starting':task_starting,
            'task_due':task_due,
            'task_completed':task_completed,
            'delayed_by':delayed_by,
            'reminder_on_whatsapp':reminder_on_whatsapp,
            'project_name':project_name,
            'project_description':project_description,
            'user':request.user,
            'show_hide_column_restriction':show_hide_column_restriction,
            'number_of_days_list':number_of_days_list,
            'number_of_hours_list':number_of_hours_list,
            'x_days_before_start_date':x_days_before_start_date,
            'start_date_selected_day':start_date_selected_day,
            'x_hours_before_start_date':x_hours_before_start_date,
            'start_date_selected_hour':start_date_selected_hour,
            'due_in_x_days':due_in_x_days,
            'due_date_selected_day':due_date_selected_day,
            'due_in_x_hours':due_in_x_hours,
            'due_date_selected_hour':due_date_selected_hour,
            'delete_permissions': delete_permissions,
            'proj_lock': proj_lock,
        }
        return render(request, 'project_management/tree_grid_view.html', context)


class CheckFieldDuplication(LoginRequiredMixin, View):
    def get(self, request):
        field_input = request.GET.get('field_title')
        field = ProjectField.objects.filter(field_title__iexact=field_input)
        if len(field) > 0:
            error = 'Field exists'
            return JsonResponse({'success': False, 'error': error})
        else:
            return JsonResponse({'success': True})


class CreateProjectTemplate(LoginRequiredMixin, View):

    def get(self, request):

        user_groups = Group.objects.all()
        field_types = []
        for field in FIELD_TYPE:
            field_types.append(field[0])
        field_choices = FieldChoice.objects.all()
        project_field_list = [item.field_title for item in ProjectField.objects.all()]
        field_list = ['task_name', 'assignee', 'start_date', 'due_date', 'status', 'priority', 'urgency']
        project_fields = []
        for field in project_field_list:
            if str(field).lower() in field_list:
                pass
            else:
                project_fields.append(field)

        context = {
            'user_groups': user_groups,
            'field_choices': field_choices,
            'field_types': field_types,
            'project_fields': project_fields,
            'client_name': connection.tenant.name,
        }

        return render(request, 'project_management/create_project_template.html', context)


    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            fields_storage = request.POST.get('fieldsStorage')
            temp_details = request.POST.get('temp_details')
            fields_details_list = json.loads(fields_storage)
            temp_details_list = json.loads(temp_details)

            temp_title = temp_details_list[0]["temp_title"]
            temp_description = temp_details_list[0]["temp_description"]
            temp_group = temp_details_list[0]["temp_group"]

            _now = datetime.datetime.now()
            proj_temp_id = _now.strftime('%Y%m%d%H%M%S%f')
            proj_template = ProjectTemplate.objects.create(proj_temp_id=proj_temp_id,
                                                           proj_temp_status='ac',
                                                           proj_temp_title=temp_title,
                                                           proj_temp_description=temp_description,
                                                           proj_temp_created_at=_now,
                                                           proj_temp_owner=request.user,
                                                           proj_temp_edited_by=request.user,
                                                           )
            for group in temp_group:
                gp = get_object_or_404(Group, name=group)
                proj_template.proj_temp_doc_group.add(gp)
            proj_template.save()

            # get or create field choices
            field_choice_list = ['Not Started', 'In Progress', 'Completed', 'On Hold', 'Delayed', 'Low', 'Medium', 'High', 'Urgent & Important', 'Less Urgent but Important', 'Urgent but less Important', 'Neither Urgent nor Important']
            for f_choice in field_choice_list:
                get_field_choice, create_field_choice = FieldChoice.objects.get_or_create(f_choice=f_choice)

            field_origin = proj_template

            status_list = ['Not Started', 'In Progress', 'Completed', 'On Hold', 'Delayed']
            priority_list = ['Low', 'Medium', 'High']
            urgency_list = ['Urgent & Important', 'Less Urgent but Important', 'Urgent but less Important', 'Neither Urgent nor Important']
            # get or create project field
            project_field_list = [
                {'field_title': 'task_name', 'field_description': 'Task name',
                 'field_type': 'string', 'field_origin': field_origin, 'field_choice': []
                 },
                {'field_title': 'assignee', 'field_description': 'Assignee',
                 'field_type': 'multi-select', 'field_origin': field_origin, 'field_choice': []
                 },
                {'field_title': 'start_date', 'field_description': 'Start date', 'field_type': 'date',
                 'field_origin': field_origin, 'field_choice': []
                 },
                {'field_title': 'due_date', 'field_description': 'Due date',
                 'field_type': 'date', 'field_origin': field_origin, 'field_choice': []
                 },
                {'field_title': 'status', 'field_description': 'Status',
                'field_type': 'single-select', 'field_origin': field_origin,'field_choice': status_list
                 },
                {'field_title': 'priority', 'field_description': 'Priority',
                'field_type': 'single-select', 'field_origin': field_origin,'field_choice': priority_list
                 },
                {'field_title': 'urgency', 'field_description': 'Urgency',
                'field_type': 'single-select', 'field_origin': field_origin,'field_choice': urgency_list
                 }
                ]

            try:
                for i, item in enumerate(project_field_list):
                    field_position = str(i + 1)
                    field_title = item["field_title"]
                    field_description = item["field_description"]
                    field_type = item["field_type"]
                    field_origin = item["field_origin"]
                    field_choice = item["field_choice"]

                    try:
                        get_project_field, create_project_field = ProjectField.objects.get_or_create(
                            field_title=field_title,
                            field_description=field_description,
                            field_type=field_type,
                            field_origin=field_origin
                        )
                    except:
                        get_project_field = get_object_or_404(ProjectField, field_title=field_title)

                    if field_choice:
                        for f_choice in field_choice:
                            try:
                                fc = FieldChoice.objects.create(f_choice=f_choice)
                            except IntegrityError:
                                fc = get_object_or_404(FieldChoice, f_choice=f_choice)
                            get_project_field.field_choice.add(fc)

                    ProjectTemplateField.objects.create(proj_temp=proj_template,
                                                        proj_temp_field=get_project_field,
                                                        proj_temp_field_position=field_position
                                                        )

            except Exception as e:
                print("post CreateProjectTemplate error1:->", e)

            try:
                for i, item in enumerate(fields_details_list):
                    field_position = str(8+i)
                    field_title = item["field_title"]
                    field_description = item["field_description"]
                    field_type = item["field_type"]
                    field_choice_new = item["field_choice"]
                    field_choice = [ch for ch in field_choice_new if ch]
                    try:
                        get_project_field, create_project_field = ProjectField.objects.get_or_create(
                            field_title=field_title,
                            field_description=field_description,
                            field_type=field_type,
                            field_origin=field_origin
                        )
                    except:
                        get_project_field = get_object_or_404(ProjectField, field_title=field_title)

                    if field_choice:
                        for f_choice in field_choice:
                            if f_choice:
                                try:
                                    fc = FieldChoice.objects.create(f_choice=f_choice)
                                except IntegrityError:
                                    fc = get_object_or_404(FieldChoice, f_choice=f_choice)
                                get_project_field.field_choice.add(fc)

                    ProjectTemplateField.objects.create(proj_temp=proj_template,
                                                        proj_temp_field=get_project_field,
                                                        proj_temp_field_position=field_position
                                                        )

            except Exception as e:
                print("post CreateProjectTemplate error2:->", e)

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class HideColumn(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            field_details = request.POST.get('field_details')
            field_title = field_details.split("/")[0]
            proj_temp_id = field_details.split("/")[1]
            proj_temp_obj = get_object_or_404(ProjectTemplate, proj_temp_id=proj_temp_id)
            proj_field_obj = ProjectField.objects.get(field_title=field_title)
            proj_temp_field_obj = ProjectTemplateField.objects.get(proj_temp=proj_temp_obj,proj_temp_field=proj_field_obj)
            proj_temp_field_obj.proj_temp_field_hide = True
            proj_temp_field_obj.save()
            response = {
                'success': True,
                'field_title':field_title,
            }
            return JsonResponse(response, safe=False)


class ShowColumn(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            field_details = request.POST.get('field_details')
            field_title = field_details.split("/")[0]
            proj_temp_id = field_details.split("/")[1]
            proj_temp_obj = get_object_or_404(ProjectTemplate, proj_temp_id=proj_temp_id)
            proj_field_obj = ProjectField.objects.get(field_title=field_title)
            proj_temp_field_obj = ProjectTemplateField.objects.get(proj_temp=proj_temp_obj, proj_temp_field=proj_field_obj)
            proj_temp_field_obj.proj_temp_field_hide = False
            proj_temp_field_obj.save()

            response = {
                'success': True,
                'field_title':field_title,
            }
            return JsonResponse(response, safe=False)


class CreateNewProject(LoginRequiredMixin, View):

    def get(self, request):

        context = {
            'client_name': connection.tenant.name,
        }
        return render(request, 'project_management/create_new_project.html', context)


class BlankProject(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):
        proj_groups = Group.objects.all()

        context = {
            'proj_groups': proj_groups,
            'client_name': connection.tenant.name,
        }
        return render(request, 'project_management/blank_project.html', context)

    def post(self, request, *args, **kwargs):

        proj_name = request.POST['projectName']
        proj_description = request.POST['projectDescription']
        proj_owner = request.user
        project = Project.objects.create(proj_name=proj_name,
                                         proj_description=proj_description,
                                         proj_owner=proj_owner,
                                         proj_start_date=datetime.datetime.now(),
                                         )

        groups = request.POST.getlist('blank_proj_group_select')
        for group in groups:
            gp = get_object_or_404(Group, name=group)
            project.proj_doc_group.add(gp)
        project.save()

        # get or create Project Status
        project_status_choice_list = ['Not Started', 'In Progress', 'Completed', 'On Hold']
        for p_status in project_status_choice_list:
            get_project_status, create_project_status = ProjectStatus.objects.get_or_create(proj_status=p_status)

        # save project status to Not Started
        get_proj_status_obj = ProjectStatus.objects.get(proj_status="Not Started")
        project.proj_status = get_proj_status_obj
        project.save()

        _now = datetime.datetime.now()
        proj_temp_id = _now.strftime('%Y%m%d%H%M%S%f')
        proj_template = ProjectTemplate.objects.create(proj_temp_id=proj_temp_id,
                                                       proj_temp_status='ac',
                                                       proj_temp_title=proj_name+' template '+str(proj_temp_id),
                                                       proj_temp_description=proj_name+' template '+str(proj_temp_id),
                                                       proj_temp_created_at=_now,
                                                       proj_temp_owner=request.user,
                                                       proj_temp_edited_by=request.user,
                                                       )
        cg = get_object_or_404(Group, name='COMMON')
        proj_template.proj_temp_doc_group.add(cg)
        proj_template.save()

        # save project template
        project.proj_origin = proj_template
        project.save()

        # get or create field choices
        field_choice_list = ['Not Started', 'In Progress', 'Completed', 'On Hold', 'Low', 'Medium', 'High', 'Urgent & Important',
                             'Less Urgent but Important', 'Urgent but less Important', 'Neither Urgent nor Important']
        for f_choice in field_choice_list:
            get_field_choice, create_field_choice = FieldChoice.objects.get_or_create(f_choice=f_choice)

        field_origin = proj_template

        status_list = ['Not Started', 'In Progress', 'Completed', 'On Hold', 'Delayed']
        priority_list = ['Low', 'Medium', 'High']
        urgency_list = ['Urgent & Important', 'Less Urgent but Important', 'Urgent but less Important',
                        'Neither Urgent nor Important']
        # get or create project field
        project_field_list = [{'field_title':'task_name','field_description':'Task name', 'field_type':'string','field_origin':field_origin,'field_choice':[]},
                              {'field_title':'assignee','field_description':'Assignee', 'field_type':'multi-select','field_origin':field_origin,'field_choice':[]},
                              {'field_title':'start_date','field_description':'Start date', 'field_type':'date','field_origin':field_origin,'field_choice':[]},
                              {'field_title': 'due_date', 'field_description': 'Due date', 'field_type': 'date','field_origin': field_origin, 'field_choice': []},
                              {'field_title':'status','field_description':'Status', 'field_type':'single-select','field_origin':field_origin,'field_choice':status_list},
                              {'field_title':'priority','field_description':'Priority', 'field_type':'single-select','field_origin':field_origin,'field_choice':priority_list},
                              {'field_title': 'urgency', 'field_description': 'Urgency', 'field_type': 'single-select','field_origin': field_origin, 'field_choice': urgency_list}
                              ]

        try:
            for i, item in enumerate(project_field_list):
                field_position = str(i+1)
                field_title = item["field_title"]
                field_description = item["field_description"]
                field_type = item["field_type"]
                field_origin = item["field_origin"]
                field_choice = item["field_choice"]

                try:
                    get_project_field, create_project_field = ProjectField.objects.get_or_create(field_title=field_title,
                                                                                                 field_description=field_description,
                                                                                                 field_type=field_type,
                                                                                                 field_origin=field_origin)
                except:
                    get_project_field = get_object_or_404(ProjectField, field_title=field_title)

                if field_choice:
                    for f_choice in field_choice:
                        try:
                            fc = FieldChoice.objects.create(f_choice=f_choice)
                        except IntegrityError:
                            fc = get_object_or_404(FieldChoice, f_choice=f_choice)
                        get_project_field.field_choice.add(fc)

                ProjectTemplateField.objects.create(proj_temp=proj_template, proj_temp_field=get_project_field, proj_temp_field_position=field_position)

        except Exception as e:
            print("post BlankProject error:->",e)

        # get project id
        proj_id = project.proj_id

        # add new task
        my_project = get_object_or_404(Project, proj_id=proj_id)
        _now = datetime.datetime.now()

        tree_root_sort_order = Task.objects.filter(project=project, task_deleted=False, parent="0").count()
        sort_order = tree_root_sort_order + 1

        create_task = Task.objects.create(text='New Task',
                                          parent='0',
                                          project=my_project,
                                          created_by=request.user,
                                          start_date=_now,
                                          sort_order=sort_order)
        # add status
        get_status, create_status = TaskStatus.objects.get_or_create(task_status="Not Started")
        create_task.status = get_status
        create_task.save()
        # add priority
        get_priority, create_priority = TaskPriority.objects.get_or_create(task_priority="Low")
        create_task.priority = get_priority
        create_task.save()

        task_id = create_task.id

        # call create task log history func
        my_task_name = create_task.text
        my_project_name = my_project.proj_name
        task_log_status = 'TASK_ADD'
        task_log_message = 'Task ID: ' + str(task_id) + ', Task Name: ' + str(
            my_task_name) + ' added to Project ID: ' + str(proj_id) + ', Project Name: ' + str(
            my_project_name) + ' successfully.'
        createTaskLogHistory(request, proj_id, task_id, task_log_status, task_log_message)

        # call create project log history func
        my_project_name = my_project.proj_name
        my_project_id = str(proj_id)
        project_log_status = 'PROJECT_ADD'
        project_log_message = 'Project ID: ' + my_project_id + ', Project Name: ' + str(my_project_name) + ' added successfully.'
        createProjectLogHistory(request, my_project_id, project_log_status, project_log_message)

        return redirect('tree-grid-view', proj_id)


class UseTemplate(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):
        proj_groups = Group.objects.all()

        proj_template = ProjectTemplate.objects.all()

        context = {
            'proj_groups': proj_groups,
            'proj_template':proj_template,
            'client_name': connection.tenant.name,
        }
        return render(request, 'project_management/use_template.html', context)

    def post(self, request, *args, **kwargs):

        proj_name = request.POST['projectName']
        proj_description = request.POST['projectDescription']
        proj_owner = request.user
        project = Project.objects.create(proj_name=proj_name,
                                         proj_description=proj_description,
                                         proj_owner=proj_owner,
                                         proj_start_date=datetime.datetime.now(),
                                         )

        groups = request.POST.getlist('use_template_group_select')
        for group in groups:
            gp = get_object_or_404(Group, name=group)
            project.proj_doc_group.add(gp)
        project.save()

        # get or create Project Status
        project_status_choice_list = ['Not Started', 'In Progress', 'Completed', 'On Hold']
        for p_status in project_status_choice_list:
            get_project_status, create_project_status = ProjectStatus.objects.get_or_create(proj_status=p_status)

        # save project status to Not Started
        get_proj_status_obj = ProjectStatus.objects.get(proj_status="Not Started")
        project.proj_status = get_proj_status_obj
        project.save()

        library_temp_id = request.POST.get('library_template_select')
        library_template = get_object_or_404(ProjectTemplate, proj_temp_id=library_temp_id)

        _now = datetime.datetime.now()
        proj_temp_id = _now.strftime('%Y%m%d%H%M%S%f')
        proj_template = ProjectTemplate.objects.create(proj_temp_id=proj_temp_id,
                                                       proj_temp_status='ac',
                                                       proj_temp_title=proj_name+' template '+str(proj_temp_id),
                                                       proj_temp_description=proj_name+' template '+str(proj_temp_id),
                                                       proj_temp_created_at=_now,
                                                       proj_temp_owner=request.user,
                                                       proj_temp_edited_by=request.user,
                                                       )
        cg = get_object_or_404(Group, name='COMMON')
        proj_template.proj_temp_doc_group.add(cg)
        proj_template.save()

        # save project template
        project.proj_origin = proj_template
        project.save()

        library_temp_field = ProjectTemplateField.objects.filter(proj_temp=library_template).order_by('proj_template_id')

        try:
            for i, item in enumerate(library_temp_field):
                field_position = item.proj_temp_field_position
                get_project_field = item.proj_temp_field
                proj_temp_field_hide = item.proj_temp_field_hide

                ProjectTemplateField.objects.create(proj_temp=proj_template,
                                                    proj_temp_field=get_project_field,
                                                    proj_temp_field_position=field_position,
                                                    proj_temp_field_hide=proj_temp_field_hide)

        except Exception as e:
            print("post UseTemplate error:->",e)

        # get project id
        proj_id = project.proj_id

        # add new task
        my_project = get_object_or_404(Project, proj_id=proj_id)
        _now = datetime.datetime.now()

        tree_root_sort_order = Task.objects.filter(project=project, task_deleted=False, parent="0").count()
        sort_order = tree_root_sort_order + 1

        create_task = Task.objects.create(text='New Task',
                                          parent='0',
                                          project=my_project,
                                          created_by=request.user,
                                          start_date=_now,
                                          sort_order=sort_order)
        # add status
        get_status, create_status = TaskStatus.objects.get_or_create(task_status="Not Started")
        create_task.status = get_status
        create_task.save()
        # add priority
        get_priority, create_priority = TaskPriority.objects.get_or_create(task_priority="Low")
        create_task.priority = get_priority
        create_task.save()

        task_id = create_task.id

        # call create task log history func
        my_task_name = create_task.text
        my_project_name = my_project.proj_name
        task_log_status = 'TASK_ADD'
        task_log_message = 'Task ID: ' + str(task_id) + ', Task Name: ' + str(
            my_task_name) + ' added to Project ID: ' + str(proj_id) + ', Project Name: ' + str(
            my_project_name) + ' successfully.'
        createTaskLogHistory(request, proj_id, task_id, task_log_status, task_log_message)

        # call create project log history func
        my_project_name = my_project.proj_name
        my_project_id = str(proj_id)
        project_log_status = 'PROJECT_ADD'
        project_log_message = 'Project ID: ' + my_project_id + ', Project Name: ' + str(
            my_project_name) + ' added successfully.'
        createProjectLogHistory(request, my_project_id, project_log_status, project_log_message)

        return redirect('tree-grid-view', proj_id)


class AddNewTask(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            project_id = request.POST.get('project_id')
            project = get_object_or_404(Project, proj_id=project_id)
            _now = datetime.datetime.now()

            tree_root_sort_order = Task.objects.filter(project=project, task_deleted=False, parent="0").count()
            sort_order = tree_root_sort_order+1

            create_task = Task.objects.create(text='New Task',
                                              parent='0',
                                              project=project,
                                              created_by=request.user,
                                              start_date=_now,
                                              sort_order=sort_order,
                                              progress=0.0)
            task_id = create_task.id
            start_date = _now.strftime('%d %b, %Y')
            progress = 0

            # call create task log history func
            task_name = create_task.text
            project_name = project.proj_name
            task_log_status = 'TASK_ADD'
            task_log_message = 'Task ID: '+str(task_id)+ ', Task Name: '+str(task_name)+' added to Project ID: '+str(project_id)+', Project Name: '+str(project_name)+' successfully.'
            createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)

            response = {
                'success': True,
                'task_id':task_id,
                'start_date':start_date,
                'progress':progress,
            }
            return JsonResponse(response, safe=False)


class AddNewSubtask(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            project_id = request.POST.get('project_id')
            parent = request.POST.get('parent')
            project = get_object_or_404(Project, proj_id=project_id)
            _now = datetime.datetime.now()

            parent_sort_order = Task.objects.filter(project=project, task_deleted=False, parent=parent).count()
            sort_order = parent_sort_order + 1

            create_task = Task.objects.create(text='New Subtask',
                                              parent=parent,
                                              project=project,
                                              created_by=request.user,
                                              start_date=_now,
                                              sort_order=sort_order,
                                              progress=0.0)
            parent_id = create_task.parent
            task_id = create_task.id
            start_date = _now.strftime('%d %b, %Y')
            progress = 0

            # call create task log history func
            task_name = create_task.text
            project_name = project.proj_name
            task_log_status = 'TASK_ADD'
            task_log_message = 'Task ID: '+str(task_id)+ ', Task Name: '+str(task_name)+' added to Project ID: '+str(project_id)+', Project Name: '+str(project_name)+' successfully.'
            createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)

            response = {
                'success': True,
                'parent_id': parent_id,
                'start_date': start_date,
                'task_id': task_id,
                'progress': progress,
            }
            return JsonResponse(response, safe=False)


class RemoveTask(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            task_id = request.POST.get('task_id')
            task = Task.objects.get(pk=task_id)
            result = [[task.id]]

            def path_to_root_task(task_id):
                task = Task.objects.filter(parent=task_id, task_deleted=False)
                child = []
                for item in task:
                    child.append(item.id)
                    path_to_root_task(item.id)
                if child:
                    result.append(child)

            path_to_root_task(task.id)
            #child_task_list = list(np.concatenate(result))
            child_task_list = [item for sublist in result for item in sublist]

            for id in child_task_list:
                task_obj = Task.objects.get(pk=id)
                project_id = task_obj.project.proj_id
                project_name = task_obj.project.proj_name
                task_obj.task_deleted = True
                task_obj.save()

                # call create task log history func
                task_name = task_obj.text
                task_log_status = 'TASK_DEL'
                task_log_message = 'Task ID: ' + str(id) + ', Task Name: '+str(task_name)+' deleted from Project ID: '+str(project_id)+', Project Name: '+str(project_name)+' successfully.'
                createTaskLogHistory(request, project_id, id, task_log_status, task_log_message)

            response = {
                'task_id':task_id,
                'success': True,
            }
            return JsonResponse(response, safe=False)


class DuplicateTask(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            task_id = request.POST.get('task_id')
            task = Task.objects.get(pk=task_id)
            project = task.project

            tree_root_sort_order = Task.objects.filter(project=project, task_deleted=False, parent="0").count()
            sort_order = tree_root_sort_order + 1

            parent_task_obj = Task.objects.get(pk=task.id)
            assigned_to_list = [user.username for user in parent_task_obj.assigned_to.all()]
            parent_task_obj.pk = None
            parent_task_obj.save()

            # call create task log history func
            parent_task_id = parent_task_obj.id
            project_id = project.proj_id
            project_name = project.proj_name
            task_name = parent_task_obj.text
            task_log_status = 'TASK_DUPL'
            task_log_message = 'Task ID: ' + str(parent_task_id) + ', Task Name: '+str(task_name)+' duplicated for Project ID: ' + str(
                project_id) + ', Project Name: ' + str(project_name) + ' successfully.'
            createTaskLogHistory(request, project_id, parent_task_id, task_log_status, task_log_message)

            parent_task_obj.sort_order = sort_order
            parent_task_obj.assigned_to.clear()
            assignee_list = [item.strip() for item in assigned_to_list]
            for usr in assignee_list:
                usr = usr.strip()
                user = get_object_or_404(User, username=usr)
                parent_task_obj.assigned_to.add(user)
            parent_task_obj.save()

            def path_to_root_task(task_id):
                child_task = Task.objects.filter(parent=task_id, task_deleted=False)
                for i, item in enumerate(child_task):
                    task_obj = Task.objects.get(pk=item.id)
                    child_assigned_to_list = [user.username for user in task_obj.assigned_to.all()]

                    task_obj.pk = None
                    task_obj.save()
                    task_obj.parent = task_obj.id - (i+1)
                    task_obj.assigned_to.clear()
                    child_assignee_list = [item.strip() for item in child_assigned_to_list]
                    for usr in child_assignee_list:
                        usr = usr.strip()
                        user = get_object_or_404(User, username=usr)
                        task_obj.assigned_to.add(user)
                    task_obj.save()
                    # call create task log history func
                    child_task_id = task_obj.id
                    project_id = project.proj_id
                    project_name = project.proj_name
                    task_name = task_obj.text
                    task_log_status = 'TASK_DUPL'
                    task_log_message = 'Task ID: ' + str(child_task_id) + ', Task Name: '+str(task_name)+' duplicated for Project ID: ' + str(
                        project_id) + ', Project Name: ' + str(project_name) + ' successfully.'
                    createTaskLogHistory(request, project_id, child_task_id, task_log_status, task_log_message)
                    path_to_root_task(item.id)

            path_to_root_task(task.id)

            response = {
                'task_id':task_id,
                'success': True,
            }
            return JsonResponse(response, safe=False)


class RemoveAllTask(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            proj_id = request.POST.get('proj_id')
            project = get_object_or_404(Project, proj_id=proj_id)
            all_task = Task.objects.filter(project=project, task_deleted=False)

            for task_obj in all_task:
                project_id = task_obj.project.proj_id
                project_name = task_obj.project.proj_name
                task_obj.task_deleted = True
                task_obj.save()

                # call create task log history func
                task_id = task_obj.id
                task_name = task_obj.text
                task_log_status = 'TASK_DEL'
                task_log_message = 'Task ID: ' + str(task_id) + ', Task Name: '+str(task_name)+' deleted from Project ID: '+str(project_id)+', Project Name: '+str(project_name)+' successfully.'
                createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class ProjectCardView(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):

        user = self.request.user
        common_group = get_object_or_404(Group, name='COMMON')
        admin_group = get_object_or_404(Group, name='DOC_ADMIN')

        if admin_group in user.groups.all():
            project_obj_list = Project.objects.filter(proj_deleted=False, proj_archive=False).order_by('-proj_id').distinct()
        else:
            project_obj_list = Project.objects.filter(Q(proj_deleted=False, proj_archive=False, proj_doc_group__in=user.groups.all()) |
                                               Q(proj_deleted=False, proj_archive=False, proj_doc_group__in=[common_group]) |
                                               Q(proj_deleted=False, proj_archive=False, proj_owner=user)).order_by('-proj_id').distinct()

        project_list = []
        try:
            for item in project_obj_list:
                proj_id = item.proj_id
                proj_name = item.proj_name
                proj_description = item.proj_description
                if item.proj_status:
                    proj_status = item.proj_status.proj_status
                else:
                    proj_status = "None"
                try:
                    proj_start_date = item.proj_start_date.strftime('%Y-%m-%d')
                except:
                    proj_start_date = ""
                try:
                    proj_end_date = item.proj_end_date.strftime('%Y-%m-%d')
                except:
                    proj_end_date = ""

                if item.proj_lock:
                    proj_lock = "true"
                else:
                    proj_lock = "false"
                project_list.append({"proj_id":proj_id, "proj_name":proj_name,
                                     "proj_description":proj_description, "proj_status":proj_status,
                                     "proj_start_date":proj_start_date, "proj_end_date":proj_end_date,
                                     "proj_lock": proj_lock,
                                     })
        except Exception as e:
            print("get ProjectCardView Error:->",e)
            project_list = []

        time_zone = get_object_or_404(Profile, user=self.request.user).client_tz

        context = {
            'project_list': project_list[:4],
            'time_zone': time_zone,
            'client_name': connection.tenant.name,
            'project_obj_list':project_obj_list[4:],
            'user': user,
            'user_is_superuser_or_staff': user.is_superuser or user.is_staff,
        }
        return render(request, 'project_management/project_card_view.html', context)


class GetUpdateProjectView(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():

            id = request.GET.get('proj_id')
            project = get_object_or_404(Project, proj_id=id)
            project_list = []
            project_status_list = [item.proj_status for item in ProjectStatus.objects.all()]
            project_group_list = [gr.name for gr in Group.objects.all()]
            section_master_names = SectionMaster.objects.values_list('section_master_name', flat=True).distinct()

            # Filter proj_dropdown_field values only from valid projects
            proj_dropdown_field_list = Project.objects.filter(
                proj_deleted=False,
                proj_archive=False,
            ).exclude(
                proj_dropdown_field__isnull=True
            ).exclude(
                proj_dropdown_field__exact=""
            ).values_list('proj_dropdown_field', flat=True).distinct()

            try:
                proj_id = project.proj_id
                proj_name = project.proj_name
                proj_description = project.proj_description
                if project.proj_status:
                    proj_status = project.proj_status.proj_status
                else:
                    proj_status = "None"
                if project.proj_doc_group:
                    proj_group = [gr.name for gr in project.proj_doc_group.all()]
                else:
                    proj_group = "None"
                try:
                    proj_start_date = project.proj_start_date.strftime('%Y-%m-%d')
                except:
                    proj_start_date = ""
                try:
                    proj_end_date = project.proj_end_date.strftime('%Y-%m-%d')
                except:
                    proj_end_date = ""

                proj_signature_value = project.proj_signature
                if proj_signature_value == "" or proj_signature_value == " " or proj_signature_value == None:
                    proj_signature = ""
                else:
                    proj_signature = proj_signature_value

                proj_section_master_name_value = project.proj_section_master_name
                if proj_section_master_name_value == "" or proj_section_master_name_value == " " or proj_section_master_name_value == None:
                    proj_section_master_name = ""
                else:
                    proj_section_master_name = proj_section_master_name_value

                proj_dropdown_field_value = project.proj_dropdown_field
                if proj_dropdown_field_value == "" or proj_dropdown_field_value == " " or proj_dropdown_field_value == None:
                    proj_dropdown_field = ""
                else:
                    proj_dropdown_field = proj_dropdown_field_value

                project_list.append({"proj_id": proj_id, "proj_name": proj_name,
                                     "proj_description": proj_description, "proj_status": proj_status,
                                     "project_status_list":project_status_list,
                                     "proj_group":proj_group, "project_group_list":project_group_list,
                                     "proj_start_date": proj_start_date, "proj_end_date": proj_end_date,
                                     "proj_signature": proj_signature,
                                     "proj_master_name": proj_section_master_name,
                                     "proj_dropdown_field": proj_dropdown_field,
                                     })
            except Exception as e:
                print("get GetUpdateProjectView Error:->", e)
                project_list = []

            response = {
                'success': True,
                'project_list':project_list,
                'section_master_names': list(section_master_names),
                'proj_dropdown_field_list': list(proj_dropdown_field_list)
            }
            return JsonResponse(response, safe=False)


class PostUpdateProjectView(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            proj_id = request.POST.get('update_proj_id')
            proj_name = request.POST.get('update_proj_name')
            proj_description = request.POST.get('update_proj_description')
            proj_status = request.POST.get('update_proj_status')
            proj_start_date = request.POST.get('update_proj_start_date')
            proj_end_date = request.POST.get('update_proj_end_date')
            proj_group_list = request.POST.get('update_proj_group')
            proj_group = json.loads(proj_group_list)
            update_proj_master_name = request.POST.get('update_proj_master_name')
            update_proj_dropdown_field = request.POST.get('update_proj_dropdown_field')

            project = get_object_or_404(Project, proj_id=proj_id)

            try:
                # Track original values
                original_name = project.proj_name
                original_description = project.proj_description
                original_proj_dropdown_field = project.proj_dropdown_field
                original_master_name = project.proj_section_master_name
                original_group_set = set(project.proj_doc_group.values_list('name', flat=True))

                # --- Log status change ---
                if proj_status:
                    new_status_obj = ProjectStatus.objects.get(proj_status=proj_status)
                    if project.proj_status != new_status_obj:
                        msg = f"Project ID: {proj_id}, Project Name: {proj_name} updated successfully.<br>" \
                        f"Updated field: <span style='background-color:#ffc107; padding: 2px 5px;'>Project status changed to:{proj_status}</span>."
                        createProjectLogHistory(request, proj_id, 'PROJECT_STAT_CHNG', msg)
                        project.proj_status = new_status_obj
                else:
                    if project.proj_status is not None:
                        msg = f"Project ID: {proj_id}, Project Name: {proj_name} updated successfully.<br>"
                        f"Updated field: <span style='background-color:#ffc107; padding: 2px 5px;'>Project status changed to:None</span>."
                        createProjectLogHistory(request, proj_id, 'PROJECT_STAT_CHNG', msg)
                    project.proj_status = None

                # --- Log start date change ---
                new_start_date = None
                if proj_start_date:
                    try:
                        new_start_date = datetime.datetime.strptime(proj_start_date, "%Y-%m-%d").date()
                    except ValueError:
                        print("Invalid proj_start_date format")

                existing_start_date = project.proj_start_date.date() if project.proj_start_date else None

                if new_start_date != existing_start_date:
                    log_value = proj_start_date if proj_start_date else 'None'
                    msg = f"Project ID: {proj_id}, Project Name: {proj_name} updated successfully.<br>" \
                          f"Updated field:<span style='background-color:#ffc107; padding: 2px 5px;'>Project start date changed to:{log_value}</span>."
                    createProjectLogHistory(request, proj_id, 'PROJECT_DATE_CHNG', msg)
                    project.proj_start_date = new_start_date if new_start_date else None

                # --- Log end date change ---
                new_end_date = None
                if proj_end_date:
                    try:
                        new_end_date = datetime.datetime.strptime(proj_end_date, "%Y-%m-%d").date()
                    except ValueError:
                        print("Invalid proj_end_date format")

                existing_end_date = project.proj_end_date.date() if project.proj_end_date else None

                if new_end_date != existing_end_date:
                    log_value = proj_end_date if proj_end_date else 'None'
                    msg = f"Project ID: {proj_id}, Project Name: {proj_name} updated successfully.<br>" \
                          f"Updated field:<span style='background-color:#ffc107; padding: 2px 5px;'>Project end date changed to:{log_value}</span>."
                    createProjectLogHistory(request, proj_id, 'PROJECT_DATE_CHNG', msg)
                    project.proj_end_date = new_end_date if new_end_date else None

                # --- Apply Updates ---
                project.proj_name = proj_name
                project.proj_description = proj_description
                project.proj_dropdown_field = update_proj_dropdown_field if update_proj_dropdown_field else ""
                project.proj_section_master_name = update_proj_master_name if update_proj_master_name else ""

                # --- Update groups ---
                project.proj_doc_group.clear()
                for group in proj_group:
                    gp = get_object_or_404(Group, name=group)
                    project.proj_doc_group.add(gp)

                project.save()

                # --- Log PROJECT_UPD if basic info changed ---
                updated_group_set = set(proj_group)
                updated_fields = []

                if proj_name != original_name:
                    updated_fields.append(f"Project name changed to: {proj_name}")

                if proj_description != original_description:
                    updated_fields.append(f"Project description changed to: {proj_description}")

                if update_proj_dropdown_field != original_proj_dropdown_field:
                    updated_fields.append(f"Chassis number changed to: {update_proj_dropdown_field}")

                if update_proj_master_name != original_master_name:
                    updated_fields.append(f"Project master data model name changed to: {update_proj_master_name}")

                if original_group_set != updated_group_set:
                    updated_fields.append(f"Project group changed to: {updated_group_set}")

                if updated_fields:
                    fields_string = ", ".join(updated_fields)
                    msg = (
                        f"Project ID: {proj_id}, Project Name: {proj_name} updated successfully.<br>"
                        f"Updated fields:<span style='background-color:#ffc107; padding: 2px 5px;'>{fields_string}</span>"
                    )
                    createProjectLogHistory(request, proj_id, 'PROJECT_UPD', msg)


            except Exception as e:
                print("post PostUpdateProjectView Error:->", e)

            return JsonResponse({'success': True}, safe=False)


def project_data(request):
    if request.method == 'GET':
        user = request.user
        common_group = get_object_or_404(Group, name='COMMON')
        admin_group = get_object_or_404(Group, name='DOC_ADMIN')

        if admin_group in user.groups.all():
            project_obj_list = Project.objects.filter(proj_deleted=False, proj_archive=False).order_by('-proj_id').distinct()
        else:
            project_obj_list = Project.objects.filter(Q(proj_deleted=False, proj_archive=False, proj_doc_group__in=user.groups.all()) |
                                                      Q(proj_deleted=False, proj_archive=False, proj_doc_group__in=[common_group]) |
                                                      Q(proj_deleted=False, proj_archive=False, proj_owner=user)).order_by('-proj_id').distinct()

        project_list = []
        try:
            for item in project_obj_list:
                proj_id = item.proj_id
                proj_name = item.proj_name
                proj_description = item.proj_description
                if item.proj_status:
                    proj_status = item.proj_status.proj_status
                else:
                    proj_status = "None"
                try:
                    proj_start_date = item.proj_start_date.strftime('%Y-%m-%d')
                except:
                    proj_start_date = ""
                try:
                    proj_end_date = item.proj_end_date.strftime('%Y-%m-%d')
                except:
                    proj_end_date = ""
                project_list.append({"proj_id": proj_id, "proj_name": proj_name,
                                     "proj_description": proj_description, "proj_status": proj_status,
                                     "proj_start_date": proj_start_date, "proj_end_date": proj_end_date,
                                     })
        except Exception as e:
            print("get project_data Error:->", e)
            project_list = []

        response = {
            'success': True,
            'project_list': project_list,
        }
        return JsonResponse(response, safe=False)


class DeleteProjectCardView(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            proj_id = request.POST.get('delete_proj_id')

            try:
                project = get_object_or_404(Project, proj_id=proj_id)
                project.proj_deleted = True

                # call create project log history func
                my_project_name = str(project.proj_name)
                my_project_id = str(project.proj_id)
                project_log_status = 'PROJECT_DEL'
                project_log_message = 'Project ID: ' + my_project_id + ', Project Name: ' + my_project_name + ' deleted successfully.'
                createProjectLogHistory(request, my_project_id, project_log_status, project_log_message)

                project.save()
            except Exception as e:
                print("post DeleteProjectCardView Error:->", e)

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class EditTreeGridRow(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            row = request.POST.get('row')
            row_data = json.loads(row)
            task_id = row_data["id"]
            task = Task.objects.get(id=task_id)

            project_template = task.project.proj_origin
            project_template_field_list = ProjectTemplateField.objects.filter(proj_temp=project_template)

            title = row_data["task_name"]
            task.text = title

            try:
                start_date = row_data["start_date"]
                start_date_obj = datetime.datetime.strptime(start_date, '%d %b, %Y')
                task.start_date = start_date_obj
            except Exception as e:
                print("post EditTreeGridRow start date error:->", e)
                pass

            try:
                end_date = row_data["due_date"]
                end_date_obj = datetime.datetime.strptime(end_date, '%d %b, %Y')
                task.end_date = end_date_obj
                try:
                    def path_to_root_task(task_id):
                        new_task = Task.objects.get(pk=task_id)
                        parent_task = Task.objects.get(id=new_task.id)
                        try:
                            curr_task_date = end_date_obj.strftime('%Y/%m/%d')
                            par_task_date = parent_task.end_date.strftime('%Y/%m/%d')
                            current_task_date = datetime.datetime.strptime(curr_task_date, '%Y/%m/%d')
                            parent_task_date = datetime.datetime.strptime(par_task_date, '%Y/%m/%d')
                            delta = current_task_date - parent_task_date
                            difference = delta.days
                            if int(difference) >= 0:
                                parent_task.end_date = end_date_obj
                                parent_task.save()
                                path_to_root_task(new_task.parent)
                            else:
                                pass
                        except:
                            parent_task.end_date = end_date_obj
                            parent_task.save()
                            path_to_root_task(new_task.parent)
                    path_to_root_task(task.id)
                except Exception as e:
                    print("post EditTreeGridRow end date error1:->", e)
                    pass
            except Exception as e:
                print("post EditTreeGridRow end date error2:->",e)
                pass

            try:
                assignee = row_data["assignee"]
                assignee_list = assignee.split(",")
                task.assigned_to.clear()
                assignee_list = [item.strip() for item in assignee_list]
                for usr in assignee_list:
                    usr = usr.strip()
                    user = get_object_or_404(User,username=usr)
                    task.assigned_to.add(user)
            except Exception as e:
                print("post EditTreeGridRow assigned to error:->",e)
                pass

            try:
                task_status = row_data["status"]
                if task_status != "":
                    try:
                        assigned_status = task.status.task_status
                    except:
                        assigned_status = ""
                    get_status, create_status = TaskStatus.objects.get_or_create(task_status=task_status)
                    task.status = get_status
                    task_id = task.id
                    project_id = task.project.proj_id
                    project_name = task.project.proj_name
                    line_break = "<br>"
                    if assigned_status != task_status:
                        if task_status == 'Completed':
                            # call create task log history func
                            task_log_status = 'TASK_COMP'
                            task_log_message = "Task ID: " + str(task_id) + ", Task Name: "+str(title)+" status changed to: "+ "%s" % (line_break) +"<span style='background-color:#ffc107; padding: 2px 5px;'>"+str(task_status)+"</span> of Project ID: "+str(project_id)+", Project Name: "+str(project_name)+"."
                            createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)
                        else:
                            task_log_status = 'TASK_STAT_CHNG'
                            task_log_message = "Task ID: " + str(task_id) + ", Task Name: "+str(title)+" status changed to: " + "%s" % (line_break) + "<span style='background-color:#ffc107; padding: 2px 5px;'>" + str(task_status) + "</span> of Project ID: "+str(project_id)+", Project Name: "+str(project_name)+"."
                            createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)
                else:
                    pass
            except Exception as e:
                print("post EditTreeGridRow status error:->", e)
                pass

            try:
                task_priority = row_data["priority"]
                if task_priority != "":
                    get_priority, create_priority = TaskPriority.objects.get_or_create(task_priority=task_priority)
                    task.priority = get_priority
                else:
                    pass
            except Exception as e:
                print("post EditTreeGridRow priority error:->", e)
                pass

            task.save()

            db_column_list = ['task_name', 'assignee', 'start_date', 'due_date', 'status', 'priority']

            column_details_dict = {}
            for item in project_template_field_list:
                column_details = {}
                field_title = item.proj_temp_field.field_title
                field_type = item.proj_temp_field.field_type
                if field_title in db_column_list:
                    pass
                else:
                    try:
                        field_value = row_data[field_title]
                        if field_type == 'single-select' and field_value:
                            get_field_obj = ProjectField.objects.get(field_title=field_title)
                            try:
                                get_choice = FieldChoice.objects.get(f_choice=field_value)
                            except Exception as e:
                                print("post EditTreeGridRow single select error:->",e)
                                get_choice = FieldChoice.objects.create(f_choice=field_value)
                            get_field_obj.field_choice.add(get_choice)
                            get_field_obj.save()

                            column_details[field_title] = field_value
                            task.save()

                        elif field_type == 'multi-select':
                            get_field_obj = ProjectField.objects.get(field_title=field_title)
                            field_value_split_list = field_value.split(",")
                            field_value_list = [x for x in field_value_split_list if x]
                            for f_choice in field_value_list:
                                f_choice = f_choice.strip()
                                try:
                                    get_choice = FieldChoice.objects.get(f_choice=f_choice)
                                except Exception as e:
                                    print("post EditTreeGridRow multi select error:->",e)
                                    get_choice = FieldChoice.objects.create(f_choice=f_choice)
                                get_field_obj.field_choice.add(get_choice)
                                get_field_obj.save()

                            column_details[field_title] = field_value
                            task.save()
                        elif field_type == 'bool':
                            column_details[field_title] = str(field_value).lower()
                            task.save()
                        else:
                            if field_value == None:
                                field_value = ""
                            column_details[field_title] = field_value
                            task.save()
                    except:
                        pass
                column_details_dict.update(column_details)

            column_details_dict = {'column_details':column_details_dict}
            task.task_extra_column = column_details_dict
            task.save()

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class GetUpdateTaskData(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():

            user_list = [user.username for user in User.objects.all()]
            task_id = request.GET.get('task_id')
            task = Task.objects.get(id=task_id)

            proj_temp_obj = task.project.proj_origin
            get_proj_temp_field_obj = ProjectTemplateField.objects.filter(proj_temp=proj_temp_obj)
            column_editable_dict = {}
            for item in get_proj_temp_field_obj:
                column_name = item.proj_temp_field.field_title
                column_editable = "true"
                user_groups = request.user.groups.all()
                try:
                    if len(set(item.proj_temp_field_editable.all())) >= 1:
                        if len(set(item.proj_temp_field_editable.all()).intersection(set(user_groups))) >= 1:
                            column_editable = "false"
                        else:
                            column_editable = "true"
                        column_editable_dict[column_name] = column_editable
                    else:
                        column_editable_dict[column_name] = column_editable
                except:
                    pass

            update_task_data = []
            task_name = task.text
            update_task_data.append({'id':'task_name', 'name':'Task name', 'value':task_name, 'type':'string', 'choices':[], 'editable':'true'})
            try:
                assignee = [user.username for user in task.assigned_to.all()]
                update_task_data.append({'id':'assignee', 'name': 'Assignee', 'value': assignee, 'type': 'multi-select', 'choices': user_list, 'editable':column_editable_dict['assignee']})
            except:
                assignee = []
                update_task_data.append(
                    {'id':'assignee', 'name': 'Assignee', 'value': assignee, 'type': 'multi-select', 'choices': user_list, 'editable':column_editable_dict['assignee']})

            try:
                start_date = task.start_date.strftime('%Y-%m-%d')
                update_task_data.append(
                    {'id':'start_date', 'name': 'Start date', 'value': start_date, 'type': 'date',
                     'choices': [], 'editable':column_editable_dict['start_date']})
            except:
                start_date = ""
                update_task_data.append(
                    {'id':'start_date', 'name': 'Start date', 'value': start_date, 'type': 'date',
                     'choices': [], 'editable':column_editable_dict['start_date']})

            try:
                due_date = task.end_date.strftime('%Y-%m-%d')
                update_task_data.append(
                    {'id':'due_date', 'name': 'Due date', 'value': due_date, 'type': 'date',
                     'choices': [], 'editable':column_editable_dict['due_date']})
            except:
                due_date = ""
                update_task_data.append(
                    {'id':'due_date', 'name': 'Due date', 'value': due_date, 'type': 'date',
                     'choices': [], 'editable':column_editable_dict['due_date']})

            # get or create status choices
            status_list = ['Not Started', 'In Progress', 'Completed', 'On Hold', 'Delayed']
            for s in status_list:
                get_status_choice, create_status_choice = TaskStatus.objects.get_or_create(task_status=s)

            task_status_list = [item.task_status for item in TaskStatus.objects.all()]
            try:
                task_status = task.status.task_status
                update_task_data.append(
                    {'id':'status', 'name': 'Status', 'value': task_status, 'type': 'single-select', 'choices': task_status_list, 'editable':column_editable_dict['status']})
            except:
                task_status = ""
                update_task_data.append(
                    {'id':'status', 'name': 'Status', 'value': task_status, 'type': 'single-select', 'choices': task_status_list, 'editable':column_editable_dict['status']})

            # get or create priority choices
            priority_list = ['Low', 'Medium', 'High']
            for p in priority_list:
                get_priority_choice, create_priority_choice = TaskPriority.objects.get_or_create(task_priority=p)
            task_priority_list = [item.task_priority for item in TaskPriority.objects.all()]
            try:
                task_priority = task.priority.task_priority
                update_task_data.append(
                    {'id':'priority', 'name': 'Priority', 'value': task_priority, 'type': 'single-select', 'choices': task_priority_list, 'editable':column_editable_dict['priority']})
            except:
                task_priority = ""
                update_task_data.append(
                    {'id':'priority', 'name': 'Priority', 'value': task_priority, 'type': 'single-select',
                     'choices': task_priority_list, 'editable':column_editable_dict['priority']})

            try:
                task_description = task.note
                update_task_data.append(
                    {'id':'note', 'name': 'Task description', 'value': task_description, 'type': 'textarea',
                     'choices': [], 'editable':'true'})
            except:
                task_description = ""
                update_task_data.append(
                    {'id':'note', 'name': 'Task description', 'value': task_description, 'type': 'textarea',
                     'choices': [], 'editable':'true'})

            task_extra_column_list = task.task_extra_column
            project_template = task.project.proj_origin
            project_template_field_list = ProjectTemplateField.objects.filter(proj_temp=project_template)

            db_column_list = ['task_name', 'assignee', 'start_date', 'due_date', 'status', 'priority']

            extra_columns_list = []
            for item in project_template_field_list:
                field_title = item.proj_temp_field.field_title
                if field_title in db_column_list:
                    pass
                else:
                    extra_columns_list.append(field_title)

            if task_extra_column_list:
                extra_column_dict = dict(task_extra_column_list["column_details"])
                for k, v in extra_column_dict.items():
                    get_field_obj = ProjectField.objects.get(field_title=k)
                    field_title = get_field_obj.field_title
                    field_name = " ".join(field_title.split("_")).capitalize()
                    field_type = get_field_obj.field_type
                    if field_type == 'single-select' or field_type == 'multi-select':
                        v = [item.strip() for item in v.split(",")]
                        field_choices = [item.f_choice for item in get_field_obj.field_choice.all()]
                    elif field_type == 'people':
                        v = [item.strip() for item in v.split(",")]
                        field_choices = user_list
                    elif field_type == 'bool':
                        bool_choice_list = ["true", "false"]
                        field_choices = bool_choice_list
                    elif field_type == 'date':
                        try:
                            v = datetime.datetime.strptime(v, '%d %b, %Y')
                            v = v.strftime('%Y-%m-%d')
                        except:
                            v = ""
                        field_choices = []
                    else:
                        field_choices = []
                    update_task_data.append(
                        {'id': field_title, 'name': field_name, 'value': v, 'type': field_type,
                         'choices': field_choices, 'editable':column_editable_dict[field_title]})
                    if field_title in extra_columns_list:
                        extra_columns_list.remove(field_title)
                    else:
                        pass

            for col in extra_columns_list:
                get_field_obj = ProjectField.objects.get(field_title=col)
                field_title = get_field_obj.field_title
                field_name = " ".join(field_title.split("_")).capitalize()
                field_type = get_field_obj.field_type
                value = ""
                if field_type == 'single-select' or field_type == 'multi-select':
                    field_choices = [item.f_choice for item in get_field_obj.field_choice.all()]
                elif field_type == 'people':
                    field_choices = user_list
                elif field_type == 'bool':
                    bool_choice_list = ["true", "false"]
                    field_choices = bool_choice_list
                else:
                    field_choices = []
                update_task_data.append(
                    {'id': field_title, 'name': field_name, 'value': value, 'type': field_type,
                     'choices': field_choices, 'editable':column_editable_dict[field_title]})

            # add restriction on due date
            min_date = str(datetime.date.today())

            response = {
                'success': True,
                'update_task_data':update_task_data,
                'min_date':min_date,
            }
            return JsonResponse(response, safe=False)


class PostUpdateTaskData(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            row = request.POST.get('formData')
            row_data = json.loads(row)
            task_id = row_data["id"]
            task = Task.objects.get(id=task_id)

            project_template = task.project.proj_origin
            project_template_field_list = ProjectTemplateField.objects.filter(proj_temp=project_template)


            title = row_data["task_name"]
            task.text = title

            try:
                start_date = row_data["start_date"]
                start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d')
                task.start_date = start_date_obj
            except Exception as e:
                print("post PostUpdateTaskData start date error:->", e)
                pass

            try:
                end_date = row_data["due_date"]
                end_date_obj = datetime.datetime.strptime(end_date, '%Y-%m-%d')
                task.end_date = end_date_obj
                try:
                    def path_to_root_task(task_id):
                        new_task = Task.objects.get(pk=task_id)
                        parent_task = Task.objects.get(id=new_task.id)
                        try:
                            curr_task_date = end_date_obj.strftime('%Y/%m/%d')
                            par_task_date = parent_task.end_date.strftime('%Y/%m/%d')
                            current_task_date = datetime.datetime.strptime(curr_task_date, '%Y/%m/%d')
                            parent_task_date = datetime.datetime.strptime(par_task_date, '%Y/%m/%d')
                            delta = current_task_date - parent_task_date
                            difference = delta.days
                            if int(difference) >= 0:
                                parent_task.end_date = end_date_obj
                                parent_task.save()
                                path_to_root_task(new_task.parent)
                            else:
                                pass
                        except:
                            parent_task.end_date = end_date_obj
                            parent_task.save()
                            path_to_root_task(new_task.parent)
                    path_to_root_task(task.id)
                except Exception as e:
                    print("post PostUpdateTaskData end date error1:->", e)
                    pass
            except Exception as e:
                print("post PostUpdateTaskData end date error2:->",e)
                pass

            try:
                assignee = row_data["assignee"]
                assignee_list = assignee
                task.assigned_to.clear()
                assignee_list = [item.strip() for item in assignee_list]
                for usr in assignee_list:
                    usr = usr.strip()
                    user = get_object_or_404(User,username=usr)
                    task.assigned_to.add(user)
            except Exception as e:
                print("post PostUpdateTaskData assigned to error:->",e)
                pass

            try:
                task_status = row_data["status"]
                if task_status !="":
                    try:
                        assigned_status = task.status.task_status
                    except:
                        assigned_status = ""
                    get_status, create_status = TaskStatus.objects.get_or_create(task_status=task_status)
                    task.status = get_status
                    task_id = task.id
                    project_id = task.project.proj_id
                    project_name = task.project.proj_name
                    line_break = "<br>"
                    if assigned_status != task_status:
                        if task_status == 'Completed':
                            # call create task log history func
                            task_log_status = 'TASK_COMP'
                            task_log_message = "Task ID: " + str(task_id) + ", Task Name: "+str(title)+" status changed to: "+ "%s" % (line_break) +"<span style='background-color:#ffc107; padding: 2px 5px;'>"+str(task_status)+"</span> of Project ID: "+str(project_id)+", Project Name: "+str(project_name)+"."
                            createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)
                        else:
                            task_log_status = 'TASK_STAT_CHNG'
                            task_log_message = "Task ID: " + str(task_id) + ", Task Name: "+str(title)+" status changed to: " + "%s" % (line_break) + "<span style='background-color:#ffc107; padding: 2px 5px;'>" + str(task_status) + "</span> of Project ID: "+str(project_id)+", Project Name: "+str(project_name)+"."
                            createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)
                else:
                   pass
            except Exception as e:
                print("post PostUpdateTaskData status error:->", e)
                pass

            try:
                task_priority = row_data["priority"]
                if task_priority != "":
                    get_priority, create_priority = TaskPriority.objects.get_or_create(task_priority=task_priority)
                    task.priority = get_priority
                else:
                    pass
            except Exception as e:
                print("post PostUpdateTaskData priority error:->", e)
                pass

            try:
                note = row_data["note"]
                task.note = note
            except Exception as e:
                print("post PostUpdateTaskData note task description error:->", e)
                pass

            task.save()

            db_column_list = ['task_name', 'assignee', 'start_date', 'due_date', 'status', 'priority']

            column_details_dict = {}
            for item in project_template_field_list:
                column_details = {}
                field_title = item.proj_temp_field.field_title
                field_type = item.proj_temp_field.field_type
                if field_title in db_column_list:
                    pass
                else:
                    try:
                        field_value = row_data[field_title]
                        if field_type == 'single-select' and field_value:
                            get_field_obj = ProjectField.objects.get(field_title=field_title)
                            try:
                                get_choice = FieldChoice.objects.get(f_choice=field_value)
                            except Exception as e:
                                print("post PostUpdateTaskData single select error:->",e)
                                get_choice = FieldChoice.objects.create(f_choice=field_value)
                            get_field_obj.field_choice.add(get_choice)
                            get_field_obj.save()

                            column_details[field_title] = field_value
                            task.save()

                        elif field_type == 'multi-select':
                            get_field_obj = ProjectField.objects.get(field_title=field_title)
                            field_value_split_list = field_value
                            field_value_list = [x for x in field_value_split_list if x]
                            for f_choice in field_value_list:
                                f_choice = f_choice.strip()
                                try:
                                    get_choice = FieldChoice.objects.get(f_choice=f_choice)
                                except Exception as e:
                                    print("post PostUpdateTaskData multi select error:->",e)
                                    get_choice = FieldChoice.objects.create(f_choice=f_choice)
                                get_field_obj.field_choice.add(get_choice)
                                get_field_obj.save()

                            column_details[field_title] = ', '.join([str(elem) for elem in field_value])
                            task.save()
                        elif field_type == 'people':
                            column_details[field_title] = ', '.join([str(elem) for elem in field_value])
                            task.save()
                        elif field_type == 'date':
                            field_value = datetime.datetime.strptime(field_value, '%Y-%m-%d')
                            field_value = field_value.strftime('%d %b, %Y')
                            column_details[field_title] = field_value
                            task.save()
                        else:
                            if field_value == None:
                                field_value = ""
                            column_details[field_title] = field_value
                            task.save()
                    except:
                        pass
                column_details_dict.update(column_details)

            column_details_dict = {'column_details':column_details_dict}
            task.task_extra_column = column_details_dict
            task.save()

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class SendTaskCompletedEmail(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            task_id = request.POST.get('task_id')
            task = Task.objects.get(id=task_id)
            project_id = task.project.proj_id
            project_name = task.project.proj_name
            current_task_status = task.status.task_status
            try:
                proj_task_notification_obj = get_object_or_404(ProjectTaskNotification,
                                                               notify_proj_id=task.project.proj_id)
                notify = proj_task_notification_obj.notify
                notify_when_choice_list = [item.notify_choice for item in
                                           proj_task_notification_obj.notify_when_choice.all()]
                email_list = []
                if notify == True and 'Task completed' in notify_when_choice_list and current_task_status == 'Completed':
                    notify_groups = proj_task_notification_obj.notify_groups.all()
                    my_users = User.objects.filter(is_active=True)
                    for user in my_users:
                        user_groups = user.groups.all()
                        for group in notify_groups:
                            if group in user_groups:
                                if user.email and user.email not in email_list:
                                    email_list.append(user.email)
                    notify_other_list = proj_task_notification_obj.notify_other
                    for item in notify_other_list:
                        if item not in email_list:
                            email_list.append(item)
                    try:
                        task_status = task.status.task_status
                    except:
                        task_status = ''
                    try:
                        task_start_date = task.start_date.strftime('%d-%m-%Y')
                    except:
                        task_start_date = ''
                    try:
                        task_end_date = task.end_date.strftime('%d-%m-%Y')
                    except:
                        task_end_date = ''
                    project_task_details_list = [
                        {'project_id': project_id, 'project_name': project_name,
                         'data': [{'message_title': 'Task completed', 'delayed_by_column': 'False',
                                   'task_data': [{'task_id':task.id, 'task_name':task.text,
                                                   'status':task_status,
                                                   'start_date': task_start_date,
                                                   'due_date':task_end_date
                                                   }]
                                   }]}]
                    if project_task_details_list[0]["data"]:
                        email_subject = "Project Task Notification - Task Completed"
                        notification_data = [{'email_list': email_list, 'project_task_details_list': project_task_details_list}]
                        get_n_data, created_n_data = NotificationData.objects.get_or_create(notification_data_proj_id=project_id,
                                                                                            notification_data=notification_data)
                    else:
                        pass
            except:
                pass

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class UrgencyMatrixView(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):

        project_obj = get_object_or_404(Project, proj_id=kwargs["proj_id"])
        project_id = project_obj.proj_id
        project_name = project_obj.proj_name

        context = {
            'client_name': connection.tenant.name,
            'project_id': project_id,
            'project_name':project_name,
        }
        return render(request, 'project_management/urgency_matrix_view.html', context)


class UrgencyMatrixAjaxView(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():
            duration_value = request.GET.get('selected_value')
            proj_id = request.GET.get('proj_id')
            project = get_object_or_404(Project, proj_id=proj_id)
            proj_temp = project.proj_origin

            proj_temp_field_list = []
            urgency_matrix_error_msg = ''

            proj_temp_field_obj = ProjectTemplateField.objects.filter(proj_temp=proj_temp, proj_temp_field_hide=False)
            for item in proj_temp_field_obj:
                proj_temp_field_list.append(item.proj_temp_field.field_title)

            urgent_and_important = []
            less_urgent_but_important = []
            urgent_but_less_important = []
            neither_urgent_nor_important = []
            if 'urgency' in proj_temp_field_list:
                task_id_list = []
                days = int(duration_value)
                today_date = date.today()
                end_date = today_date + timedelta(days=days)

                filtered_tasks = Task.objects.filter(project=project,
                                                     end_date__lte=end_date, task_deleted=False).exclude(
                    status__task_status='Completed').order_by('-id')
                for task in filtered_tasks:
                    task_id = task.id
                    assigned_to = task.assigned_to.all()
                    user = request.user
                    username = request.user.username
                    try:
                        task_status = task.status.task_status
                    except:
                        task_status = ""
                    if user in assigned_to and task_status and task_id not in task_id_list:
                        task_id_list.append(task_id)
                    try:
                        task_extra_column_list = task.task_extra_column
                        if task_extra_column_list:
                            extra_column_dict = task_extra_column_list["column_details"]
                            try:
                                accountable_column_value = extra_column_dict["accountable"]
                                if accountable_column_value:
                                    accountable_list = accountable_column_value.split(",")
                                    accountable_column_value_list = [value.strip() for value in accountable_list if
                                                                     value]
                                    if username in accountable_column_value_list and task_id not in task_id_list:
                                        task_id_list.append(task_id)
                            except Exception as e:
                                print("accountable err:->", e)
                            try:
                                consulted_column_value = extra_column_dict["consulted"]
                                if consulted_column_value:
                                    consulted_list = consulted_column_value.split(",")
                                    consulted_column_value_list = [value.strip() for value in consulted_list if value]
                                    if username in consulted_column_value_list and task_id not in task_id_list:
                                        task_id_list.append(task_id)
                            except Exception as e:
                                print("consulted err:->", e)
                            try:
                                informed_column_value = extra_column_dict["informed"]
                                if informed_column_value:
                                    informed_list = informed_column_value.split(",")
                                    informed_column_value_list = [value.strip() for value in informed_list if value]
                                    if username in informed_column_value_list and task_id not in task_id_list:
                                        task_id_list.append(task_id)
                            except Exception as e:
                                print("informed err:->", e)
                    except:
                        pass

                try:
                    for task_id in task_id_list:
                        task = Task.objects.get(id=task_id)
                        task_name = task.text
                        task_extra_column_list = task.task_extra_column
                        if task_extra_column_list:
                            urgency_column_dict = task_extra_column_list["column_details"]
                            urgency_column_value = urgency_column_dict["urgency"]
                            if urgency_column_value == 'Urgent & Important':
                                urgent_and_important.append({'task_id': task_id, 'task_name': task_name})
                            elif urgency_column_value == 'Less Urgent but Important':
                                less_urgent_but_important.append({'task_id': task_id, 'task_name': task_name})
                            elif urgency_column_value == 'Urgent but less Important':
                                urgent_but_less_important.append({'task_id': task_id, 'task_name': task_name})
                            elif urgency_column_value == 'Neither Urgent nor Important':
                                neither_urgent_nor_important.append({'task_id': task_id, 'task_name': task_name})
                            else:
                                pass
                except Exception as e:
                    print("get ShowUrgencyMatrix Error:->", e)
                    urgency_matrix_error_msg = 'Tasks should have a Custom Field named "Urgency" with a dropdown value of either Urgent & Important, Less Urgent but Important, Urgent but less Important, or Neither Urgent nor Important.'
            else:
                urgency_matrix_error_msg = 'Tasks should have a Custom Field named "Urgency" with a dropdown value of either Urgent & Important, Less Urgent but Important, Urgent but less Important, or Neither Urgent nor Important.'

            response = {
                'success': True,
                'urgent_and_important': urgent_and_important,
                'less_urgent_but_important': less_urgent_but_important,
                'urgent_but_less_important': urgent_but_less_important,
                'neither_urgent_nor_important': neither_urgent_nor_important,
                'urgency_matrix_error_msg': urgency_matrix_error_msg,
            }
            return JsonResponse(response, safe=False)


class ProjectsDeleted(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):

        user = self.request.user
        common_group = get_object_or_404(Group, name='COMMON')
        admin_group = get_object_or_404(Group, name='DOC_ADMIN')

        if admin_group in user.groups.all():
            deleted_project_obj_list = Project.objects.filter(proj_deleted=True).order_by('-proj_id').distinct()
        else:
            deleted_project_obj_list = Project.objects.filter(Q(proj_deleted=True, proj_doc_group__in=user.groups.all()) |
                                               Q(proj_deleted=True, proj_doc_group__in=[common_group]) |
                                               Q(proj_deleted=True, proj_owner=user)).order_by('-proj_id').distinct()

        time_zone = get_object_or_404(Profile, user=self.request.user).client_tz

        context = {
            'time_zone': time_zone,
            'client_name': connection.tenant.name,
            'deleted_project_obj_list':deleted_project_obj_list,
            'user': user,
        }
        return render(request, 'project_management/projects_deleted.html', context)


class RestoreDeletedProject(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            proj_id = request.POST.get('proj_id')

            try:
                project = get_object_or_404(Project, proj_id=proj_id)
                project.proj_deleted = False

                # call create project log history func
                my_project_name = str(project.proj_name)
                my_project_id = str(project.proj_id)
                project_log_status = 'PROJECT_REST'
                project_log_message = 'Project ID: ' + my_project_id + ', Project Name: ' + my_project_name + ' restored successfully.'
                createProjectLogHistory(request, my_project_id, project_log_status, project_log_message)

                project.save()
            except:
                pass

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class MoveToArchive(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            proj_id = request.POST.get('archive_proj_id')

            try:
                project = get_object_or_404(Project, proj_id=proj_id)
                project.proj_archive = True

                # call create project log history func
                my_project_name = str(project.proj_name)
                my_project_id = str(project.proj_id)
                project_log_status = 'PROJECT_ARCH'
                project_log_message = 'Project ID: ' + my_project_id + ', Project Name: ' + my_project_name + ' archived successfully.'
                createProjectLogHistory(request, my_project_id, project_log_status, project_log_message)

                project.save()
            except Exception as e:
                print("post MoveToArchive Error:->", e)

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class ProjectsArchive(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):

        user = self.request.user
        common_group = get_object_or_404(Group, name='COMMON')
        admin_group = get_object_or_404(Group, name='DOC_ADMIN')

        if admin_group in user.groups.all():
            archive_project_obj_list = Project.objects.filter(proj_archive=True).order_by('-proj_id').distinct()
        else:
            archive_project_obj_list = Project.objects.filter(Q(proj_archive=True, proj_doc_group__in=user.groups.all()) |
                                               Q(proj_archive=True, proj_doc_group__in=[common_group]) |
                                               Q(proj_archive=True, proj_owner=user)).order_by('-proj_id').distinct()

        time_zone = get_object_or_404(Profile, user=self.request.user).client_tz

        context = {
            'time_zone': time_zone,
            'client_name': connection.tenant.name,
            'archive_project_obj_list':archive_project_obj_list,
            'user': user,
        }
        return render(request, 'project_management/projects_archive.html', context)


class RestoreArchiveProject(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            proj_id = request.POST.get('proj_id')

            try:
                project = get_object_or_404(Project, proj_id=proj_id)
                project.proj_archive = False

                # call create project log history func
                my_project_name = str(project.proj_name)
                my_project_id = str(project.proj_id)
                project_log_status = 'PROJECT_REST'
                project_log_message = 'Project ID: ' + my_project_id + ', Project Name: ' + my_project_name + ' restored successfully.'
                createProjectLogHistory(request, my_project_id, project_log_status, project_log_message)

                project.save()
            except:
                pass

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class PermanentlyDeleteProject(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            proj_id = request.POST.get('proj_id')

            try:
                project = get_object_or_404(Project, proj_id=proj_id)
                tasks = Task.objects.filter(project=project, task_deleted=False)
                for task in tasks:
                    task.delete()

                # call create project log history func
                my_project_name = str(project.proj_name)
                my_project_id = str(project.proj_id)
                project_log_status = 'PROJECT_DEL'
                project_log_message = 'Project ID: ' + my_project_id + ', Project Name: ' + my_project_name + ' deleted permanently.'
                createProjectLogHistory(request, my_project_id, project_log_status, project_log_message)

                project.delete()
            except:
                pass

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class MyTasks(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):

        do_today_task_list = []
        do_later_task_list = []
        today_date = date.today()
        end_date = today_date + timedelta(days=7)

        do_today_tasks = Task.objects.filter(end_date__lte=end_date, task_deleted=False).exclude(status__task_status='Completed').order_by('-id')
        for task in do_today_tasks:
            assigned_to = task.assigned_to.all()
            user = request.user
            username = request.user.username
            try:
                task_status = task.status.task_status
            except:
                task_status = ""
            if user in assigned_to and task_status and task not in do_today_task_list:
                do_today_task_list.append(task)
            try:
                task_extra_column_list = task.task_extra_column
                if task_extra_column_list:
                    extra_column_dict = task_extra_column_list["column_details"]
                    try:
                        accountable_column_value = extra_column_dict["accountable"]
                        if accountable_column_value:
                            accountable_list = accountable_column_value.split(",")
                            accountable_column_value_list = [value.strip() for value in accountable_list if
                                                             value]
                            if username in accountable_column_value_list and task not in do_today_task_list:
                                do_today_task_list.append(task)
                    except Exception as e:
                        print("accountable err:->", e)
                    try:
                        consulted_column_value = extra_column_dict["consulted"]
                        if consulted_column_value:
                            consulted_list = consulted_column_value.split(",")
                            consulted_column_value_list = [value.strip() for value in consulted_list if value]
                            if username in consulted_column_value_list and task not in do_today_task_list:
                                do_today_task_list.append(task)
                    except Exception as e:
                        print("consulted err:->", e)
                    try:
                        informed_column_value = extra_column_dict["informed"]
                        if informed_column_value:
                            informed_list = informed_column_value.split(",")
                            informed_column_value_list = [value.strip() for value in informed_list if value]
                            if username in informed_column_value_list and task not in do_today_task_list:
                                do_today_task_list.append(task)
                    except Exception as e:
                        print("informed err:->", e)
            except:
                pass

        do_later_tasks = Task.objects.filter(end_date__gt=end_date, task_deleted=False).exclude(status__task_status='Completed').order_by('-id')
        for task in do_later_tasks:
            assigned_to = task.assigned_to.all()
            user = request.user
            username = request.user.username
            try:
                task_status = task.status.task_status
            except:
                task_status = ""
            if user in assigned_to and task_status and task not in do_later_task_list:
                do_later_task_list.append(task)
            try:
                task_extra_column_list = task.task_extra_column
                if task_extra_column_list:
                    extra_column_dict = task_extra_column_list["column_details"]
                    try:
                        accountable_column_value = extra_column_dict["accountable"]
                        if accountable_column_value:
                            accountable_list = accountable_column_value.split(",")
                            accountable_column_value_list = [value.strip() for value in accountable_list if
                                                             value]
                            if username in accountable_column_value_list and task not in do_later_task_list:
                                do_later_task_list.append(task)
                    except Exception as e:
                        print("accountable err:->", e)
                    try:
                        consulted_column_value = extra_column_dict["consulted"]
                        if consulted_column_value:
                            consulted_list = consulted_column_value.split(",")
                            consulted_column_value_list = [value.strip() for value in consulted_list if value]
                            if username in consulted_column_value_list and task not in do_later_task_list:
                                do_later_task_list.append(task)
                    except Exception as e:
                        print("consulted err:->", e)
                    try:
                        informed_column_value = extra_column_dict["informed"]
                        if informed_column_value:
                            informed_list = informed_column_value.split(",")
                            informed_column_value_list = [value.strip() for value in informed_list if value]
                            if username in informed_column_value_list and task not in do_later_task_list:
                                do_later_task_list.append(task)
                    except Exception as e:
                        print("informed err:->", e)
            except:
                pass

        time_zone = get_object_or_404(Profile, user=self.request.user).client_tz

        context = {
            'time_zone': time_zone,
            'client_name': connection.tenant.name,
            'do_today_task_list':do_today_task_list,
            'do_later_task_list':do_later_task_list,
        }
        return render(request, 'project_management/my_tasks.html', context)


class ImportCSVFile(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):
        proj_groups = Group.objects.all()

        context = {
            'proj_groups': proj_groups,
            'client_name': connection.tenant.name,
        }
        return render(request, 'project_management/import_csv_file.html', context)

    def post(self, request, *args, **kwargs):

        project_name = request.POST['project_name']
        project_description = request.POST['project_description']
        project_groups = request.POST.getlist('import_csv_file_group_select')

        request.session['project_name'] = project_name
        request.session['project_description'] = project_description
        request.session['project_groups'] = project_groups

        return redirect('upload-csv-file')


ALLOWED_EXTENSIONS = ['.csv']

class UploadCSVFile(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):

        context = {
            'client_name': connection.tenant.name,
        }
        return render(request, 'project_management/upload_csv_file.html', context)

    def post(self, request, *args, **kwargs):

        if request.method == 'POST':

            proj_file = request.FILES.get('files')
            proj_name = request.session['project_name']

            myfile = proj_file
            myfile_name = myfile.name
            f_name, f_ext = os.path.splitext(myfile_name)
            path = MEDIA_ROOT + '/' + connection.tenant.name + '/project_csv_files'
            if not os.path.isdir(path):
                os.mkdir(path)

            if f_ext not in ALLOWED_EXTENSIONS:
                messages.error(request, 'Oops! File does not have an approved extension. Please upload ".csv" file.')
                return redirect('upload-csv-file')
            else:
                path = connection.tenant.name + '/project_csv_files/' + myfile_name
                count = 0
                while True:
                    if count > 0:
                        new_file_name = myfile_name.replace(f_ext, '') + '(' + str(count) + ')' + f_ext
                    else:
                        new_file_name = myfile_name

                    path = connection.tenant.name + '/project_csv_files/' + new_file_name
                    check_file = default_storage.exists(path)
                    if not check_file:
                        myfile_name = new_file_name
                        break
                    count += 1

                file_name = default_storage.save(path, myfile)
                new_path = file_name

                rows = []
                with open(MEDIA_ROOT + '/' + path, 'r') as file:
                    csvreader = csv.reader(file)
                    header = next(csvreader)
                    for row in csvreader:
                        rows.append(row)
                    file.close()

                new_header = ["_".join(h.lower().split(" ")) for h in header]
                field_list = [item.field_title for item in ProjectField.objects.all()]

                result = []
                for row in rows:
                    col_val_dict = dict(zip(new_header, row))
                    result.append(col_val_dict)

                field_not_exists_list = []
                for h in new_header:
                    if h in field_list:
                        pass
                    else:
                        field_not_exists_list.append(h)
                if len(field_not_exists_list) != 0:
                    messages.error(request, 'Either column names '+str(field_not_exists_list)+ ' not matching or not exists in to the system.')
                    return redirect('upload-csv-file')
                else:
                    context = {
                        'proj_name': proj_name,
                        'new_header':new_header,
                        'data':rows,
                        'path':path,
                    }
                    return render(request, 'project_management/view_upload_csv_file.html', context)


class SaveUploadCSVFile(LoginRequiredMixin, View):

    def post(self, request, *args, **kwargs):
        proj_name = request.session['project_name']
        proj_description = request.session['project_description']
        proj_groups = request.session['project_groups']
        proj_owner = request.user
        path = request.POST.get('path')

        if 'cancel_upload' in request.POST:
            default_storage.delete(path)
            return redirect('create-new-project')
        if 'save_upload' in request.POST:
            rows = []
            with open(MEDIA_ROOT + '/' + path, 'r') as file:
                csvreader = csv.reader(file)
                header = next(csvreader)
                for row in csvreader:
                    rows.append(row)
                file.close()

            new_header = ["_".join(h.lower().split(" ")) for h in header]

            result = []
            for row in rows:
                col_val_dict = dict(zip(new_header, row))
                result.append(col_val_dict)

            project = Project.objects.create(proj_name=proj_name,
                                             proj_description=proj_description,
                                             proj_owner=proj_owner,
                                             proj_start_date=datetime.datetime.now(),
                                             )
            groups = proj_groups
            for group in groups:
                gp = get_object_or_404(Group, name=group)
                project.proj_doc_group.add(gp)
            project.save()

            try:
                # call create project log history func
                my_project_name = str(project.proj_name)
                my_project_id = str(project.proj_id)
                project_log_status = 'PROJECT_ADD'
                project_log_message = 'Project ID: ' + my_project_id + ', Project Name: ' + my_project_name + ' added successfully.'
                createProjectLogHistory(request, my_project_id, project_log_status, project_log_message)
            except:
                pass

            # get or create Project Status
            project_status_choice_list = ['Not Started', 'In Progress', 'Completed', 'On Hold']
            for p_status in project_status_choice_list:
                get_project_status, create_project_status = ProjectStatus.objects.get_or_create(proj_status=p_status)

            # save project status to Not Started
            get_proj_status_obj = ProjectStatus.objects.get(proj_status="Not Started")
            project.proj_status = get_proj_status_obj
            project.save()

            _now = datetime.datetime.now()
            proj_temp_id = _now.strftime('%Y%m%d%H%M%S%f')
            proj_template = ProjectTemplate.objects.create(proj_temp_id=proj_temp_id,
                                                           proj_temp_status='ac',
                                                           proj_temp_title=proj_name + ' template ' + str(proj_temp_id),
                                                           proj_temp_description=proj_name + ' template ' + str(proj_temp_id),
                                                           proj_temp_created_at=_now,
                                                           proj_temp_owner=request.user,
                                                           proj_temp_edited_by=request.user,
                                                           )
            cg = get_object_or_404(Group, name='COMMON')
            proj_template.proj_temp_doc_group.add(cg)
            proj_template.save()

            # save project template
            project.proj_origin = proj_template
            project.save()

            # get or create field choices
            field_choice_list = ['Not Started', 'In Progress', 'Completed', 'On Hold', 'Low', 'Medium', 'High']
            for f_choice in field_choice_list:
                get_field_choice, create_field_choice = FieldChoice.objects.get_or_create(f_choice=f_choice)

            field_origin = proj_template

            status_list = ['Not Started', 'In Progress', 'Completed', 'On Hold', 'Delayed']
            priority_list = ['Low', 'Medium', 'High']

            # get or create project field
            project_field_list = new_header

            try:
                for i, item in enumerate(project_field_list):
                    proj_field = ProjectField.objects.get(field_title=item)
                    field_position = str(i + 1)
                    field_title = proj_field.field_title
                    field_description = proj_field.field_description
                    field_type = proj_field.field_type
                    field_origin = field_origin
                    if field_title == 'status':
                        field_choice = status_list
                    elif field_title == 'priority':
                        field_choice = priority_list
                    else:
                        field_choice = [f.f_choice for f in proj_field.field_choice.all()]

                    try:
                        get_project_field, create_project_field = ProjectField.objects.get_or_create(
                            field_title=field_title,
                            field_description=field_description,
                            field_type=field_type,
                            field_origin=field_origin)
                    except:
                        get_project_field = get_object_or_404(ProjectField, field_title=field_title)

                    if field_choice:
                        for f_choice in field_choice:
                            try:
                                fc = FieldChoice.objects.create(f_choice=f_choice)
                            except IntegrityError:
                                fc = get_object_or_404(FieldChoice, f_choice=f_choice)
                            get_project_field.field_choice.add(fc)

                    ProjectTemplateField.objects.create(proj_temp=proj_template,
                                                        proj_temp_field=get_project_field,
                                                        proj_temp_field_position=field_position)

            except Exception as e:
                print("post UploadCSVFile error:->", e)

            # get project id
            proj_id = project.proj_id

            project_template = project.proj_origin
            project_template_field_list = ProjectTemplateField.objects.filter(proj_temp=project_template)

            for j, res_dict in enumerate(result):
                sort_order = str(j+1)
                task = Task.objects.create(parent='0',
                                            project=project,
                                            created_by=request.user,
                                            sort_order=sort_order)

                try:
                    # call create task log history func
                    task_id = task.id
                    task_name = task.text
                    project_id = task.project.proj_id
                    project_name = task.project.proj_name
                    task_log_status = 'TASK_ADD'
                    task_log_message = 'Task ID: '+str(task_id)+ ', Task Name: '+str(task_name)+' added to Project ID: '+str(project_id)+', Project Name: '+str(project_name)+' successfully.'
                    createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)
                except:
                    pass

                title = res_dict["task_name"]
                task.text = title

                _now = datetime.datetime.now()
                try:
                    start_date = res_dict["start_date"]
                    if start_date:
                        fuzzy_start_date = parse(start_date, fuzzy=True)
                        fuzzy_start_date_str = str(fuzzy_start_date)[:10]
                        start_date_val = datetime.datetime.strptime(fuzzy_start_date_str, '%Y-%m-%d')
                        start_date_obj = start_date_val.strftime('%Y-%m-%d')
                        task.start_date = start_date_obj
                    else:
                        default_start_date = _now.strftime('%Y-%m-%d')
                        task.start_date = default_start_date
                except Exception as e:
                    print("post UploadCSVFile start date error:->", e)
                    pass

                try:
                    end_date = res_dict["due_date"]
                    fuzzy_end_date = parse(end_date, fuzzy=True)
                    fuzzy_end_date_str = str(fuzzy_end_date)[:10]
                    end_date_val = datetime.datetime.strptime(fuzzy_end_date_str, '%Y-%m-%d')
                    end_date_obj = end_date_val.strftime('%Y-%m-%d')
                    task.end_date = end_date_obj
                except Exception as e:
                    print("post UploadCSVFile end date error2:->", e)
                    pass

                try:
                    assignee = res_dict["assignee"]
                    assignee_list = assignee.split(",")
                    task.assigned_to.clear()
                    assignee_list = [item.strip() for item in assignee_list]
                    for usr in assignee_list:
                        usr = usr.strip()
                        user = get_object_or_404(User, username=usr)
                        task.assigned_to.add(user)
                except Exception as e:
                    print("post UploadCSVFile assigned to error:->", e)
                    pass

                try:
                    task_status = res_dict["status"]
                    if task_status != "":
                        get_status, create_status = TaskStatus.objects.get_or_create(task_status=task_status)
                        task.status = get_status
                    else:
                        pass
                except Exception as e:
                    print("post UploadCSVFile status error:->", e)
                    pass

                try:
                    task_priority = res_dict["priority"]
                    if task_priority != "":
                        get_priority, create_priority = TaskPriority.objects.get_or_create(
                            task_priority=task_priority)
                        task.priority = get_priority
                    else:
                        pass
                except Exception as e:
                    print("post UploadCSVFile priority error:->", e)
                    pass

                task.save()

                db_column_list = ['task_name', 'assignee', 'start_date', 'due_date', 'status', 'priority']

                column_details_dict = {}
                for item in project_template_field_list:
                    column_details = {}
                    field_title = item.proj_temp_field.field_title
                    field_type = item.proj_temp_field.field_type
                    if field_title in db_column_list:
                        pass
                    else:
                        try:
                            field_value = res_dict[field_title]
                            if field_type == 'single-select' and field_value:
                                get_field_obj = ProjectField.objects.get(field_title=field_title)
                                try:
                                    get_choice = FieldChoice.objects.get(f_choice=field_value)
                                except Exception as e:
                                    print("post UploadCSVFile single select error:->", e)
                                    get_choice = FieldChoice.objects.create(f_choice=field_value)
                                get_field_obj.field_choice.add(get_choice)
                                get_field_obj.save()

                                column_details[field_title] = field_value
                                task.save()

                            elif field_type == 'multi-select':
                                get_field_obj = ProjectField.objects.get(field_title=field_title)
                                field_value_split_list = field_value.split(",")
                                field_value_list = [x for x in field_value_split_list if x]
                                if len(field_value_list) != 0:
                                    for f_choice in field_value_list:
                                        f_choice = f_choice.strip()
                                        try:
                                            get_choice = FieldChoice.objects.get(f_choice=f_choice)
                                        except Exception as e:
                                            print("post UploadCSVFile multi select error:->", e)
                                            get_choice = FieldChoice.objects.create(f_choice=f_choice)
                                        get_field_obj.field_choice.add(get_choice)
                                        get_field_obj.save()

                                column_details[field_title] = field_value
                                task.save()
                            elif field_type == 'people':
                                column_details[field_title] = field_value
                                task.save()
                            elif field_type == 'date':
                                field_value = datetime.datetime.strptime(field_value, '%d-%m-%Y')
                                field_value = field_value.strftime('%d %b, %Y')
                                column_details[field_title] = field_value
                                task.save()
                            else:
                                if field_value == None:
                                    field_value = ""
                                column_details[field_title] = field_value
                                task.save()
                        except:
                            pass
                    column_details_dict.update(column_details)

                column_details_dict = {'column_details': column_details_dict}
                task.task_extra_column = column_details_dict
                task.save()

            return redirect('tree-grid-view', proj_id)


class CreateProjectField(LoginRequiredMixin, View):

    def get(self, request):

        field_types = []
        for field in FIELD_TYPE:
            field_types.append(field[0])
        field_choices = FieldChoice.objects.all()
        project_field_list = [item.field_title for item in ProjectField.objects.all()]
        field_list = ['task_name', 'assignee', 'start_date', 'due_date', 'status', 'priority', 'urgency']
        project_fields = []
        for field in project_field_list:
            if str(field).lower() in field_list:
                pass
            else:
                project_fields.append(field)

        context = {
            'field_choices': field_choices,
            'field_types': field_types,
            'project_fields': project_fields,
            'client_name': connection.tenant.name,
        }

        return render(request, 'project_management/create_project_field.html', context)

    def post(self, request):
        field_title = request.POST.get('proj_field_title')
        proj_field_description = request.POST.get('proj_field_description')
        proj_field_type = request.POST.get('proj_field_select')
        proj_field_choices = request.POST.getlist('proj_field_select_choices')
        proj_field_title="_".join(field_title.split(" ")).lower()
        try:
            get_project_field, create_project_field = ProjectField.objects.get_or_create(
                field_title=proj_field_title,
                field_description=proj_field_description,
                field_type=proj_field_type,
            )
        except:
            get_project_field = get_object_or_404(ProjectField, field_title=proj_field_title)

        if proj_field_choices:
            for f_choice in proj_field_choices:
                try:
                    fc = FieldChoice.objects.create(f_choice=f_choice)
                except IntegrityError:
                    fc = get_object_or_404(FieldChoice, f_choice=f_choice)
                get_project_field.field_choice.add(fc)
                get_project_field.save()

        messages.success(request, 'Field name "'+field_title+'" created successfully!')
        return redirect('create-project-field')


class GetFieldChoiceValues(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():
            field_choice_value_exist = "False"
            entered_field_choice_values_lower = ""
            try:
                db_field_choice_values = [str(item.f_choice).lower() for item in FieldChoice.objects.all()]
                entered_field_choice_values = request.GET.get('field_choice_values')
                entered_field_choice_values_lower += entered_field_choice_values.lower()
                if entered_field_choice_values_lower in db_field_choice_values:
                    field_choice_value_exist = "True"

            except:
                field_choice_value_exist = "False"

            response = {
                'success': True,
                'field_choice_value_exist': field_choice_value_exist,
                'entered_field_choice_values_lower': entered_field_choice_values_lower,
            }
            return JsonResponse(response, safe=False)


class PostFieldChoiceValues(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            field_choice_value = request.POST.get('input_field_choice_value')

            try:
                f_choice = FieldChoice.objects.create(f_choice=field_choice_value)
            except Exception as e:
                print("PostFieldChoiceValues Error:->", e)

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class GetProjectTaskLogHistory(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():
            # project task log history
            task_history_details = {}
            task_logs = {}
            req_data = request.GET.get('proj_id')
            proj_id = req_data
            task_log_history_obj = TaskLogHistory.objects.filter(project_id = proj_id)
            task_log_id = [item.task_log_id for item in task_log_history_obj]
            task_id = [item.task_id for item in task_log_history_obj]
            task_log_message = [item.task_log_message for item in task_log_history_obj]
            task_log_created_by_list = [item.task_log_created_by.username for item in task_log_history_obj ]
            task_log_created_at_list = [item.task_log_created_at for item in task_log_history_obj]
            task_log_status_list = [item.get_task_log_status_display() for item in task_log_history_obj]

            for i in range(0,len(task_log_created_by_list)):
                task_log = {
                    'Log_ID':task_log_id[i],
                    'Task_ID':task_id[i],
                    'Log_Message':task_log_message[i],
                    'Created_By':task_log_created_by_list[i],
                    'Created_At':task_log_created_at_list[i],
                    'Task_Log_Status':task_log_status_list[i],
                }
                task_logs["task_log"+str(i+1)] = task_log
            task_history_details["task_logs"] = task_logs

            response = {
                'success': True,
                'task_history_details': task_history_details,
            }
            return JsonResponse(response, safe=False)


class Export_to_Excel(LoginRequiredMixin, View):

    def get(self, request):

        local_timezone = tzlocal.get_localzone()
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Document Log History.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Document Log History"

        # Add headers
        headers = ["Log ID", "Username", "Date and Time", "Log Message", "Log Status", "Object ID"]
        ws.append(headers)

        # Save the workbook to the HttpResponse
        wb.save(response)
        return response


class ReportingDashboard(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):

        user = self.request.user
        common_group = get_object_or_404(Group, name='COMMON')
        admin_group = get_object_or_404(Group, name='DOC_ADMIN')

        try:
            completed_task_status = TaskStatus.objects.get(task_status='Completed')
            if admin_group in user.groups.all():
                completed_tasks_count = Task.objects.filter(Q(project__proj_deleted=False, project__proj_archive=False,
                                                              task_deleted=False, status=completed_task_status)).distinct().count()
            else:
                completed_tasks_count = Task.objects.filter(Q(project__proj_deleted=False, project__proj_archive=False,
                                                              project__proj_doc_group__in=user.groups.all(),
                                                              task_deleted=False, status=completed_task_status) |
                                                            Q(project__proj_deleted=False, project__proj_archive=False,
                                                              project__proj_doc_group__in=[common_group],
                                                              task_deleted=False, status=completed_task_status) |
                                                            Q(project__proj_deleted=False, project__proj_archive=False, project__proj_owner=user,
                                                              task_deleted=False,
                                                              status=completed_task_status)).distinct().count()
        except Exception as e:
            completed_tasks_count = 0

        try:
            completed_task_status = TaskStatus.objects.get(task_status='Completed')
            if admin_group in user.groups.all():
                incomplete_tasks_count = Task.objects.filter(project__proj_deleted=False, project__proj_archive=False, task_deleted=False).exclude(status=completed_task_status).distinct().count()
            else:
                incomplete_tasks_count = Task.objects.filter(Q(project__proj_deleted=False, project__proj_archive=False,
                                                              project__proj_doc_group__in=user.groups.all(), task_deleted=False) |
                                                            Q(project__proj_deleted=False, project__proj_archive=False,
                                                              project__proj_doc_group__in=[common_group], task_deleted=False) |
                                                            Q(project__proj_deleted=False, project__proj_archive=False, project__proj_owner=user,
                                                              task_deleted=False)).exclude(status=completed_task_status).distinct().count()
        except Exception as e:
            incomplete_tasks_count = 0

        try:
            completed_task_status = TaskStatus.objects.get(task_status='Completed')
            today_date = date.today()
            if admin_group in user.groups.all():
                overdue_tasks_count = Task.objects.filter(project__proj_deleted=False, project__proj_archive=False,
                                                          end_date__lt=today_date, task_deleted=False).exclude(status=completed_task_status).distinct().count()
            else:
                overdue_tasks_count = Task.objects.filter(Q(project__proj_deleted=False, project__proj_archive=False,
                                                              project__proj_doc_group__in=user.groups.all(),
                                                              end_date__lt=today_date, task_deleted=False) |
                                                            Q(project__proj_deleted=False, project__proj_archive=False,
                                                              project__proj_doc_group__in=[common_group],
                                                              end_date__lt=today_date, task_deleted=False) |
                                                            Q(project__proj_deleted=False, project__proj_archive=False, project__proj_owner=user,
                                                              end_date__lt=today_date, task_deleted=False)).exclude(status=completed_task_status).distinct().count()
        except:
            overdue_tasks_count = 0

        try:
            today_date = date.today()
            completed_task_status = TaskStatus.objects.get(task_status='Completed')
            if admin_group in user.groups.all():
                upcoming_tasks_count = Task.objects.filter(project__proj_deleted=False, project__proj_archive=False,
                                                           start_date__gte=today_date, task_deleted=False).exclude(status=completed_task_status).distinct().count()
            else:
                upcoming_tasks_count = Task.objects.filter(Q(project__proj_deleted=False, project__proj_archive=False,
                                                            project__proj_doc_group__in=user.groups.all(),
                                                            start_date__gte=today_date, task_deleted=False) |
                                                          Q(project__proj_deleted=False, project__proj_archive=False,
                                                            project__proj_doc_group__in=[common_group],
                                                            start_date__gte=today_date, task_deleted=False) |
                                                          Q(project__proj_deleted=False, project__proj_archive=False, project__proj_owner=user,
                                                            start_date__gte=today_date, task_deleted=False)).exclude(status=completed_task_status).distinct().count()
        except:
            upcoming_tasks_count = 0

        try:
            if admin_group in user.groups.all():
                total_tasks_count = Task.objects.filter(project__proj_deleted=False, project__proj_archive=False, task_deleted=False).distinct().count()
            else:
                total_tasks_count = Task.objects.filter(Q(project__proj_deleted=False, project__proj_archive=False,
                                                               project__proj_doc_group__in=user.groups.all(),
                                                               task_deleted=False) |
                                                             Q(project__proj_deleted=False, project__proj_archive=False,
                                                               project__proj_doc_group__in=[common_group], task_deleted=False) |
                                                             Q(project__proj_deleted=False, project__proj_archive=False, project__proj_owner=user,
                                                               task_deleted=False)).distinct().count()
        except:
            total_tasks_count = 0

        # add code related to highcharts
        project_labels, overdue_tasks_data, incomplete_tasks_data, upcoming_tasks_data, completed_tasks_data = [], [], [], [], []

        total_projects = Project.objects.filter(proj_deleted=False, proj_archive=False)
        for proj_obj in total_projects:
            project_name = proj_obj.proj_name
            # add code related to overdue task highcharts
            try:
                completed_task_status = TaskStatus.objects.get(task_status='Completed')
                today_date = date.today()
                if admin_group in user.groups.all():
                    proj_overdue_tasks_count = Task.objects.filter(project=proj_obj, end_date__lt=today_date, task_deleted=False).exclude(status=completed_task_status).distinct().count()
                else:
                    proj_overdue_tasks_count = Task.objects.filter(Q(project=proj_obj,
                                                                  project__proj_doc_group__in=user.groups.all(),
                                                                  end_date__lt=today_date, task_deleted=False) |
                                                                Q(project=proj_obj,
                                                                  project__proj_doc_group__in=[common_group],
                                                                  end_date__lt=today_date, task_deleted=False) |
                                                                Q(project=proj_obj, project__proj_owner=user,
                                                                  end_date__lt=today_date, task_deleted=False)).exclude(status=completed_task_status).distinct().count()
            except:
                proj_overdue_tasks_count = 0
            overdue_tasks_data.append(proj_overdue_tasks_count)

            # add code related to incomplete task highcharts
            try:
                completed_task_status = TaskStatus.objects.get(task_status='Completed')
                if admin_group in user.groups.all():
                    proj_incomplete_tasks_count = Task.objects.filter(project=proj_obj, task_deleted=False).exclude(
                        status=completed_task_status).distinct().count()
                else:
                    proj_incomplete_tasks_count = Task.objects.filter(
                        Q(project=proj_obj,
                          project__proj_doc_group__in=user.groups.all(), task_deleted=False) |
                        Q(project=proj_obj,
                          project__proj_doc_group__in=[common_group], task_deleted=False) |
                        Q(project=proj_obj, project__proj_owner=user,
                          task_deleted=False)).exclude(status=completed_task_status).distinct().count()
            except Exception as e:
                proj_incomplete_tasks_count = 0
            incomplete_tasks_data.append(proj_incomplete_tasks_count)

            # add code related to upcoming task highcharts
            try:
                today_date = date.today()
                completed_task_status = TaskStatus.objects.get(task_status='Completed')
                if admin_group in user.groups.all():
                    proj_upcoming_tasks_count = Task.objects.filter(project=proj_obj,
                                                               start_date__gte=today_date,
                                                               task_deleted=False).exclude(status=completed_task_status).distinct().count()
                else:
                    proj_upcoming_tasks_count = Task.objects.filter(
                        Q(project=proj_obj,
                          project__proj_doc_group__in=user.groups.all(),
                          start_date__gte=today_date, task_deleted=False) |
                        Q(project=proj_obj,
                          project__proj_doc_group__in=[common_group],
                          start_date__gte=today_date, task_deleted=False) |
                        Q(project=proj_obj, project__proj_owner=user,
                          start_date__gte=today_date, task_deleted=False)).exclude(status=completed_task_status).distinct().count()
            except:
                proj_upcoming_tasks_count = 0
            upcoming_tasks_data.append(proj_upcoming_tasks_count)

            # add code related to completed task highcharts
            try:
                completed_task_status = TaskStatus.objects.get(task_status='Completed')
                if admin_group in user.groups.all():
                    proj_completed_tasks_count = Task.objects.filter(
                        Q(project=proj_obj,
                          task_deleted=False, status=completed_task_status)).distinct().count()
                else:
                    proj_completed_tasks_count = Task.objects.filter(
                        Q(project=proj_obj,
                          project__proj_doc_group__in=user.groups.all(),
                          task_deleted=False, status=completed_task_status) |
                        Q(project=proj_obj,
                          project__proj_doc_group__in=[common_group],
                          task_deleted=False, status=completed_task_status) |
                        Q(project=proj_obj, project__proj_owner=user,
                          task_deleted=False,
                          status=completed_task_status)).distinct().count()
            except Exception as e:
                proj_completed_tasks_count = 0
            completed_tasks_data.append(proj_completed_tasks_count)

            project_labels.append(project_name)

        time_zone = get_object_or_404(Profile, user=self.request.user).client_tz

        context = {
            "time_zone": time_zone,
            'client_name': connection.tenant.name,
            "completed_tasks_count": completed_tasks_count,
            "incomplete_tasks_count": incomplete_tasks_count,
            "overdue_tasks_count": overdue_tasks_count,
            "upcoming_tasks_count": upcoming_tasks_count,
            "total_tasks_count": total_tasks_count,
            "project_labels": project_labels,
            "overdue_tasks_data": overdue_tasks_data,
            "incomplete_tasks_data": incomplete_tasks_data,
            "upcoming_tasks_data": upcoming_tasks_data,
            "completed_tasks_data": completed_tasks_data,
        }
        return render(request, 'project_management/reporting_dashboard.html', context)


class MoreCharts(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):

        user = self.request.user
        common_group = get_object_or_404(Group, name='COMMON')
        admin_group = get_object_or_404(Group, name='DOC_ADMIN')

        # add code related to highcharts
        project_labels, total_tasks_data = [], []

        total_projects = Project.objects.filter(proj_deleted=False, proj_archive=False)
        for proj_obj in total_projects:
            project_name = proj_obj.proj_name
            # add code related to total task highcharts
            try:
                if admin_group in user.groups.all():
                    proj_total_tasks_count = Task.objects.filter(project=proj_obj,
                                                            task_deleted=False).distinct().count()
                else:
                    proj_total_tasks_count = Task.objects.filter(Q(project=proj_obj,
                                                              project__proj_doc_group__in=user.groups.all(),
                                                              task_deleted=False) |
                                                            Q(project=proj_obj,
                                                              project__proj_doc_group__in=[common_group],
                                                              task_deleted=False) |
                                                            Q(project=proj_obj,
                                                              project__proj_owner=user,
                                                              task_deleted=False)).distinct().count()
            except:
                proj_total_tasks_count = 0
            total_tasks_data.append(proj_total_tasks_count)

            project_labels.append(project_name)

        # add code related to project status highcharts
        project_status_data_list, project_status_labels = [], []

        if admin_group in user.groups.all():
            projects = Project.objects.filter(proj_deleted=False, proj_archive=False).distinct()
        else:
            projects = Project.objects.filter(Q(proj_deleted=False, proj_archive=False,
                                               proj_doc_group__in=user.groups.all()) |
                                             Q(proj_deleted=False, proj_archive=False,
                                               proj_doc_group__in=[common_group]) |
                                             Q(proj_deleted=False, proj_archive=False,
                                               proj_owner=user)).distinct()
        for item in projects:
            try:
                proj_status = item.proj_status.proj_status
            except:
                proj_status = "None"
            project_status_labels.append(proj_status)
        project_status_data = dict(Counter(project_status_labels))

        # Convert Key-Value Dictionary to Lists of List
        project_status_data_type = [[key, val] for key, val in project_status_data.items()]
        project_status_data_list.append(project_status_data_type)
        project_data_list = [item for sublist in project_status_data_list for item in sublist]

        # add code related to assignee highcharts
        if admin_group in user.groups.all():
            total_tasks = Task.objects.filter(Q(project__proj_deleted=False, project__proj_archive=False,
                                                          task_deleted=False)).distinct()
        else:
            total_tasks = Task.objects.filter(Q(project__proj_deleted=False, project__proj_archive=False,
                                                          project__proj_doc_group__in=user.groups.all(),
                                                          task_deleted=False) |
                                                        Q(project__proj_deleted=False, project__proj_archive=False,
                                                          project__proj_doc_group__in=[common_group],
                                                          task_deleted=False) |
                                                        Q(project__proj_deleted=False, project__proj_archive=False,
                                                          project__proj_owner=user,
                                                          task_deleted=False)).distinct()

        task_assignee_data_list, task_assignee_labels = [], []
        for task in total_tasks:
            assignee_nested_list = []
            try:
                for user in task.assigned_to.all():
                    username = user.username
                    assignee_nested_list.append(username)
                task_assignee_labels.append(assignee_nested_list)
            except:
                pass
        flatList = [element for innerList in task_assignee_labels for element in innerList]
        task_assignee_data = dict(Counter(flatList))

        task_assignee_labels_list, task_assignee_data_list = [], []
        for key, val in task_assignee_data.items():
            task_assignee_labels_list.append(key)
            task_assignee_data_list.append(val)

        # add code related to upcoming tasks by assignee this week highcharts
        today = date.today()
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        this_week_task_assignee_data_list, this_week_task_assignee_labels = [], []
        for task in total_tasks:
            this_week_assignee_nested_list = []
            try:
                task_status = task.status.task_status
            except:
                task_status = ''
            try:
                task_end_date = task.start_date.date()
                if task_end_date >= today and task_end_date <= end and task_status != 'Completed':
                    for user in task.assigned_to.all():
                        username = user.username
                        this_week_assignee_nested_list.append(username)
                    this_week_task_assignee_labels.append(this_week_assignee_nested_list)
            except:
                pass
        this_week_flat_list = [element for innerList in this_week_task_assignee_labels for element in innerList]
        this_week_task_assignee_data = dict(Counter(this_week_flat_list))

        this_week_task_assignee_labels_list, this_week_task_assignee_data_list = [], []
        for key, val in this_week_task_assignee_data.items():
            this_week_task_assignee_labels_list.append(key)
            this_week_task_assignee_data_list.append(val)

        time_zone = get_object_or_404(Profile, user=self.request.user).client_tz

        context = {
            "time_zone": time_zone,
            'client_name': connection.tenant.name,
            "project_labels": project_labels,
            "total_tasks_data": total_tasks_data,
            "project_data_list": project_data_list,
            "task_assignee_labels_list": task_assignee_labels_list,
            "task_assignee_data_list": task_assignee_data_list,
            "this_week_task_assignee_labels_list":this_week_task_assignee_labels_list,
            "this_week_task_assignee_data_list":this_week_task_assignee_data_list,
        }
        return render(request, 'project_management/more_charts.html', context)


class GetEditableColumnRestriction(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():

            proj_id = request.GET.get('proj_id')
            project_obj = get_object_or_404(Project, proj_id=proj_id)
            editable_column_name_list = []
            proj_temp_obj = project_obj.proj_origin

            proj_temp_editable_column_obj = ProjectTemplateField.objects.filter(proj_temp=proj_temp_obj).extra(
                select={'int_from_position': 'CAST(proj_temp_field_position AS INTEGER)'}
            ).order_by('int_from_position', 'proj_temp_field_position')

            for item in proj_temp_editable_column_obj:
                # check if task name field is true then break loop and go to the next field
                check_task_name = " ".join(item.proj_temp_field.field_title.split("_")).lower()
                if check_task_name == "task name":
                    continue
                field_id = item.proj_temp_field.field_title
                field_title = " ".join(item.proj_temp_field.field_title.split("_")).capitalize()
                proj_template_id = item.proj_template_id
                editable_column_name_list.append(
                    {'id': field_id, 'text': field_title, 'proj_template_id': proj_template_id})

            response = {
                'success': True,
                'editable_column_name_list':editable_column_name_list,
            }
            return JsonResponse(response, safe=False)


class GetEditableColumnGroups(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():

            proj_template_id = request.GET.get('proj_template_id')
            user_groups = [gr.name for gr in Group.objects.all()]
            try:
                proj_temp_editable_group_obj = ProjectTemplateField.objects.get(proj_template_id=proj_template_id)
                column_groups = [group.name for group in proj_temp_editable_group_obj.proj_temp_field_editable.all()]
            except:
                column_groups = []

            response = {
                'success': True,
                'user_groups':user_groups,
                'column_groups': column_groups,
                'proj_template_id':proj_template_id,
            }
            return JsonResponse(response, safe=False)


class PostEditableColumnGroups(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            proj_temp_field_id = request.POST.get('get_proj_temp_field_id')
            selected_column_group_options = request.POST.get('selected_column_group_options')
            selected_column_groups = json.loads(selected_column_group_options)
            try:
                proj_temp_editable_group_obj = ProjectTemplateField.objects.get(proj_template_id=proj_temp_field_id)
                proj_temp_editable_group_obj.proj_temp_field_editable.clear()
                for group in selected_column_groups:
                    gp = get_object_or_404(Group, name=group)
                    proj_temp_editable_group_obj.proj_temp_field_editable.add(gp)
                proj_temp_editable_group_obj.save()
            except:
                pass

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class DeletedTasks(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):

        user = self.request.user
        common_group = get_object_or_404(Group, name='COMMON')
        admin_group = get_object_or_404(Group, name='DOC_ADMIN')

        try:
            if admin_group in user.groups.all():
                deleted_task_objs = Task.objects.filter(task_deleted=True, project__proj_deleted=False, project__proj_archive=False).order_by('-id').distinct()
            else:
                deleted_task_objs = Project.objects.filter(Q(task_deleted=True, project__proj_deleted=False, project__proj_archive=False, project__proj_doc_group__in=user.groups.all()) |
                                                   Q(task_deleted=True, project__proj_deleted=False, project__proj_archive=False, project__proj_doc_group__in=[common_group]) |
                                                   Q(task_deleted=True, project__proj_deleted=False, project__proj_archive=False, created_by=user)).order_by('-id').distinct()
        except:
            deleted_task_objs = []

        deleted_task_obj_list = []
        for task in deleted_task_objs:
            if task.parent != "0":
                parent_task_id = task.parent
                parent_task = Task.objects.get(pk=parent_task_id)
                if parent_task.task_deleted:
                    pass
                else:
                    deleted_task_obj_list.append(task)
            else:
                deleted_task_obj_list.append(task)

        time_zone = get_object_or_404(Profile, user=self.request.user).client_tz

        context = {
            'time_zone': time_zone,
            'client_name': connection.tenant.name,
            'deleted_task_obj_list':deleted_task_obj_list,
            'user': user,
        }
        return render(request, 'project_management/deleted_tasks.html', context)


class RestoreDeletedTask(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            task_id = request.POST.get('task_id')
            task = Task.objects.get(pk=task_id)

            if task.parent != "0":
                parent_task_id = task.parent
                parent_task = Task.objects.get(pk=parent_task_id)
                if parent_task.task_deleted:
                    response = {
                        'success': False,
                        'error': f'The parent task is deleted for this child task, so first restore the parent task ID:({parent_task_id})',
                    }
                    return JsonResponse(response, safe=False)

            result = [[task.id]]

            def path_to_root_task(task_id):
                task = Task.objects.filter(parent=task_id)
                child = []
                for item in task:
                    child.append(item.id)
                    path_to_root_task(item.id)
                if child:
                    result.append(child)

            path_to_root_task(task.id)
            child_task_list = [item for sublist in result for item in sublist]

            for id in child_task_list:
                task_obj = Task.objects.get(pk=id)
                task_obj.task_deleted = False
                task_obj.save()

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class DeleteTaskPermanently(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            task_id = request.POST.get('task_id')
            task = Task.objects.get(pk=task_id)
            result = [[task.id]]

            def path_to_root_task(task_id):
                task = Task.objects.filter(parent=task_id)
                child = []
                for item in task:
                    child.append(item.id)
                    path_to_root_task(item.id)
                if child:
                    result.append(child)

            path_to_root_task(task.id)
            child_task_list = [item for sublist in result for item in sublist]

            for id in child_task_list:
                task_obj = Task.objects.get(pk=id)
                task_obj.delete()

            response = {
                'success': True,
            }
            return JsonResponse(response, safe=False)


class AddNewSection(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            project_id = request.POST.get('project_id')
            project = get_object_or_404(Project, proj_id=project_id)
            _now = datetime.datetime.now()

            tree_root_sort_order = Task.objects.filter(project=project, task_deleted=False, parent="0").count()
            sort_order = tree_root_sort_order+1

            create_section = Task.objects.create(text='New Section',
                                              parent='0',
                                              project=project,
                                              created_by=request.user,
                                              start_date=_now,
                                              sort_order=sort_order,
                                              progress=0.0,
                                              is_section=True)
            task_id = create_section.id

            # call create task log history func
            task_name = create_section.text
            project_name = project.proj_name
            task_log_status = 'TASK_ADD'
            task_log_message = 'Section Task ID: '+str(task_id)+ ', Section Task Name: '+str(task_name)+' added to Project ID: '+str(project_id)+', Project Name: '+str(project_name)+' successfully.'
            createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)

            response = {
                'success': True,
                'task_id':task_id,
            }
            return JsonResponse(response, safe=False)


class GetSectionLabel(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():

            section_task_id = request.GET.get("section_task_id")
            task = get_object_or_404(Task, id=section_task_id)

            # Ensure task is section task, that means is_section = True
            section_task = task.is_section
            if section_task:
                is_section = "True"
            else:
                is_section = "False"

            response = {
                "success": True,
                "section_label": task.section_label or {},
                "is_section": is_section,
            }
            return JsonResponse(response, safe=False)


class AddSectionLabel(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            section_task_id = request.POST.get("section_task_id")
            key = request.POST.get("section_label_key").strip()
            value = request.POST.get("section_label_value").strip()

            # Ensure required fields are provided
            if not section_task_id or not key or not value:
                response = {
                    "success": False,
                    "error": "All fields are required."
                }
                return JsonResponse(response, safe=False)

            task = get_object_or_404(Task, id=section_task_id)

            # Ensure task is section task, that means is_section = True
            section_task = task.is_section
            if not section_task:
                response = {
                    "success": False,
                    "error": "You can not add label, because this is not a section task!"
                }
                return JsonResponse(response, safe=False)

            # Ensure section_label exists
            if task.section_label is None:
                task.section_label = {}

            # Check if the key already exists (case-insensitive check)
            existing_keys = {k.lower(): k for k in task.section_label.keys()}  # Map lowercase to original keys
            if key.lower() in existing_keys:
                response = {
                    "success": False,
                    "error": "Section label name already exists, please try another name!",
                }
                return JsonResponse(response, safe=False)

            # Add new label
            task.section_label[key] = value
            task.save()

            response = {
                "success": True,
                "section_label": task.section_label,
            }
            return JsonResponse(response, safe=False)


class UpdateSectionLabel(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            section_task_id = request.POST.get("section_task_id")
            old_key = request.POST.get("old_section_label_key")
            new_key = request.POST.get("new_section_label_key").strip()
            new_value = request.POST.get("new_section_label_value").strip()

            if not new_key or not new_value:
                return JsonResponse({"success": False, "error": "Section label name and value cannot be empty."})

            task = get_object_or_404(Task, id=section_task_id)

            if task.section_label is None:
                return JsonResponse({"success": False, "error": "No section label found."})

            section_label = task.section_label  # Preserve order

            if old_key not in section_label:
                return JsonResponse({"success": False, "error": "Original label key not found."})

            # Case-Insensitive Check for Duplicate Keys**
            existing_keys_lower = {k.lower(): k for k in section_label.keys()}  # Create case-insensitive mapping

            if new_key.lower() in existing_keys_lower and existing_keys_lower[new_key.lower()] != old_key:
                return JsonResponse(
                    {"success": False, "error": "Section label name already exists, please try another name!"})

            if old_key == new_key:
                # Just update the value without changing order
                section_label[old_key] = new_value
            else:
                # Update key while keeping order
                updated_label = collections.OrderedDict()
                for k, v in section_label.items():
                    if k == old_key:
                        updated_label[new_key] = new_value  # Replace with new key-value
                    else:
                        updated_label[k] = v  # Preserve other entries
                section_label = updated_label

            task.section_label = section_label
            task.save()
            response = {
                "success": True,
                "section_label": task.section_label,
            }
            return JsonResponse(response, safe=False)


class DeleteSectionLabel(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            section_task_id = request.POST.get("section_task_id")
            key = request.POST.get("section_label_key")

            task = get_object_or_404(Task, id=section_task_id)

            if task.section_label and key in task.section_label:
                del task.section_label[key]
                task.save()
                response = {
                    "success": True,
                }
                return JsonResponse(response, safe=False)

            response = {
                "success": False,
                "error":"Section Label key not found!",
            }
            return JsonResponse(response, safe=False)


class AddUpdateDeleteSignature(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            try:
                task_id = request.POST.get('task_id')
                action = request.POST.get('action')  # "add", "update", or "delete"
                col_id = request.POST.get('col_id')
                user = request.user
                task = get_object_or_404(Task, id=task_id)

                task_name = task.text
                project_name = task.project.proj_name
                project_id = task.project.proj_id


                # Ensure `task_extra_column` is initialized**
                if task.task_extra_column is None:
                    task.task_extra_column = {"column_details": {}}

                # Load existing task_extra_column**
                existing_column_details = task.task_extra_column.get("column_details", {})

                # Ensure `field_value` and `message` are always defined**
                field_value = ""  # Default value
                message = "Invalid action!"  # Default error message

                if action == "add" or action == "update":
                    if action == "add":
                        task_log_status = 'SIGNATURE_ADD'
                        str_added_or_updated = 'added'
                    else:
                        task_log_status = 'SIGNATURE_UPD'
                        str_added_or_updated = 'updated'
                    try:
                        time_zone = get_object_or_404(Profile, user=user).client_tz
                    except:
                        time_zone = "Asia/Kolkata"

                    local_time = localtime(now()).astimezone(pytz.timezone(time_zone))
                    current_datetime = local_time.strftime("%d %B, %Y | %I:%M:%S %p")
                    signature = f"{user.username} | {current_datetime}"
                    field_value = signature
                    message = "Signature added successfully!" if action == "add" else "Signature updated successfully!"

                    # Update the dictionary while preserving existing keys**
                    existing_column_details[col_id] = field_value

                    # call create task log history func
                    task_log_message = 'Signature ' + str_added_or_updated +' successfully for Task ID: ' + str(task_id) \
                                       + ', Task Name: ' + str(task_name) + ', Column Name: ' + str(col_id) \
                                       + ' and Project ID: ' + str(project_id) + ', Project Name: ' + str(project_name)
                    createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)
                elif action == "delete":
                    # call create task log history func
                    task_log_status = 'SIGNATURE_DEL'
                    task_log_message = 'Signature deleted successfully for Task ID: ' + str(task_id) \
                                       + ', Task Name: ' + str(task_name) + ', Column Name: ' + str(col_id) \
                                       + ' and Project ID: ' + str(project_id) + ', Project Name: ' + str(project_name)
                    createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)

                    # Remove only the specified key, keep others**
                    field_value = ""
                    message = "Signature deleted successfully!"
                    if col_id in existing_column_details:
                        del existing_column_details[col_id]

                # Update the task object without losing existing data**
                task.task_extra_column["column_details"] = existing_column_details
                task.save()


                return JsonResponse({
                    'success': True,
                    'task_id': task.id,
                    'signature': field_value,
                    'message': message
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})


class MakeSectionORTask(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            task_id = request.POST.get('task_id')
            task = Task.objects.get(pk=task_id)

            if task.is_section == True:
                task.is_section = False
            else:
                task.is_section = True
            task.save()

            response = {
                'task_id':task_id,
                'success': True,
            }
            return JsonResponse(response, safe=False)


class GetSectionMetadata(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():

            section_metadata_task_id = request.GET.get("section_metadata_task_id")
            task = get_object_or_404(Task, id=section_metadata_task_id)

            # Ensure task is section task, that means is_section = True
            section_task = task.is_section
            if section_task:
                is_section = "True"
            else:
                is_section = "False"

            # Fetch metadata from SectionTemplateMetadata
            metadata_objects = SectionTemplateMetadata.objects.filter(section_task_id=section_metadata_task_id)

            # Construct column_details JSON
            column_details = []
            for metadata in metadata_objects:
                column_details.append({
                    "name": metadata.section_task_metadata.field_title,  # Get field title
                    "type": metadata.section_task_metadata.field_type,  # Get field type
                    "choices": metadata.section_temp_meta_choices if metadata.section_temp_meta_choices else [],
                    "id": metadata.section_temp_meta_id,  # Unique ID
                    "position": int(
                        metadata.section_temp_meta_position) if metadata.section_temp_meta_position else None,
                })

            # Sort column details by position
            column_details.sort(key=lambda x: x["position"] if x["position"] is not None else float("inf"))

            # Construct response
            response = {
                "success": True,
                "section_metadata": {"column_details": column_details},
                "is_section": is_section,
            }
            return JsonResponse(response, safe=False)


class GetSectionMetadataFields(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == "GET" and request.is_ajax():
            # Fetch all ProjectField records
            project_fields = ProjectField.objects.all().values("field_id", "field_title", "field_type")

            # Convert to list of dictionaries
            field_options = list(project_fields)

            return JsonResponse({"success": True, "fields": field_options}, safe=False)


class GetMetadataFieldChoices(View):

    def get(self, request, field_id):

        if request.method == "GET" and request.is_ajax():

            try:
                # Fetch the selected ProjectField and its choices
                field = get_object_or_404(ProjectField, field_id=field_id)
                if field.field_type in ["single-select", "multi-select"]:
                    choices = field.field_choice.all().values_list("f_choice", flat=True)
                else:
                    choices = []
                return JsonResponse({"success": True, "choices": list(choices)})
            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)}, status=400)


class AddSectionMetadata(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == "POST":

            metadata_name = request.POST.get("section_metadata_key", "").strip()
            metadata_type = request.POST.get("section_metadata_type", "").strip()
            task_id = request.POST.get("section_metadata_task_id", "").strip()
            choices = request.POST.getlist("select_choices[]")

            if not metadata_name:
                return JsonResponse({"success": False, "error": "Metadata name is required!"})

            # Handle numeric metadata_name (treat as field_id)
            if metadata_name.isnumeric():
                field_id = int(metadata_name)
                field = get_object_or_404(ProjectField, field_id=field_id)
                metadata_name = field.field_title  # Convert ID to actual field name

            # Check if metadata already exists
            project_field = ProjectField.objects.filter(field_title=metadata_name).first()
            if not project_field:
                field_description = " ".join(metadata_name.split("_")).capitalize()
                project_field = ProjectField.objects.create(
                    field_title=metadata_name,
                    field_type=metadata_type,
                    field_description=field_description
                )

            # Fetch or create FieldChoice objects and assign them to ProjectField
            if choices:
                field_choices = []
                for choice in choices:
                    field_choice, created = FieldChoice.objects.get_or_create(f_choice=choice)
                    field_choices.append(field_choice)

                # Add choices to project_field's ManyToMany field
                project_field.field_choice.add(*field_choices)

            # Check if section metadata already exists to prevent duplicates
            existing_metadata = SectionTemplateMetadata.objects.filter(
                section_task_metadata=project_field,
                section_task_id=task_id
            ).first()

            if existing_metadata:
                return JsonResponse({"success": False, "error": "Metadata already exists!"})

            # Get the highest section_temp_meta_position for the given task_id
            max_position = SectionTemplateMetadata.objects.filter(section_task_id=task_id).aggregate(
                Max('section_temp_meta_position')
            )['section_temp_meta_position__max']

            # Increment position (default to 1 if none exist)
            new_position = int(max_position) + 1 if max_position else 1

            # Create new metadata with incremental position

            section_metadata = SectionTemplateMetadata.objects.create(
                section_task_metadata=project_field,
                section_task_id=task_id,
                section_temp_meta_position=new_position,
                section_temp_meta_choices=choices if choices else None
            )

            SectionMetadataValues.objects.create(
                section_task_id=task_id,
                section_task_metadata=project_field,
                section_metadata_value="",  # Empty value initially
                section_meta_val_position=1
            )

            # Fetch metadata from SectionTemplateMetadata
            metadata_objects = SectionTemplateMetadata.objects.filter(section_task_id=task_id)

            # Construct column_details JSON
            column_details = []
            for metadata in metadata_objects:
                column_details.append({
                    "name": metadata.section_task_metadata.field_title,  # Get field title
                    "type": metadata.section_task_metadata.field_type,  # Get field type
                    "choices": metadata.section_temp_meta_choices if metadata.section_temp_meta_choices else [],
                    "id": metadata.section_temp_meta_id,  # Unique ID
                    "position": int(
                        metadata.section_temp_meta_position) if metadata.section_temp_meta_position else None,
                })

            # Sort column details by position
            column_details.sort(key=lambda x: x["position"] if x["position"] is not None else float("inf"))

            return JsonResponse({
                "success": True,
                "message": "Metadata added successfully!",
                "section_metadata": {"column_details": column_details},
            })

        return JsonResponse({"success": False, "message": "Invalid request!"})


class DeleteSectionMetadata(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == "POST" and request.is_ajax():

            try:
                section_temp_meta_id = request.POST.get("section_temp_meta_id")
                section_metadata_task_id = request.POST.get("section_metadata_task_id")

                # Check if metadata exists
                metadata = SectionTemplateMetadata.objects.filter(
                    section_temp_meta_id=section_temp_meta_id,
                    section_task_id=section_metadata_task_id
                ).first()

                if not metadata:
                    return JsonResponse({"success": False, "error": "Metadata not found!"})

                # Delete related SectionMetadataValues (using section_task_metadata reference)
                SectionMetadataValues.objects.filter(section_task_id=section_metadata_task_id, section_task_metadata=metadata.section_task_metadata).delete()

                # Delete metadata entry
                metadata.delete()

                # Reset section_temp_meta_position for the same task_id
                remaining_metadata = SectionTemplateMetadata.objects.filter(
                    section_task_id=section_metadata_task_id
                ).order_by("section_temp_meta_position")

                for index, meta in enumerate(remaining_metadata, start=1):
                    meta.section_temp_meta_position = int(index)  # Convert to int to match PositiveIntegerField
                    meta.save()

                # Fetch metadata from SectionTemplateMetadata
                metadata_objects = SectionTemplateMetadata.objects.filter(section_task_id=section_metadata_task_id)

                # Construct column_details JSON
                column_details = []
                for metadata in metadata_objects:
                    column_details.append({
                        "name": metadata.section_task_metadata.field_title,  # Get field title
                        "type": metadata.section_task_metadata.field_type,  # Get field type
                        "choices": metadata.section_temp_meta_choices if metadata.section_temp_meta_choices else [],
                        "id": metadata.section_temp_meta_id,  # Unique ID
                        "position": int(
                            metadata.section_temp_meta_position) if metadata.section_temp_meta_position else None,
                    })

                # Sort column details by position
                column_details.sort(key=lambda x: x["position"] if x["position"] is not None else float("inf"))

                return JsonResponse({
                    "success": True,
                    "message": "Section metadata deleted successfully!",
                    "section_metadata": {"column_details": column_details}
                })

            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)})

        return JsonResponse({"success": False, "error": "Invalid request method"})


class UpdateSectionMetadataOrder(LoginRequiredMixin, View):

    def post(self, request, *args, **kwargs):

        if request.method == "POST" and request.is_ajax():

            try:
                data = json.loads(request.POST.get("ordered_metadata", "[]"))
                section_metadata_task_id = request.POST.get("section_metadata_task_id")

                if not data:
                    return JsonResponse({"success": False, "error": "No metadata received"})

                for item in data:
                    metadata_id = item.get("id")
                    new_position = int(item.get("position"))

                    if not metadata_id or not new_position:
                        continue  # Skip invalid data

                    # Ensure we're updating records only for the specified task ID
                    updated_rows = SectionTemplateMetadata.objects.filter(
                        section_temp_meta_id=metadata_id,
                        section_task_id=section_metadata_task_id
                    ).update(section_temp_meta_position=new_position)

                    if updated_rows == 0:
                        print(f"Warning: Metadata ID {metadata_id} not found or not updated.")

                return JsonResponse({"success": True, "message": "Metadata order updated successfully!"})

            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)})

        return JsonResponse({"success": False, "error": "Invalid request method"})


class NestedGridMetadataView(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == "GET" and request.is_ajax():

            if not request.is_ajax():
                return JsonResponse({"success": False, "error": "Invalid request!"})

            task_id = request.GET.get("task_id")
            try:
                task = Task.objects.get(id=task_id)
            except Task.DoesNotExist:
                return JsonResponse({"success": False, "error": "Task not found!"})

            # Fetch section metadata details from SectionTemplateMetadata
            column_details_qs = SectionTemplateMetadata.objects.filter(section_task_id=task_id).order_by(
                "section_temp_meta_position")

            user_list = [user.username for user in User.objects.all()]
            project = task.project
            proj_section_master_name = project.proj_section_master_name

            columns = []
            for item in column_details_qs:
                field_id = item.section_task_metadata.field_title
                field_title = " ".join(field_id.split("_"))
                order = item.section_temp_meta_id
                position = int(item.section_temp_meta_position)
                type = item.section_task_metadata.field_type

                # Initialize config dict if missing
                if item.section_temp_meta_config is None:
                    item.section_temp_meta_config = {}
                section_temp_meta_config = item.section_temp_meta_config or {}  # Ensure it's a dictionary
                width = section_temp_meta_config.get("width", None)  # Access 'width' safely

                if type == 'string':
                    string_column = {
                        'id': field_id,
                        'type': "string",
                        'editorType': "input",
                        'width': width,
                        'order': order,
                        'position': position,
                        'header': [{'text': field_title, 'align': 'left'}],
                    }
                    columns.append(string_column)
                elif type == 'number':
                    number_column = {
                        'id': field_id,
                        'type': "number",
                        'format': "#.00",
                        'width': width,
                        'order': order,
                        'position': position,
                        'header': [{'text': field_title, 'align': 'left'}],
                    }
                    columns.append(number_column)
                elif type == 'textarea':
                    textarea_column = {
                        'id': field_id,
                        'type': "string",
                        'editorType': "textarea",
                        'width': width,
                        'header': [{'text': field_title, 'align': 'left'}],
                        'order': order,
                        'position': position,
                    }
                    columns.append(textarea_column)
                elif type == 'bool':
                    bool_column = {
                        'id': field_id,
                        'type': 'boolean',
                        'header': [{'text': field_title, 'align': 'left'}],
                        'order': order,
                        'width': width,
                        'position': position,
                    }
                    columns.append(bool_column)
                elif type == 'date':
                    date_column = {
                        'id': field_id,
                        'width': width,
                        'header': [{'text': field_title, 'align': 'left'}],
                        'type': 'date',
                        'format': '%d %M, %Y',
                        'align': 'center',
                        'order': order,
                        'position': position,
                    }
                    columns.append(date_column)
                elif type == 'single-select':
                    try:
                        matching_section_master = SectionMaster.objects.get(
                            section_master_name=proj_section_master_name,
                            section_master_metadata_name=field_id,
                            is_sub_choice=False,
                        )
                        if matching_section_master.is_completed:
                            options = list(matching_section_master.section_master_choice.order_by("-id").values_list("master_choice", flat=True))
                        else:
                            options = []
                    except SectionMaster.DoesNotExist:
                        options = item.section_temp_meta_choices
                    single_select_column = {
                        'id': field_id,
                        'width': width,
                        'header': [{'text': field_title, 'align': 'left'}, {'content': 'selectFilter'}],
                        'editorType': 'combobox',
                        'options': options,
                        'order': order,
                        'position': position,
                    }
                    columns.append(single_select_column)
                elif type == 'multi-select':
                    options = item.section_temp_meta_choices
                    multi_select_column = {
                        'id': field_id,
                        'width': width,
                        'header': [{'text': field_title, 'align': 'left'}, {'content': 'selectFilter'}],
                        'editorType': 'multiselect',
                        'options': options,
                        'order': order,
                        'position': position,
                    }
                    columns.append(multi_select_column)
                elif type == 'people':
                    options = user_list
                    multi_select_column = {
                        'id': field_id,
                        'width': width,
                        'header': [{'text': field_title, 'align': 'left'}, {'content': 'selectFilter'}],
                        'editorType': 'multiselect',
                        'options': options,
                        'customType': 'people',
                        'order': order,
                        'position': position,
                    }
                    columns.append(multi_select_column)
                elif type == 'signature':
                    string_column = {
                        'id': field_id,
                        'type': "string",
                        'editorType': "input",
                        'width': width,
                        'header': [{'text': field_title, 'align': 'left'}],
                        'customType': 'signature',
                        'order': order,
                        'position': position,
                    }
                    columns.append(string_column)
                else:
                    pass

            # Fetch task data values from SectionMetadataValues
            task_data_qs = SectionMetadataValues.objects.filter(section_task_id=task_id).order_by(
                "section_meta_val_position")

            task_data_dict = {}

            for item in task_data_qs:
                position = int(item.section_meta_val_position)
                field_name = item.section_task_metadata.field_title  # Get field title
                field_value = item.section_metadata_value  # Get metadata value
                order = int(item.section_meta_val_position)

                if position not in task_data_dict:
                    task_data_dict[position] = {"position": position, "order": order}

                task_data_dict[position][field_name] = field_value


            # Convert dictionary to list
            task_data_list = sorted(task_data_dict.values(), key=lambda x: x["position"])

            try:
                column_config_details = task.section_metadata.get("column_config_details", [])
            except:
                column_config_details = []

            time_zone = get_object_or_404(Profile, user=request.user).client_tz

            context = {
                'columns': columns,
                'task_data_list': task_data_list,
                'time_zone': time_zone,
                'client_name': connection.tenant.name,
                'column_config_details': column_config_details
            }
            return JsonResponse({"success": True, "context": context})


class UpdateColumnOrder(LoginRequiredMixin, View):

    def post(self, request):

        try:
            data = json.loads(request.body.decode('utf-8'))  # Parse JSON data
            new_order = data.get("new_order", [])  # List of column IDs and positions
            task_id = data.get("task_id")

            # Ensure task exists
            task = get_object_or_404(Task, id=task_id)

            # Update section_temp_meta_position for each column
            for column in new_order:
                column_id = column["id"]
                new_position = int(column["position"])

                # Update the column order in SectionTemplateMetadata
                SectionTemplateMetadata.objects.filter(section_temp_meta_id=column_id, section_task_id=task_id).update(
                    section_temp_meta_position=new_position
                )

            return JsonResponse({"success": True, "message": "Column order updated!"})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)


class UpdateRowOrder(LoginRequiredMixin, View):

    def post(self, request):

        try:
            data = json.loads(request.body.decode("utf-8"))  # Parse JSON data
            new_order = data.get("new_order", [])  # List of rows with updated positions
            task_id = data.get("task_id")

            # Ensure the task exists
            task = get_object_or_404(Task, id=task_id)

            # Get all relevant SectionMetadataValues for this task
            section_metadata_values = SectionMetadataValues.objects.filter(section_task_id=task_id)

            # Create a mapping of current_position -> new_position
            position_mapping = {int(row.get("order")): int(row.get("position")) for row in new_order}

            # Fetch all rows that need to be updated
            rows_to_update = section_metadata_values.filter(
                section_meta_val_position__in=position_mapping.keys()
            )

            # Update each row individually to ensure proper swaps
            for row in rows_to_update:
                current_position = row.section_meta_val_position
                if current_position in position_mapping:
                    row.section_meta_val_position = position_mapping[current_position]
                    row.save()  # Save individually to avoid conflicts

            return JsonResponse({"success": True, "message": "Row order updated successfully!"})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)


class ModifyNestedGridRow(LoginRequiredMixin, View):

    def post(self, request):

        try:
            data = json.loads(request.body)
            task_id = data.get("task_id")
            position = int(data.get("position"))
            action = data.get("action")

            # Ensure the task exists
            task = get_object_or_404(Task, id=task_id)

            # Fetch all section metadata values related to this task
            section_metadata_values = list(SectionMetadataValues.objects.filter(
                section_task_id=task_id).order_by("section_meta_val_position"))

            # Fetch only the rows related to this task_id
            matching_rows = SectionMetadataValues.objects.filter(
                section_task_id=task_id,
                section_meta_val_position=position
            )

            # Find row index safely using filter (avoid missing row issues)
            # matching_rows = [row for row in section_metadata_values if int(row.section_meta_val_position) == position]
            for row in matching_rows:
                print("row id:->",row.section_meta_val_id)
            if not matching_rows:
                return JsonResponse({"success": False, "error": f"Row with position {position} not found"}, status=404)

            if action == "insert_above":
                new_position = position
                self.shift_positions_down(section_metadata_values, start_from=position)

                for item in matching_rows:
                    self.create_new_row(task_id, item.section_task_metadata, new_position)

            elif action == "insert_below":
                new_position = position + 1
                self.shift_positions_down(section_metadata_values, start_from=new_position)

                for item in matching_rows:
                    self.create_new_row(task_id, item.section_task_metadata, new_position)

            elif action == "clear_row":
                for item in matching_rows:
                    item.section_metadata_value = ""  # Clear only data fields
                    item.save()

            elif action == "delete_row":
                if position == 1:
                    return JsonResponse({"success": False, "error": "You cannot delete the first row!"}, status=400)

                # Delete the row
                for item in matching_rows:
                    item.delete()

                # Shift positions up after deletion
                self.shift_positions_up(section_metadata_values, start_from=position)

            # Fetch updated task data values from SectionMetadataValues
            task_data_qs = SectionMetadataValues.objects.filter(section_task_id=task_id).order_by("section_meta_val_position")

            # Convert to JSON format
            task_data_list = self.convert_to_json(task_data_qs)

            return JsonResponse({"success": True, "updated_data": task_data_list, "message": f"Action '{action}' performed successfully!"})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    def shift_positions_down(self, queryset, start_from):
        """
        Shifts all positions downward (increases position) starting from a given index.
        """
        SectionMetadataValues.objects.filter(
            section_task_id=queryset[0].section_task_id,
            section_meta_val_position__gte=start_from
        ).update(
            section_meta_val_position=Cast(F('section_meta_val_position'), IntegerField()) + 1
        )

    def shift_positions_up(self, queryset, start_from):
        """
        Shifts all positions upward (decreases position) starting from a given index.
        """
        SectionMetadataValues.objects.filter(
            section_task_id=queryset[0].section_task_id,
            section_meta_val_position__gt=start_from
        ).update(
            section_meta_val_position=Cast(F('section_meta_val_position'), IntegerField()) - 1
        )

    def create_new_row(self, task_id, section_task_metadata, position):
        """
        Creates a new blank row at the specified position.
        """
        SectionMetadataValues.objects.create(
            section_task_id=task_id,
            section_meta_val_position=position,  # Store as string
            section_metadata_value="",
            section_task_metadata=section_task_metadata
        )

    def convert_to_json(self, queryset):
        """
        Converts queryset data into JSON format similar to `section_metadata["data"]`
        """
        task_data_dict = {}

        for item in queryset:
            position = int(item.section_meta_val_position)
            field_name = item.section_task_metadata.field_title  # Get field title
            field_value = item.section_metadata_value  # Get metadata value

            if position not in task_data_dict:
                task_data_dict[position] = {"position": position}

            task_data_dict[position][field_name] = field_value

        return sorted(task_data_dict.values(), key=lambda x: x["position"])


# class UpdateSectionMetadataValues(LoginRequiredMixin, View):
#
#     def post(self, request):
#
#         if request.method == "POST" and request.is_ajax():
#
#             try:
#                 task_id = request.POST.get("task_id")
#                 position = request.POST.get("position")
#                 column_id = request.POST.get("column_id")
#                 value = request.POST.get("value")
#                 section_task_metadata = ProjectField.objects.get(field_title=column_id)
#                 section_metadata_values = SectionMetadataValues.objects.update_or_create(section_task_metadata=section_task_metadata,
#                                                                                          section_task_id=task_id,
#                                                                                          section_meta_val_position=int(position),
#                                                                                          defaults={'section_metadata_value': value})
#
#                 return JsonResponse({
#                     "success": True,
#                     "message": "Section metadata values updated successfully!",
#                 })
#
#             except Exception as e:
#                 return JsonResponse({"success": False, "error": str(e)})
#
#         return JsonResponse({"success": False, "error": "Invalid request method"})


class UpdateSectionMetadataValues(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == "POST" and request.is_ajax():
            try:
                task_id = request.POST.get("task_id")
                position = request.POST.get("position")
                column_id = request.POST.get("column_id")
                value = request.POST.get("value")

                section_task_metadata = ProjectField.objects.get(field_title=column_id)

                # Fetch the existing value (if exists)
                existing = SectionMetadataValues.objects.filter(
                    section_task_metadata=section_task_metadata,
                    section_task_id=task_id,
                    section_meta_val_position=int(position)
                ).first()

                if existing:
                    if existing.section_metadata_value != value:
                        old_value = existing.section_metadata_value
                        existing.section_metadata_value = value
                        existing.save()

                        # Extra safety (if override is skipped)
                        MetadataTracking.objects.create(
                            section_metadata_value=existing,
                            changed_field="section_metadata_value",
                            old_value=old_value,
                            new_value=value,
                            changed_by=request.user,
                            changed_at=now()
                        )
                else:
                    # New value creation
                    new_value = SectionMetadataValues.objects.create(
                        section_task_metadata=section_task_metadata,
                        section_task_id=task_id,
                        section_meta_val_position=int(position),
                        section_metadata_value=value
                    )
                    new_value.save()

                return JsonResponse({
                    "success": True,
                    "message": "Section metadata values updated successfully!",
                })

            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)})

        return JsonResponse({"success": False, "error": "Invalid request method"})


class GridAddUpdateDeleteSignature(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            try:
                task_id = request.POST.get('task_id')
                action = request.POST.get('action')  # "add", "update", or "delete"
                col_id = request.POST.get('col_id')
                row_position = request.POST.get('row_position')
                user = request.user
                task = get_object_or_404(Task, id=task_id)

                task_name = task.text
                project_name = task.project.proj_name
                project_id = task.project.proj_id

                # Ensure `field_value` and `message` are always defined**
                field_value = ""  # Default value
                message = "Invalid action!"  # Default error message

                if action == "add" or action == "update":
                    if action == "add":
                        task_log_status = 'SIGNATURE_ADD'
                        str_added_or_updated = 'added'
                    else:
                        task_log_status = 'SIGNATURE_UPD'
                        str_added_or_updated = 'updated'
                    try:
                        time_zone = get_object_or_404(Profile, user=user).client_tz
                    except:
                        time_zone = "Asia/Kolkata"

                    local_time = localtime(now()).astimezone(pytz.timezone(time_zone))
                    current_datetime = local_time.strftime("%d %B, %Y | %I:%M:%S %p")
                    signature = f"{user.username} | {current_datetime}"
                    field_value = signature
                    message = "Signature added successfully!" if action == "add" else "Signature updated successfully!"

                    # call create task log history func
                    task_log_message = 'Signature ' + str_added_or_updated +' successfully for Task ID: ' + str(task_id) \
                                       + ', Task Name: ' + str(task_name) + ', Column Name: ' + str(col_id) \
                                       + ' and Project ID: ' + str(project_id) + ', Project Name: ' + str(project_name)
                    createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)
                elif action == "delete":
                    # call create task log history func
                    task_log_status = 'SIGNATURE_DEL'
                    task_log_message = 'Signature deleted successfully for Task ID: ' + str(task_id) \
                                       + ', Task Name: ' + str(task_name) + ', Column Name: ' + str(col_id) \
                                       + ' and Project ID: ' + str(project_id) + ', Project Name: ' + str(project_name)
                    createTaskLogHistory(request, project_id, task_id, task_log_status, task_log_message)

                    # Remove only the specified key, keep others**
                    field_value = ""
                    message = "Signature deleted successfully!"

                section_task_metadata = ProjectField.objects.get(field_title=col_id)

                # Fetch the existing value (if exists)
                existing = SectionMetadataValues.objects.filter(
                    section_task_metadata=section_task_metadata,
                    section_task_id=task_id,
                    section_meta_val_position=int(row_position)
                ).first()

                if existing:
                    if existing.section_metadata_value != field_value:
                        old_value = existing.section_metadata_value
                        existing.section_metadata_value = field_value
                        existing.save()

                        # Extra safety (if override is skipped)
                        MetadataTracking.objects.create(
                            section_metadata_value=existing,
                            changed_field="section_metadata_value",
                            old_value=old_value,
                            new_value=field_value,
                            changed_by=request.user,
                            changed_at=now()
                        )
                else:
                    # New value creation
                    new_value = SectionMetadataValues.objects.update_or_create(
                        section_task_metadata=section_task_metadata,
                        section_task_id=task_id,
                        section_meta_val_position=int(row_position),
                        defaults={'section_metadata_value': field_value})
                    new_value.save()

                return JsonResponse({
                    'success': True,
                    'signature': field_value,
                    'message': message
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})


class GetSectionHeader(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():

            section_header_task_id = request.GET.get("section_header_task_id")

            if not section_header_task_id:
                return JsonResponse({"success": False, "error": "Task ID is required"}, status=400)

            # Fetch metadata from the database
            metadata = SectionTemplateMetadata.objects.filter(section_task_id=section_header_task_id).order_by(
                'section_temp_meta_position')

            # Format data for frontend
            section_metadata = [
                {
                    "id": meta.section_task_metadata.field_id,  # Assuming `id` is a field in `ProjectField`
                    "name": meta.section_task_metadata.field_title,  # Replace with actual field if different
                    "position": meta.section_temp_meta_position
                }
                for meta in metadata
            ]

            # Fetch section_metadata from Task Model (assuming section_metadata is a JSONField)
            try:
                task = Task.objects.get(id=section_header_task_id)
                section_metadata_json = task.section_metadata  # Assuming it's stored as JSONField
            except Task.DoesNotExist:
                section_metadata_json = {}

            return JsonResponse({
                "success": True,
                "section_header": section_metadata,
                "section_metadata_json": section_metadata_json
            })


class SaveSectionHeader(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            try:
                data = json.loads(request.body)
                task_id = data.get("task_id")
                new_column_config = data.get("column_config_details", [])

                task = Task.objects.get(id=task_id)

                # Initialize if empty
                if not task.section_metadata:
                    task.section_metadata = {"column_config_details": []}

                existing_config = task.section_metadata.get("column_config_details", [])

                # Get already used ids
                existing_ids = [item["id"] for item in existing_config]
                new_ids = [item["id"] for item in new_column_config]

                # Check for duplicates
                duplicates = list(set(existing_ids) & set(new_ids))
                if duplicates:
                    return JsonResponse({
                        "success": False,
                        "error": f"This metadata is already grouped under a header: {', '.join(duplicates)}"
                    })

                # Append new config to existing
                updated_config = existing_config + new_column_config
                task.section_metadata["column_config_details"] = updated_config
                task.save()

                return JsonResponse({
                    "success": True,
                    "section_metadata_json": task.section_metadata,
                    "message": "Section metadata saved successfully!"
                })

            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)})


class DeleteSectionHeaderGroup(LoginRequiredMixin, View):

    def post(self, request):

        try:
            task_id = request.POST.get('task_id')
            updated_config_json = request.POST.get('updated_config')

            if not task_id or not updated_config_json:
                return JsonResponse({"success": False, "error": "Missing data"}, status=400)

            task = Task.objects.get(id=task_id)
            task.section_metadata = json.loads(updated_config_json)
            task.save()

            return JsonResponse({"success": True})
        except Task.DoesNotExist:
            return JsonResponse({"success": False, "error": "Task not found"}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)


class AddUpdateSectionMetadataWidth(LoginRequiredMixin, View):

    def post(self, request):

        try:
            data = json.loads(request.body.decode('utf-8'))  # Parse JSON data
            task_id = data.get('task_id')
            config_col_id = data.get('config_col_id')
            config_metadata_id = data.get('config_metadata_id')
            config_metadata_position = data.get('config_metadata_position')
            config_metadata_width = data.get('config_metadata_width')

            # Validate required fields
            if not all([task_id, config_col_id, config_metadata_id, config_metadata_position, config_metadata_width]):
                return JsonResponse({"success": False, "error": "Missing data"}, status=400)

            try:
                config_metadata_width = int(config_metadata_width)
            except ValueError:
                return JsonResponse({"success": False, "error": "Invalid width value"}, status=400)

            try:
                section_metadata = SectionTemplateMetadata.objects.get(
                    section_task_id=task_id,
                    section_temp_meta_id=config_metadata_id,
                    section_temp_meta_position=config_metadata_position
                )
            except SectionTemplateMetadata.DoesNotExist:
                # Do nothing if not found
                return JsonResponse({"success": True, "message": "No update performed, metadata not found."})

            # Initialize config dict if missing
            if section_metadata.section_temp_meta_config is None:
                section_metadata.section_temp_meta_config = {}

            # Update the width
            section_metadata.section_temp_meta_config["width"] = config_metadata_width
            section_metadata.save()

            return JsonResponse({"success": True, "message": "Metadata column width updated successfully."})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)


class CopyProject(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            try:
                proj_id = request.POST.get('copy_proj_id')
                if not proj_id:
                    return JsonResponse({"success": False, "error": "Missing project ID"}, status=400)

                project = get_object_or_404(Project, proj_id=proj_id)

                proj_name = project.proj_name
                proj_description = project.proj_description
                proj_owner = request.user
                new_project = Project.objects.create(proj_name=proj_name,
                                                 proj_description=proj_description,
                                                 proj_owner=proj_owner,
                                                 proj_start_date=datetime.datetime.now(),
                                                 )

                groups = [gr.name for gr in project.proj_doc_group.all()]
                for group in groups:
                    gp = get_object_or_404(Group, name=group)
                    new_project.proj_doc_group.add(gp)
                new_project.save()

                try:
                    # call create project log history func
                    my_project_name = new_project.proj_name
                    my_project_id = str(new_project.proj_id)
                    project_log_status = 'PROJECT_COPY'
                    project_log_message = 'Project ID: ' + my_project_id + ', Project Name: ' + str(
                        my_project_name) + ' copied successfully.'
                    createProjectLogHistory(request, my_project_id, project_log_status, project_log_message)
                except:
                    pass

                # get or create Project Status
                project_status_choice_list = ['Not Started', 'In Progress', 'Completed', 'On Hold']
                for p_status in project_status_choice_list:
                    get_project_status, create_project_status = ProjectStatus.objects.get_or_create(proj_status=p_status)

                # save project status to Not Started
                get_proj_status_obj = ProjectStatus.objects.get(proj_status="Not Started")
                new_project.proj_status = get_proj_status_obj
                new_project.save()

                library_temp_id = project.proj_origin.proj_temp_id
                library_template = get_object_or_404(ProjectTemplate, proj_temp_id=library_temp_id)

                _now = datetime.datetime.now()
                proj_temp_id = _now.strftime('%Y%m%d%H%M%S%f')
                proj_template = ProjectTemplate.objects.create(proj_temp_id=proj_temp_id,
                                                               proj_temp_status='ac',
                                                               proj_temp_title=proj_name + ' template ' + str(proj_temp_id),
                                                               proj_temp_description=proj_name + ' template ' + str(
                                                                   proj_temp_id),
                                                               proj_temp_created_at=_now,
                                                               proj_temp_owner=request.user,
                                                               proj_temp_edited_by=request.user,
                                                               )
                cg = get_object_or_404(Group, name='COMMON')
                proj_template.proj_temp_doc_group.add(cg)
                proj_template.save()

                # save project template
                new_project.proj_origin = proj_template
                new_project.save()

                library_temp_field = ProjectTemplateField.objects.filter(proj_temp=library_template).order_by(
                    'proj_template_id')

                try:
                    for i, item in enumerate(library_temp_field):
                        field_position = item.proj_temp_field_position
                        get_project_field = item.proj_temp_field
                        proj_temp_field_hide = item.proj_temp_field_hide

                        ProjectTemplateField.objects.create(proj_temp=proj_template,
                                                            proj_temp_field=get_project_field,
                                                            proj_temp_field_position=field_position,
                                                            proj_temp_field_hide=proj_temp_field_hide)

                except Exception as e:
                    print("post CopyProject error:->", e)

                original_tasks = Task.objects.filter(project=project, task_deleted=False)
                task_id_map = {}  # To map old task_id -> new task_id

                new_task_start_date = datetime.datetime.now()

                # 1. Duplicate Tasks
                for task in original_tasks:
                    old_task_id = task.id

                    new_task = Task.objects.create(
                        text=task.text,
                        start_date=new_task_start_date,
                        end_date=None,
                        duration=task.duration,
                        progress=task.progress,
                        parent=task.parent,
                        sort_order=task.sort_order,
                        project=new_project,
                        status=task.status,
                        created_by=request.user,
                        note=task.note,
                        priority=task.priority,
                        task_extra_column=task.task_extra_column,
                        task_deleted=task.task_deleted,
                        is_section=task.is_section,
                        section_metadata=task.section_metadata,
                        section_label=task.section_label
                    )

                    # Assign ManyToMany
                    new_task.assigned_to.set(task.assigned_to.all())

                    task_id_map[str(old_task_id)] = str(new_task.id)

                # Step 3: Fix parent-child relationships
                for original_task in original_tasks:
                    old_id = str(original_task.id)
                    new_task_id = task_id_map[old_id]
                    new_task = Task.objects.get(id=new_task_id)
                    old_parent_id = original_task.parent

                    if old_parent_id and old_parent_id in task_id_map:
                        new_task.parent = task_id_map[old_parent_id]
                        new_task.save()

                # 2. Copy SectionTemplateMetadata for each task
                for old_task_id, new_task_id in task_id_map.items():
                    section_meta = SectionTemplateMetadata.objects.filter(section_task_id=old_task_id)
                    for meta in section_meta:
                        SectionTemplateMetadata.objects.create(
                            section_task_metadata=meta.section_task_metadata,
                            section_task_id=new_task_id,
                            section_temp_meta_position=meta.section_temp_meta_position,
                            section_temp_meta_choices=meta.section_temp_meta_choices,
                            section_temp_meta_config=meta.section_temp_meta_config or {}
                        )

                # 3. Copy SectionMetadataValues for each task
                for old_task_id, new_task_id in task_id_map.items():
                    section_values = SectionMetadataValues.objects.filter(section_task_id=old_task_id)
                    for val in section_values:
                        SectionMetadataValues.objects.create(
                            section_task_metadata=val.section_task_metadata,
                            section_metadata_value=val.section_metadata_value,
                            section_task_id=new_task_id,
                            section_meta_val_position=val.section_meta_val_position
                        )

                response = {
                    "success": True,
                    "message": "Project and all tasks duplicated successfully.",
                }
                return JsonResponse(response, safe=False)
            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)}, status=500)


class GetProjectDetailsView(LoginRequiredMixin, View):

    def get(self, request):
        project_id = request.GET.get('project_id')
        if not project_id:
            return JsonResponse({'success': False, 'error': 'No project ID provided.'})

        _now = datetime.datetime.now()
        _now_time = _now.strftime('%Y%m%d%H%M%S%f')
        try:
            project = Project.objects.get(proj_id=project_id)
            proj_name = str(project.proj_name) +" "+str(_now_time)
            proj_description = project.proj_description
            proj_group = [gr.name for gr in project.proj_doc_group.all()]
            project_group_list = [gr.name for gr in Group.objects.all()]

            return JsonResponse({
                'success': True,
                'name': proj_name,
                'description': proj_description,
                'proj_group': proj_group,
                'project_group_list': project_group_list,
            })
        except Project.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Project not found.'})


class CopyLibraryProject(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):

        proj_groups = Group.objects.all()
        projects = Project.objects.filter(proj_deleted=False, proj_archive=False).order_by('-proj_id').distinct()

        time_zone = get_object_or_404(Profile, user=self.request.user).client_tz

        context = {
            'proj_groups': proj_groups,
            'time_zone': time_zone,
            'projects': projects,
            'client_name': connection.tenant.name,
        }
        return render(request, 'project_management/copy_library_project.html', context)

    def post(self, request, *args, **kwargs):

        if request.method == 'POST' and request.is_ajax():

            try:
                proj_id = request.POST.get('projectID')
                if not proj_id:
                    return JsonResponse({"success": False, "error": "Missing project ID"}, status=400)

                project = get_object_or_404(Project, proj_id=proj_id)

                proj_name = request.POST.get('projectName')
                proj_description = request.POST.get('projectDescription')
                copy_proj_with_or_without_values = request.POST.get('copy_proj_with_or_without_values')

                proj_owner = request.user
                new_project = Project.objects.create(proj_name=proj_name,
                                                     proj_description=proj_description,
                                                     proj_owner=proj_owner,
                                                     proj_start_date=datetime.datetime.now(),
                                                     )

                project_groups = request.POST.get('project_groups')
                groups = json.loads(project_groups)
                for group in groups:
                    gp = get_object_or_404(Group, name=group)
                    new_project.proj_doc_group.add(gp)
                new_project.save()

                try:
                    # call create project log history func
                    my_project_name = new_project.proj_name
                    my_project_id = str(new_project.proj_id)
                    project_log_status = 'PROJECT_COPY'
                    project_log_message = 'Project ID: ' + my_project_id + ', Project Name: ' + str(
                        my_project_name) + ' copied successfully.'
                    createProjectLogHistory(request, my_project_id, project_log_status, project_log_message)
                except:
                    pass

                # get or create Project Status
                project_status_choice_list = ['Not Started', 'In Progress', 'Completed', 'On Hold']
                for p_status in project_status_choice_list:
                    get_project_status, create_project_status = ProjectStatus.objects.get_or_create(
                        proj_status=p_status)

                # save project status to Not Started
                get_proj_status_obj = ProjectStatus.objects.get(proj_status="Not Started")
                new_project.proj_status = get_proj_status_obj
                new_project.save()

                library_temp_id = project.proj_origin.proj_temp_id
                library_template = get_object_or_404(ProjectTemplate, proj_temp_id=library_temp_id)

                _now = datetime.datetime.now()
                proj_temp_id = _now.strftime('%Y%m%d%H%M%S%f')
                proj_template = ProjectTemplate.objects.create(proj_temp_id=proj_temp_id,
                                                               proj_temp_status='ac',
                                                               proj_temp_title=proj_name + ' template ' + str(
                                                                   proj_temp_id),
                                                               proj_temp_description=proj_name + ' template ' + str(
                                                                   proj_temp_id),
                                                               proj_temp_created_at=_now,
                                                               proj_temp_owner=request.user,
                                                               proj_temp_edited_by=request.user,
                                                               )
                cg = get_object_or_404(Group, name='COMMON')
                proj_template.proj_temp_doc_group.add(cg)
                proj_template.save()

                # save project template
                new_project.proj_origin = proj_template
                new_project.save()

                library_temp_field = ProjectTemplateField.objects.filter(proj_temp=library_template).order_by(
                    'proj_template_id')

                try:
                    for i, item in enumerate(library_temp_field):
                        field_position = item.proj_temp_field_position
                        get_project_field = item.proj_temp_field
                        proj_temp_field_hide = item.proj_temp_field_hide

                        ProjectTemplateField.objects.create(proj_temp=proj_template,
                                                            proj_temp_field=get_project_field,
                                                            proj_temp_field_position=field_position,
                                                            proj_temp_field_hide=proj_temp_field_hide)

                except Exception as e:
                    print("post CopyProject error:->", e)

                original_tasks = Task.objects.filter(project=project, task_deleted=False)
                task_id_map = {}  # To map old task_id -> new task_id

                new_task_start_date = datetime.datetime.now()

                # 1. Duplicate Tasks
                for task in original_tasks:
                    old_task_id = task.id

                    if copy_proj_with_or_without_values == "with_values":
                        task_extra_column = task.task_extra_column
                    else:
                        # Strip all values but keep keys
                        task_extra_column = {}
                        if isinstance(task.task_extra_column, dict):
                            for key, value in task.task_extra_column.items():
                                if isinstance(value, dict):
                                    task_extra_column[key] = {k: "" for k in value.keys()}
                                else:
                                    task_extra_column[key] = ""

                    new_task = Task.objects.create(
                        text=task.text,
                        start_date=new_task_start_date,
                        end_date=task.end_date if copy_proj_with_or_without_values == "with_values" else None,
                        duration=task.duration,
                        progress=task.progress,
                        parent=task.parent,
                        sort_order=task.sort_order,
                        project=new_project,
                        status=task.status if copy_proj_with_or_without_values == "with_values" else None,
                        created_by=request.user,
                        note=task.note if copy_proj_with_or_without_values == "with_values" else None,
                        priority=task.priority if copy_proj_with_or_without_values == "with_values" else None,
                        task_extra_column=task_extra_column,
                        task_deleted=task.task_deleted,
                        is_section=task.is_section,
                        section_metadata=task.section_metadata,
                        section_label=task.section_label
                    )

                    # Set assigned_to based on with/without values
                    if copy_proj_with_or_without_values == "with_values":
                        new_task.assigned_to.set(task.assigned_to.all())
                    else:
                        new_task.assigned_to.clear()

                    task_id_map[str(old_task_id)] = str(new_task.id)

                # 2. Fix parent-child relationships
                for original_task in original_tasks:
                    old_id = str(original_task.id)
                    new_task_id = task_id_map[old_id]
                    new_task = Task.objects.get(id=new_task_id)
                    old_parent_id = original_task.parent

                    if old_parent_id and old_parent_id in task_id_map:
                        new_task.parent = task_id_map[old_parent_id]
                        new_task.save()

                # 3. Copy SectionTemplateMetadata for each task
                for old_task_id, new_task_id in task_id_map.items():
                    section_meta = SectionTemplateMetadata.objects.filter(section_task_id=old_task_id)
                    for meta in section_meta:
                        SectionTemplateMetadata.objects.create(
                            section_task_metadata=meta.section_task_metadata,
                            section_task_id=new_task_id,
                            section_temp_meta_position=meta.section_temp_meta_position,
                            section_temp_meta_choices=meta.section_temp_meta_choices,
                            section_temp_meta_config=meta.section_temp_meta_config or {}
                        )

                # 4. Copy SectionMetadataValues for each task
                for old_task_id, new_task_id in task_id_map.items():
                    section_values = SectionMetadataValues.objects.filter(section_task_id=old_task_id)
                    for val in section_values:
                        SectionMetadataValues.objects.create(
                            section_task_metadata=val.section_task_metadata,
                            section_metadata_value=val.section_metadata_value if copy_proj_with_or_without_values == "with_values" else "",
                            section_task_id=new_task_id,
                            section_meta_val_position=val.section_meta_val_position
                        )
                new_project_id = str(new_project.proj_id)
                response = {
                    "success": True,
                    "new_project_id":new_project_id,
                    "message": "Project and all tasks duplicated successfully.",
                }
                return JsonResponse(response, safe=False)
            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)}, status=500)


class GetSectionMetadataName(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():

            metadata_tracking_task_id = request.GET.get("metadata_tracking_task_id")

            if not metadata_tracking_task_id:
                return JsonResponse({"success": False, "error": "Task ID is required"}, status=400)

            # Fetch metadata from the database
            metadata = SectionTemplateMetadata.objects.filter(section_task_id=metadata_tracking_task_id).order_by(
                'section_temp_meta_position')

            # Format data for frontend
            section_metadata = [
                {
                    "id": meta.section_task_metadata.field_id,  # Assuming `id` is a field in `ProjectField`
                    "name": meta.section_task_metadata.field_title,  # Replace with actual field if different
                    "position": meta.section_temp_meta_position
                }
                for meta in metadata
            ]

            # Fetch section_metadata from Task Model (assuming section_metadata is a JSONField)
            try:
                task = Task.objects.get(id=metadata_tracking_task_id)
                section_metadata_json = task.section_metadata  # Assuming it's stored as JSONField
            except Task.DoesNotExist:
                section_metadata_json = {}

            return JsonResponse({
                "success": True,
                "section_header": section_metadata,
                "section_metadata_json": section_metadata_json
            })


class GetSectionMetadataValues(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():

            project_field_id = request.GET.get("project_field_id")
            task_id = request.GET.get("task_id")
            selected_row_position = request.GET.get("selected_row_position")

            if not project_field_id or not task_id:
                return JsonResponse({"success": False, "error": "Missing parameters"}, status=400)

            try:
                # Get ProjectField object
                project_field = ProjectField.objects.get(field_id=project_field_id)

                # Fetch distinct values for that field and task
                if selected_row_position:
                    values_qs = SectionMetadataValues.objects.filter(
                        section_task_metadata=project_field,
                        section_task_id=task_id,
                        section_meta_val_position=int(selected_row_position)
                    ).exclude(section_metadata_value__isnull=True).exclude(section_metadata_value__exact='').order_by('section_meta_val_position')
                else:
                    values_qs = SectionMetadataValues.objects.filter(
                        section_task_metadata=project_field,
                        section_task_id=task_id
                    ).exclude(section_metadata_value__isnull=True).exclude(section_metadata_value__exact='').order_by('section_meta_val_position')

                values = list(values_qs.values('section_metadata_value', 'section_meta_val_position'))

                return JsonResponse({"success": True, "values": values})

            except ProjectField.DoesNotExist:
                return JsonResponse({"success": False, "error": "Invalid Metadata ID"}, status=404)


class FetchMetadataTrackingValues(LoginRequiredMixin, View):

    def post(self, request):

        try:
            data = json.loads(request.body)
            task_id = data.get("task_id")
            position = int(data.get("position"))
            field_id = data.get("field_id")

            section_task_metadata = ProjectField.objects.get(field_id=field_id)
            # Ensure the task exists
            task = get_object_or_404(Task, id=task_id)

            # Fetch only the rows related to this task_id
            matching_row = SectionMetadataValues.objects.get(
                section_task_id=task_id,
                section_task_metadata = section_task_metadata,
                section_meta_val_position=position
            )

            if not matching_row:
                return JsonResponse({"success": False, "error": f"Row with position {position} not found"}, status=404)

            history = matching_row.metadata_history.order_by('-changed_at')
            changes = []

            try:
                time_zone = get_object_or_404(Profile, user=self.request.user).client_tz
            except:
                time_zone = "Asia/Kolkata"

            for h in history:
                local_dt = h.changed_at.astimezone(pytz.timezone(time_zone))
                changes.append({
                    "old_value": h.old_value,
                    "new_value": h.new_value,
                    "changed_by": str(h.changed_by),
                    "changed_at": local_dt.strftime("%d %b, %Y %I:%M:%S %p")
                })

            return JsonResponse({"success": True, "changes": changes})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)


class AddUpdateProjectSignature(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            try:
                proj_id = request.POST.get('add_update_proj_signature_id')
                signature_status = request.POST.get('signature_status')
                user = request.user
                project = get_object_or_404(Project, proj_id=proj_id)

                try:
                    time_zone = get_object_or_404(Profile, user=user).client_tz
                except:
                    time_zone = "Asia/Kolkata"

                local_time = localtime(now()).astimezone(pytz.timezone(time_zone))
                current_datetime = local_time.strftime("%d %B, %Y | %I:%M:%S %p")
                signature = f"{user.username} | {current_datetime}"
                proj_signature = signature
                project.proj_signature = proj_signature
                project.save()

                # call create project log history func
                project_id = str(proj_id)
                project_name = str(project.proj_name)
                if signature_status == "Add":
                    project_log_status = 'PROJECT_SIGNATURE_ADD'
                    str_added_or_updated = 'added'
                else:
                    project_log_status = 'PROJECT_SIGNATURE_UPD'
                    str_added_or_updated = 'updated'
                project_log_message = 'Signature ' + str_added_or_updated + ' successfully for Project ID: ' + str(project_id) + ', Project Name: ' + str(project_name)
                createProjectLogHistory(request, project_id, project_log_status, project_log_message)

                return JsonResponse({
                    'success': True,
                    'proj_signature': proj_signature,
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})


class DeleteProjectSignature(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():

            try:
                proj_id = request.POST.get('delete_proj_signature_id')
                project = get_object_or_404(Project, proj_id=proj_id)
                project.proj_signature = ""
                project.save()

                # call create project log history func
                project_id = str(proj_id)
                project_name = str(project.proj_name)
                project_log_status = 'PROJECT_SIGNATURE_DEL'

                project_log_message = 'Signature deleted successfully for Project ID: ' + str(
                    project_id) + ', Project Name: ' + str(project_name)
                createProjectLogHistory(request, project_id, project_log_status, project_log_message)

                return JsonResponse({
                    'success': True,
                    'proj_signature': "",
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})


class ManageSectionMaster(LoginRequiredMixin, View):

    def get(self, request):

        user = self.request.user
        names = SectionMaster.objects.values_list('section_master_name', flat=True).distinct()
        fields = ProjectField.objects.values_list('field_title', flat=True)
        choices = SectionMasterChoice.objects.all().order_by('-id')  # Include all master choices

        context = {
            'client_name': connection.tenant.name,
            'names': list(names),
            'fields': list(fields),
            'user': user,
            'choices': choices,  # Pass to template
        }

        return render(request, 'project_management/manage_section_master.html', context)


class UpdateSpecialChar(View):

        def post(self, request, pk):

            if request.method == "POST" and request.is_ajax():
                try:
                    data = json.loads(request.body)
                    special_char = data.get("special_char", "").strip()
                    obj = SectionMasterChoice.objects.get(id=pk)
                    obj.special_char_choice = special_char
                    obj.save()
                    return JsonResponse({"success": True})
                except SectionMasterChoice.DoesNotExist:
                    return JsonResponse({"success": False, "error": "Choice not found."})


class DeleteSpecialChar(View):

    def post(self, request, pk):

        if request.method == "POST" and request.is_ajax():
            try:
                obj = SectionMasterChoice.objects.get(id=pk)
                obj.special_char_choice = None
                obj.save()
                return JsonResponse({"success": True})
            except SectionMasterChoice.DoesNotExist:
                return JsonResponse({"success": False, "error": "Choice not found."})


class CreateSectionMaster(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            data = request.POST
            name = data.get('section_master_name')
            metadata = data.get('section_master_metadata_name')
            choices = json.loads(data.get('choices', '[]'))
            section_sub_metadata_name = data.get('section_sub_metadata_name')
            section_sub_choice_name = data.get('section_sub_choice_name')

            # Check if the combination already exists
            exists = SectionMaster.objects.filter(
                section_master_name=name,
                section_master_metadata_name=metadata,
                section_sub_metadata_name=section_sub_metadata_name,
                section_sub_choice=section_sub_choice_name,
            ).exists()

            if exists:
                return JsonResponse({
                    'status': 'error',
                    'message': f'"{metadata}", "{section_sub_choice_name}" already exists for "{name}".<br>Please choose a different section master metadata name.'
                }, status=400)

            section_master = SectionMaster.objects.create(
                section_master_name=name,
                section_master_metadata_name=metadata,
                section_sub_metadata_name=section_sub_metadata_name,
                section_sub_choice=section_sub_choice_name,
            )

            if section_sub_choice_name:
                section_master.is_sub_choice = True
            else:
                section_master.is_sub_choice = False

            # Reverse the order of choices before processing
            for choice_str in reversed(choices):
                choice, _ = SectionMasterChoice.objects.get_or_create(master_choice=choice_str)
                section_master.section_master_choice.add(choice)

            section_master.save()
            return JsonResponse({'status': 'success'})

        return JsonResponse({'error': 'Invalid request'}, status=400)


class SectionMasterListView(LoginRequiredMixin, View):

    def get(self, request):

        data = []
        for sm in SectionMaster.objects.all().order_by("-id"):  # Latest SectionMasters first
            choices = sm.section_master_choice.order_by("-id").values_list("master_choice", flat=True)
            data.append({
                "id": sm.id,
                "section_master_name": sm.section_master_name,
                "section_master_metadata_name": sm.section_master_metadata_name,
                "choices": list(choices),
                "section_sub_metadata_name": sm.section_sub_metadata_name,
                "section_sub_choice_name": sm.section_sub_choice,
                "status": "Completed" if sm.is_completed else "In Progress",
                "is_completed": sm.is_completed,
            })
        return JsonResponse(data, safe=False)


class ToggleSectionMasterStatusView(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            section_id = request.POST.get("section_id")
            try:
                section = SectionMaster.objects.get(id=section_id)
                section.is_completed = not section.is_completed
                section.save()
                return JsonResponse({
                    "success": True,
                    "new_status": "Completed" if section.is_completed else "In Progress",
                    "is_completed": section.is_completed,
                })
            except SectionMaster.DoesNotExist:
                return JsonResponse({"success": False, "error": "SectionMaster not found"})

        return JsonResponse({"success": False, "error": "Invalid request method."})


class AddChoice(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            section_master_id = request.POST.get("section_master_id")
            choice_value = request.POST.get("choice", "").strip()

            if not section_master_id or not choice_value:
                return JsonResponse({"success": False, "error": "Missing data"})

            section_master = get_object_or_404(SectionMaster, id=section_master_id)

            # Check if the choice already exists (case-insensitive)
            if section_master.section_master_choice.filter(master_choice__iexact=choice_value).exists():
                return JsonResponse({
                    "success": False,
                    "error": f'Choice "{choice_value}" is already added. Please choose a different choice name.'
                })

            # Get or create the SectionMasterChoice object
            choice_obj, created = SectionMasterChoice.objects.get_or_create(master_choice=choice_value)

            # Link it to the SectionMaster
            section_master.section_master_choice.add(choice_obj)

            return JsonResponse({"success": True, "created": created})

        return JsonResponse({"success": False, "error": "Invalid request method."})


class RemoveChoice(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            section_master_id = request.POST.get("section_master_id")
            choice_value = request.POST.get("choice")

            if not section_master_id or not choice_value:
                return JsonResponse({"success": False, "error": "Missing data"})

            section_master = get_object_or_404(SectionMaster, id=section_master_id)
            try:
                choice_obj = SectionMasterChoice.objects.get(master_choice=choice_value)
                section_master.section_master_choice.remove(choice_obj)

                # # Optional: Delete the choice if no other section uses it
                # if choice_obj.sectionmaster_set.count() == 0:
                #     choice_obj.delete()

                return JsonResponse({"success": True})
            except SectionMasterChoice.DoesNotExist:
                return JsonResponse({"success": False, "error": "Choice not found"})


class DeleteSectionMaster(LoginRequiredMixin, View):

    def post(self, request):

        if request.method == 'POST' and request.is_ajax():
            section_master_id = request.POST.get("section_master_id")

            if not section_master_id:
                return JsonResponse({"success": False, "error": "ID missing"})

            try:
                section_master = SectionMaster.objects.get(id=section_master_id)
                section_master.delete()

                # # Optional: get related choices before deletion
                # related_choices = list(section_master.section_master_choice.all())

                # # Optional: remove unused choices
                # for choice in related_choices:
                #     if choice.sectionmaster_set.count() == 0:
                #         choice.delete()

                return JsonResponse({"success": True})
            except SectionMaster.DoesNotExist:
                return JsonResponse({"success": False, "error": "Not found"})


def get_matching_choices(request):
    query = request.GET.get("q", "")
    if query:
        matches = SectionMasterChoice.objects.filter(master_choice__icontains=query).values_list("master_choice", flat=True).distinct()[:5]
        return JsonResponse({"results": list(matches)})
    return JsonResponse({"results": []})


def get_existing_choices(request):

    metadata_name = request.GET.get("metadata_name")

    if not metadata_name:
        return JsonResponse({"success": False, "section_masters": []})

    try:
        section_masters = SectionMaster.objects.filter(
            section_master_metadata_name=metadata_name
        ).prefetch_related('section_master_choice')

        result = []
        for sm in section_masters:
            result.append({
                "id": sm.id,
                "section_master_name": sm.section_master_name,
                "section_sub_metadata_name": sm.section_sub_metadata_name,
                "section_sub_choice": sm.section_sub_choice,
                "section_master_metadata_name": sm.section_master_metadata_name,
                "choices": list(sm.section_master_choice.order_by("id").values_list("master_choice", flat=True)),
            })

        return JsonResponse({"success": True, "section_masters": result})

    except Exception as e:
        print("error:->", e)
        return JsonResponse({"success": False, "section_masters": [], "error": str(e)})


def get_section_sub_metadata_choices(request):

    section_master_name = request.GET.get("section_master_name")
    metadata_name = request.GET.get("metadata_name")

    if not metadata_name:
        return JsonResponse({"success": False, "choices": []})

    try:
        # Get all matching section masters
        filters = {
            "section_master_name": section_master_name,
            "section_master_metadata_name": metadata_name,
        }

        sm = SectionMaster.objects.filter(**filters).first()

        if sm:
            choices = list(sm.section_master_choice.order_by("id").values_list("master_choice", flat=True))
            return JsonResponse({"success": True, "choices": choices})
        else:
            return JsonResponse({"success": True, "choices": []})

    except Exception as e:
        return JsonResponse({"success": False, "choices": [], "error": str(e)})


@csrf_exempt
def get_nested_combobox_choices(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid method"})

    try:
        data = json.loads(request.body)
        metadata_name = data.get("section_sub_metadata_name")
        column_data = data.get("sub_column_data", {})
        task_id = data.get("project_row_id")
        task = get_object_or_404(Task, id=task_id)
        project = task.project
        proj_section_master_name = project.proj_section_master_name

        if not metadata_name or not column_data:
            return JsonResponse({"success": False, "choices": [], "error": "Missing metadata or column data"})

        activity_value = column_data.get("ACTIVITY", "").strip()

        sm = SectionMaster.objects.filter(
            section_master_name=proj_section_master_name,
            section_master_metadata_name=metadata_name,
            section_sub_metadata_name="ACTIVITY",
            section_sub_choice=activity_value,
            is_sub_choice=True
        ).first()

        if sm:
            if sm.is_completed:
                choices = list(sm.section_master_choice.order_by("id").values_list("master_choice", flat=True))
            else:
                choices = []
            return JsonResponse({"success": True, "choices": choices})

        # fallback
        matching_section_master = SectionMaster.objects.get(
            section_master_name=proj_section_master_name,
            section_master_metadata_name=metadata_name,
            is_sub_choice=False,
        )
        if matching_section_master.is_completed:
            choices = list(matching_section_master.section_master_choice.order_by("-id").values_list("master_choice", flat=True))
        else:
            choices = []
        return JsonResponse({"success": True, "choices": choices})

    except Exception as e:
        print("error:->", e)
        project_field = ProjectField.objects.filter(field_title=metadata_name).first()
        column_details_qs = SectionTemplateMetadata.objects.filter(
            section_task_metadata=project_field,
            section_task_id=task_id
        ).order_by("section_temp_meta_position")

        for item in column_details_qs:
            return JsonResponse({"success": True, "choices": item.section_temp_meta_choices})


@csrf_exempt
def get_special_char_choice(request):

    if request.method == "POST":

        try:
            data = json.loads(request.body)
            master_choice = data.get("master_choice")

            smc = SectionMasterChoice.objects.filter(master_choice=master_choice).first()
            if smc and smc.special_char_choice:
                return JsonResponse({"success": True, "special_char_choice": smc.special_char_choice})
            else:
                return JsonResponse({"success": True, "special_char_choice": ""})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Invalid method"}, status=405)


# @csrf_exempt
# def get_nested_combobox_choices(request):
#
#     if request.method != "POST":
#         return JsonResponse({"success": False, "error": "Invalid method"})
#
#     try:
#         data = json.loads(request.body)
#         metadata_name = data.get("section_sub_metadata_name")
#         column_data = data.get("sub_column_data", {})
#         task_id = data.get("project_row_id")
#         task = get_object_or_404(Task, id=task_id)
#         project = task.project
#         proj_section_master_name = project.proj_section_master_name
#         if not metadata_name or not column_data:
#             return JsonResponse({"success": False, "choices": [], "error": "Missing metadata or column data"})
#
#         # Search through each key-value to find matching SectionMaster
#         for key, value in column_data.items():
#             sm = SectionMaster.objects.filter(
#                 section_master_metadata_name=metadata_name,
#                 is_sub_choice=True,
#                 section_sub_metadata_name=key,
#                 section_sub_choice=value
#             ).first()
#
#             if sm:
#                 print("Matched SectionMaster:", sm)
#                 choices = list(sm.section_master_choice.order_by("id").values_list("master_choice", flat=True))
#                 print("Matched choices:->", choices)
#                 return JsonResponse({"success": True, "choices": choices})
#         # No match found
#
#         matching_section_master = SectionMaster.objects.get(
#             section_master_name=proj_section_master_name,
#             section_master_metadata_name=metadata_name,
#             is_sub_choice=False,
#         )
#         choices = list(matching_section_master.section_master_choice.order_by("-id").values_list("master_choice", flat=True))
#         if choices:
#             print("choices:->", choices)
#             return JsonResponse({"success": True, "choices": choices})
#     except Exception as e:
#         print("error:->",e)
#         project_field = ProjectField.objects.filter(field_title=metadata_name).first()
#         # Fetch section metadata details from SectionTemplateMetadata
#         column_details_qs = SectionTemplateMetadata.objects.filter(section_task_metadata=project_field,
#                                                                    section_task_id=task_id). \
#             order_by("section_temp_meta_position")
#         print("column_details_qs:->", column_details_qs)
#         for item in column_details_qs:
#             choices = item.section_temp_meta_choices
#             return JsonResponse({"success": True, "choices": choices})


class GetProjectLogHistory(LoginRequiredMixin, View):

    def get(self, request):

        if request.method == 'GET' and request.is_ajax():
            # project log history
            project_history_details = {}
            project_logs = {}
            req_data = request.GET.get('proj_id')
            proj_id = req_data
            project_log_history_obj = ProjectLogHistory.objects.filter(project_id = proj_id)
            project_log_id = [item.project_log_id for item in project_log_history_obj]
            project_id = [item.project_id for item in project_log_history_obj]
            project_log_message = [item.project_log_message for item in project_log_history_obj]
            project_log_created_by_list = [item.project_log_created_by.username for item in project_log_history_obj ]
            project_log_created_at_list = [item.project_log_created_at for item in project_log_history_obj]
            project_log_status_list = [item.get_project_log_status_display() for item in project_log_history_obj]

            for i in range(0,len(project_log_created_by_list)):
                project_log = {
                    'Log_ID':project_log_id[i],
                    'Project_ID':project_id[i],
                    'Log_Message':project_log_message[i],
                    'Created_By':project_log_created_by_list[i],
                    'Created_At':project_log_created_at_list[i],
                    'Project_Log_Status':project_log_status_list[i],
                }
                project_logs["project_log"+str(i+1)] = project_log
            project_history_details["project_logs"] = project_logs

            response = {
                'success': True,
                'project_history_details': project_history_details,
            }
            return JsonResponse(response, safe=False)


# Exports all tasks into one Excel sheet.
def apply_border_to_range(ws, start_row, start_col, end_row, end_col, border):
    """Apply borders to all edges of a merged cell range."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            left = border.left if col == start_col else Side(border_style=None)
            right = border.right if col == end_col else Side(border_style=None)
            top = border.top if row == start_row else Side(border_style=None)
            bottom = border.bottom if row == end_row else Side(border_style=None)
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def export_excel_to_single_sheet(request):

    proj_id = request.GET.get('proj_id')
    if not proj_id:
        return JsonResponse({'success': False, 'error': 'Project ID not provided.'})

    try:
        project = Project.objects.get(proj_id=proj_id)
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found.'})

    name = re.sub(r'[:\\/?*\[\]]', '', str(project.proj_name))[:20] or "Sheet"

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title=name)
    row_num = 1

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    col_width = 17
    max_columns_used = 1

    all_tasks = Task.objects.filter(project_id=proj_id, task_deleted=False).order_by('sort_order')
    task_dict = defaultdict(list)
    root_tasks = []

    for task in all_tasks:
        meta_count = SectionTemplateMetadata.objects.filter(section_task_id=str(task.id)).count()
        max_columns_used = max(max_columns_used, meta_count)
        if task.parent and task.parent != '0':
            task_dict[task.parent].append(task)
        else:
            root_tasks.append(task)

    total_width = 150
    num_excel_columns = total_width // col_width
    for i in range(1, num_excel_columns + 1):
        ws.column_dimensions[get_column_letter(i)].width = col_width

    # Row 1: Project Name
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_excel_columns)
    cell = ws.cell(row=row_num, column=1)
    cell.value = f"Project Name: {project.proj_name}"
    cell.font = Font(size=14, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    apply_border_to_range(ws, row_num, 1, row_num, num_excel_columns, thin_border)
    row_num += 1

    # Row 2: Project Description
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_excel_columns)
    cell = ws.cell(row=row_num, column=1)
    cell.value = f"Project Description: {project.proj_description}"
    cell.font = Font(size=12, italic=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    apply_border_to_range(ws, row_num, 1, row_num, num_excel_columns, thin_border)
    row_num += 1

    # Row 3: Chassis Number
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_excel_columns)
    cell = ws.cell(row=row_num, column=1)
    cell.value = f"Chassis Number: {project.proj_dropdown_field or ''}"
    cell.font = Font(size=12, italic=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    apply_border_to_range(ws, row_num, 1, row_num, num_excel_columns, thin_border)
    row_num += 1

    # Row 4: Date and Time in merged cell
    try:
        user_profile = Profile.objects.get(user=request.user)
        user_tz_name = user_profile.client_tz or 'Asia/Kolkata'
    except Profile.DoesNotExist:
        user_tz_name = 'Asia/Kolkata'

    try:
        user_tz = timezone(user_tz_name)
    except UnknownTimeZoneError:
        user_tz = timezone('Asia/Kolkata')

    localized_now = django_now().astimezone(user_tz)
    now_str = localized_now.strftime("Project export Date | Time: %d %B, %Y | %I:%M:%S %p")

    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_excel_columns)
    cell = ws.cell(row=row_num, column=1)
    cell.value = now_str
    cell.font = Font(size=11, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    apply_border_to_range(ws, row_num, 1, row_num, num_excel_columns, thin_border)
    row_num += 1

    # Row 5: Model name and status
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_excel_columns)
    model_status = f"Project model name: {project.proj_section_master_name or ''} | Project status: {project.proj_status.proj_status if project.proj_status else ''}"
    cell = ws.cell(row=row_num, column=1)
    cell.value = model_status
    cell.font = Font(size=11, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    apply_border_to_range(ws, row_num, 1, row_num, num_excel_columns, thin_border)
    row_num += 1

    # Row 6: Start date and end date
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_excel_columns)
    start_date = project.proj_start_date.strftime("%d %b %Y") if project.proj_start_date else ""
    end_date = project.proj_end_date.strftime("%d %b %Y") if project.proj_end_date else ""
    start_end = f"Project start date: {start_date} | Project end date: {end_date}"
    cell = ws.cell(row=row_num, column=1)
    cell.value = start_end
    cell.font = Font(size=11, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    apply_border_to_range(ws, row_num, 1, row_num, num_excel_columns, thin_border)
    row_num += 1

    # Row 7: Signature
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_excel_columns)
    signature = f"Project signature: {project.proj_signature or ''}"
    cell = ws.cell(row=row_num, column=1)
    cell.value = signature
    cell.font = Font(size=11, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    apply_border_to_range(ws, row_num, 1, row_num, num_excel_columns, thin_border)
    row_num += 2

    # Row 8: Company Name Header
    if getattr(connection.tenant, 'name', '').lower() in ["acgl", "reliance", "macys"]:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_excel_columns)
        cell = ws.cell(row=row_num, column=1)
        cell.value = "AUTOMOBILE CORPORATION OF GOA LTD."
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        apply_border_to_range(ws, row_num, 1, row_num, num_excel_columns, thin_border)
        row_num += 1

    def get_merge_plan(num_columns):
        if num_columns == 1:
            return [(1, 8)]
        elif num_columns == 2:
            return [(1, 4), (5, 4)]
        elif num_columns == 3:
            return [(1, 2), (3, 4), (7, 2)]
        elif num_columns == 4:
            return [(1, 2), (3, 3), (6, 1), (7, 2)]
        elif num_columns == 5:
            return [(1, 1), (2, 3), (5, 1), (6, 1), (7, 2)]
        elif num_columns == 6:
            return [(1, 1), (2, 2), (4, 1), (5, 1), (6, 1), (7, 2)]
        elif num_columns == 7:
            return [(1, 1), (2, 2), (4, 1), (5, 1), (6, 1), (7, 1), (8, 1)]
        elif num_columns == 8:
            return [(1, 1), (2, 1), (3, 1), (4, 1), (5, 1), (6, 1), (7, 1), (8, 1)]
        else:
            return [(i + 1, 1) for i in range(num_columns)]

    def write_task_single_sheet(task, ws, row_num, indent_level=0):
        metadata_fields = SectionTemplateMetadata.objects.filter(section_task_id=str(task.id)).order_by('section_temp_meta_position')
        column_fields = [meta.section_task_metadata for meta in metadata_fields]
        column_headers = [field.field_title for field in column_fields]
        num_columns = len(column_headers)

        merge_plan = get_merge_plan(num_columns)

        # Task title row
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_excel_columns)
        cell = ws.cell(row=row_num, column=1)
        cell.value = ("   " * indent_level) + task.text
        cell.font = Font(size=14, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        apply_border_to_range(ws, row_num, 1, row_num, num_excel_columns, thin_border)
        row_num += 1

        # Column headers
        max_estimated_height = 15  # Default height for single-line header
        special_col_indices = []

        for col_index, (start_col, span) in enumerate(merge_plan):
            ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=start_col + span - 1)
            cell = ws.cell(row=row_num, column=start_col)
            header_text = column_headers[col_index]
            cell.value = header_text
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
            apply_border_to_range(ws, row_num, start_col, row_num, start_col + span - 1, thin_border)

            # Check for SPL CHAR-type column
            if header_text.strip().lower().replace('.', '') == 'spl char':
                special_col_indices.append(col_index)

            # Estimate wrapped lines: assuming ~1 char fits per unit width
            est_chars_per_line = col_width * span  # Approximate characters that can fit
            if est_chars_per_line > 0:
                lines = -(-len(header_text) // est_chars_per_line)  # Ceiling division
                lines = math.ceil(lines)
                row_height = lines * 15  # Assume 15 units per line
                max_estimated_height = max(max_estimated_height, row_height)

        # Set final row height based on max-estimated line wrap
        ws.row_dimensions[row_num].height = max_estimated_height
        row_num += 1

        # Data rows
        values_qs = SectionMetadataValues.objects.filter(section_task_id=str(task.id))
        values_by_position = defaultdict(dict)
        for val in values_qs:
            values_by_position[val.section_meta_val_position][val.section_task_metadata_id] = val.section_metadata_value

        for position in sorted(values_by_position.keys()):
            row_values = values_by_position[position]
            max_estimated_height = 15  # base row height

            for col_index, (start_col, span) in enumerate(merge_plan):
                value = row_values.get(column_fields[col_index].pk, "")
                text = str(value) if value is not None else ""

                # Apply SPL CHAR transformation if needed
                if col_index in special_col_indices:
                    low_text = text.lower()
                    if low_text == 'minor':
                        text = "◯"
                    elif low_text == 'major':
                        text = "◑"
                    elif low_text == 'critical':
                        text = "⬤"

                # Merge and write value
                ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num,
                               end_column=start_col + span - 1)
                cell = ws.cell(row=row_num, column=start_col)
                cell.value = text
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                apply_border_to_range(ws, row_num, start_col, row_num, start_col + span - 1, thin_border)

                # Estimate line wrapping for height adjustment
                est_chars_per_line = col_width * span
                if est_chars_per_line > 0:
                    est_lines = -(-len(text) // est_chars_per_line)  # Ceiling division
                    row_height = est_lines * 15
                    max_estimated_height = max(max_estimated_height, row_height)

            # Set row height after all merged columns are processed
            ws.row_dimensions[row_num].height = max_estimated_height
            row_num += 1

        row_num += 2
        for child in task_dict.get(str(task.id), []):
            row_num = write_task_single_sheet(child, ws, row_num, indent_level + 1)
        return row_num

    for root_task in root_tasks:
        row_num = write_task_single_sheet(root_task, ws, row_num)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Project_{proj_id}_GroupedTasks.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# Exports child tasks underneath their parent in the same Excel sheet
def grouped_hierarchy_clean_sheet_name(name, used_names, fallback="Sheet"):
    name = re.sub(r'[:\\/?*\[\]]', '', name)
    name = name.strip()
    base_name = name[:31] if name else fallback
    unique_name = base_name
    suffix = 1
    while unique_name in used_names:
        suffix_str = f"_{suffix}"
        unique_name = (base_name[:31 - len(suffix_str)] + suffix_str)[:31]
        suffix += 1
    used_names.add(unique_name)
    return unique_name


def export_excel_to_grouped_hierarchy(request):
    proj_id = request.GET.get('proj_id')
    if not proj_id:
        return JsonResponse({'success': False, 'error': 'Project ID not provided.'})

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    used_sheet_names = set()

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Fetch all tasks
    all_tasks = Task.objects.filter(project_id=proj_id, task_deleted=False).order_by('sort_order')

    # Group tasks into parent/child structure
    children_map = defaultdict(list)
    top_level_tasks = []

    for task in all_tasks:
        if task.parent and task.parent != '0':
            children_map[task.parent].append(task)
        else:
            top_level_tasks.append(task)

    def write_task_grouped_hierarchy(ws, task, row_num, indent_level=0):
        # Fetch metadata fields sorted by position
        metadata_fields = SectionTemplateMetadata.objects.filter(
            section_task_id=str(task.id)
        ).order_by('section_temp_meta_position')

        column_fields = [meta.section_task_metadata for meta in metadata_fields]
        column_headers = [field.field_title for field in column_fields]

        num_columns = len(column_headers)
        if num_columns == 0:
            num_columns = 1  # Prevent merge error if no metadata

        # TASK NAME ROW
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_columns)
        indent_prefix = "    " * indent_level
        task_cell = ws.cell(row=row_num, column=1, value=indent_prefix + task.text)
        task_cell.font = Font(size=14, bold=True, color="FFFFFF")
        task_cell.alignment = Alignment(horizontal='center')
        task_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        row_num += 1

        # COLUMN HEADER ROW
        for col_index, header in enumerate(column_headers, start=1):
            cell = ws.cell(row=row_num, column=col_index, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
            cell.border = thin_border

        row_num += 1

        # Group metadata values by position
        values_qs = SectionMetadataValues.objects.filter(section_task_id=str(task.id))
        values_by_position = defaultdict(dict)
        for val in values_qs:
            values_by_position[val.section_meta_val_position][val.section_task_metadata_id] = val.section_metadata_value

        # Write each data row
        for position in sorted(values_by_position.keys()):
            row_values = values_by_position[position]
            for col_index, field in enumerate(column_fields, start=1):
                value = row_values.get(field.pk, "")
                cell = ws.cell(row=row_num, column=col_index, value=value)
                cell.alignment = Alignment(horizontal='left', wrap_text = True)
                cell.border = thin_border
            row_num += 1

        row_num += 2  # space before next task

        # Recursively write child tasks
        for child in children_map.get(str(task.id), []):
            row_num = write_task_grouped_hierarchy(ws, child, row_num, indent_level + 1)

        return row_num

    # Write each top-level task and its children in its own sheet
    for task in top_level_tasks:
        sheet_name = grouped_hierarchy_clean_sheet_name(task.text or f"Task-{task.id}", used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)
        row_num = 1
        row_num = write_task_grouped_hierarchy(ws, task, row_num)

        # Auto-size columns
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    continue
            column_letter = get_column_letter(col_idx)
            # ws.column_dimensions[column_letter].width = max_length + 5
            ws.column_dimensions[column_letter].width = min(max(8, max_length + 2), 40)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Project_{proj_id}_Tasks_With_Subtasks.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# Exports each task (parent or child) into its own Excel sheet.
def separate_sheets_clean_sheet_name(name, used_names, fallback="Sheet"):
    name = re.sub(r'[:\\/?*\[\]]', '', name).strip()
    base_name = name[:31] if name else fallback
    unique_name = base_name
    suffix = 1
    while unique_name in used_names:
        suffix_str = f"_{suffix}"
        unique_name = (base_name[:31 - len(suffix_str)] + suffix_str)[:31]
        suffix += 1
    used_names.add(unique_name)
    return unique_name


# Write one task to its own sheet
def write_task_separate_sheets(wb, task, used_sheet_names, border):
    sheet_name = separate_sheets_clean_sheet_name(task.text or f"Task-{task.id}", used_sheet_names)
    ws = wb.create_sheet(title=sheet_name)
    row_num = 1

    metadata_fields = SectionTemplateMetadata.objects.filter(
        section_task_id=str(task.id)
    ).order_by('section_temp_meta_position')

    column_fields = [meta.section_task_metadata for meta in metadata_fields]
    column_headers = [field.field_title for field in column_fields]

    num_columns = len(column_headers)
    if num_columns == 0:
        num_columns = 1  # Avoid merge error

    # Task name row
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_columns)
    task_cell = ws.cell(row=row_num, column=1, value=task.text)
    task_cell.font = Font(size=14, bold=True, color="FFFFFF")
    task_cell.alignment = Alignment(horizontal='center')
    task_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    row_num += 1

    # Column headers
    for col_index, header in enumerate(column_headers, start=1):
        cell = ws.cell(row=row_num, column=col_index, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
        cell.border = border
    row_num += 1

    # Metadata values
    values_qs = SectionMetadataValues.objects.filter(section_task_id=str(task.id))
    values_by_position = defaultdict(dict)
    for val in values_qs:
        values_by_position[val.section_meta_val_position][val.section_task_metadata_id] = val.section_metadata_value

    for position in sorted(values_by_position.keys()):
        row_values = values_by_position[position]
        for col_index, field in enumerate(column_fields, start=1):
            value = row_values.get(field.pk, "")
            cell = ws.cell(row=row_num, column=col_index, value=value)
            cell.alignment = Alignment(horizontal='left', wrap_text = True)
            cell.border = border
        row_num += 1

    # Auto-size columns
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                continue
        column_letter = get_column_letter(col_idx)
        # ws.column_dimensions[column_letter].width = max_length + 5
        ws.column_dimensions[column_letter].width = min(max(8, max_length + 2), 40)


def export_excel_to_separate_sheets(request):
    proj_id = request.GET.get('proj_id')
    if not proj_id:
        return JsonResponse({'success': False, 'error': 'Project ID not provided.'})

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    used_sheet_names = set()

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    all_tasks = Task.objects.filter(project_id=proj_id, task_deleted=False).order_by('sort_order')

    # Build parent-children map
    children_map = defaultdict(list)
    task_dict = {}
    for task in all_tasks:
        task_dict[str(task.id)] = task
        if task.parent and task.parent != '0':
            children_map[task.parent].append(task)

    def write_in_order(task):
        write_task_separate_sheets(wb, task, used_sheet_names, thin_border)
        for child in children_map.get(str(task.id), []):
            write_in_order(child)

    # Start writing from top-level tasks
    for task in all_tasks:
        if not task.parent or task.parent == '0':
            write_in_order(task)

    # Return XLSX
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Project_{proj_id}_All_Tasks_Ordered.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# pdf file conversion, pagesize A4 and landscape
def export_to_pdf_file(request):

    proj_id = request.GET.get("proj_id")
    mode = request.GET.get("mode")
    if not proj_id:
        return JsonResponse({'success': False, 'error': 'Project ID not provided.'})

    try:
        project = Project.objects.get(proj_id=proj_id)
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found.'})

    pagesize = A4 if mode == 'portrait' else landscape(A4)
    width, height = pagesize

    try:
        user_profile = Profile.objects.get(user=request.user)
        tz_name = user_profile.client_tz or "Asia/Kolkata"
    except Profile.DoesNotExist:
        tz_name = "Asia/Kolkata"

    try:
        user_tz = timezone(tz_name)
    except UnknownTimeZoneError:
        user_tz = timezone("Asia/Kolkata")

    local_time = django_now().astimezone(user_tz)
    formatted_time = local_time.strftime("Project export Date | Time: %d %B, %Y | %I:%M:%S %p")

    # Register DejaVuSans font
    font_path = MEDIA_ROOT + '/DejaVuSans.ttf'  # adjust path if needed
    pdfmetrics.registerFont(TTFont("DejaVu", font_path))
    italic_font_path = MEDIA_ROOT + '/DejaVuSans-Oblique.ttf'  # adjust path if needed
    pdfmetrics.registerFont(TTFont("DejaVu-Italic", italic_font_path))

    # Styles
    styles = getSampleStyleSheet()
    wrap_style = ParagraphStyle("Wrap", parent=styles["Normal"], fontName="DejaVu", fontSize=8, leading=10)
    heading_style = ParagraphStyle("Heading", fontName="DejaVu", fontSize=14, textColor=colors.white,
                                   alignment=1, backColor=colors.HexColor("#1F4E78"),
                                   leading=17, spaceAfter=1)
    desc_style = ParagraphStyle("Desc", fontName="DejaVu-Italic", fontSize=11, textColor=colors.white,
                                alignment=1, backColor=colors.HexColor("#2E75B6"),
                                leading=14, spaceAfter=1)
    time_style = ParagraphStyle("Time", fontName="DejaVu", fontSize=10, textColor=colors.white,
                                alignment=1, backColor=colors.HexColor("#4F81BD"),
                                leading=12, spaceAfter=1)
    signature_style = ParagraphStyle("Signature", fontName="DejaVu", fontSize=10, textColor=colors.white,
                                alignment=1, backColor=colors.HexColor("#4F81BD"),
                                leading=12, spaceAfter=5)
    company_style = ParagraphStyle("Company", fontName="DejaVu", fontSize=13, textColor=colors.white,
                                   alignment=1, backColor=colors.HexColor("#203864"),
                                   leading=16)

    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#9BBB59")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'DejaVu'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
    ])

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=23, rightMargin=24, topMargin=24, bottomMargin=24)
    doc.title = project.proj_name  # Add this line to set the PDF metadata title
    elements = []

    # Row 1: Project Name
    elements.append(Paragraph(f"Project Name: {project.proj_name}", heading_style))

    # Row 2: Project Description
    elements.append(Paragraph(f"Project Description: {project.proj_description}", desc_style))

    # Row 3: Chassis Number
    elements.append(Paragraph(f"Chassis Number: {project.proj_dropdown_field or ''}", desc_style))

    # Row 4: Timestamp
    elements.append(Paragraph(formatted_time, time_style))

    # Row 5: Model name and project status
    model_status_text = f"Project model name: {project.proj_section_master_name or ''} | Project status: {project.proj_status.proj_status if project.proj_status else ''}"
    elements.append(Paragraph(model_status_text, time_style))

    # Row 6: Start and End Dates
    start_date = project.proj_start_date.strftime("%d %b %Y") if project.proj_start_date else ""
    end_date = project.proj_end_date.strftime("%d %b %Y") if project.proj_end_date else ""
    date_range_text = f"Project start date: {start_date} | Project end date: {end_date}"
    elements.append(Paragraph(date_range_text, time_style))

    # Row 7: Signature
    signature_text = f"Project signature: {project.proj_signature or ''}"
    elements.append(Paragraph(signature_text, signature_style))

    company = getattr(connection.tenant, 'name', '').lower()
    if company in ["acgl", "reliance", "macys"]:
        elements.append(Paragraph("AUTOMOBILE CORPORATION OF GOA LTD.", company_style))

    all_tasks = Task.objects.filter(project_id=proj_id, task_deleted=False).order_by('sort_order')
    task_dict = defaultdict(list)
    root_tasks = []

    for task in all_tasks:
        if task.parent and task.parent != '0':
            task_dict[task.parent].append(task)
        else:
            root_tasks.append(task)

    def write_task_to_pdf(task, indent_level=0):
        elements.append(Spacer(1, 0.2 * inch))
        title = Paragraph((" " * 4 * indent_level) + task.text, heading_style)
        elements.append(title)

        metadata_fields = SectionTemplateMetadata.objects.filter(section_task_id=str(task.id)).order_by("section_temp_meta_position")
        column_fields = [meta.section_task_metadata for meta in metadata_fields]
        column_headers = [field.field_title for field in column_fields]
        values_qs = SectionMetadataValues.objects.filter(section_task_id=str(task.id))
        values_by_position = defaultdict(dict)
        for val in values_qs:
            values_by_position[val.section_meta_val_position][val.section_task_metadata_id] = val.section_metadata_value

        data = []
        if column_headers:
            # Wrap column headers
            header_row = [Paragraph(h, wrap_style) for h in column_headers]
            data.append(header_row)

            for position in sorted(values_by_position.keys()):
                row = []
                for field in column_fields:
                    raw_val = values_by_position[position].get(field.pk, "")
                    col_name = field.field_title.strip().lower().replace('.', '')

                    if "spl char" in col_name:
                        val_lower = str(raw_val).strip().lower()
                        if val_lower == "minor":
                            display_val = "◯"
                        elif val_lower == "major":
                            display_val = "◑"
                        elif val_lower == "critical":
                            display_val = "⬤"
                        else:
                            display_val = str(raw_val) if raw_val is not None else ""
                    else:
                        display_val = str(raw_val) if raw_val is not None else ""

                    row.append(Paragraph(display_val, wrap_style))
                data.append(row)
        else:
            data.append([Paragraph("No Metadata Defined", wrap_style)])

        if data:
            col_count = len(data[0])
            col_width = (width - 60) / max(1, col_count)
            table = Table(data, colWidths=[col_width] * col_count)
            table.setStyle(table_style)
            elements.append(table)

        for child in task_dict.get(str(task.id), []):
            write_task_to_pdf(child, indent_level + 1)

    for root_task in root_tasks:
        write_task_to_pdf(root_task)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"Project_{proj_id}_GroupedTasks.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@csrf_exempt
def project_lock_unlock(request):

    if request.method == "POST":

        proj_id = request.POST.get("proj_id")
        try:
            project = Project.objects.get(proj_id=proj_id)
            project.proj_lock = not project.proj_lock
            project.save()
            return JsonResponse({"success": True, "locked": project.proj_lock})
        except Project.DoesNotExist:
            return JsonResponse({"success": False, "error": "Project not found"})

    return JsonResponse({"success": False, "error": "Invalid request"})